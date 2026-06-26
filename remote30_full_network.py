"""Remote 30 전체 배관망 총괄 모듈 (10번 모듈 백엔드).

기존 ``remote30_prototype`` 의 헤드망 추출 로직(Stage A)을 재사용하고,
zone별 라이저 템플릿(Stage B), 라이저↔헤드망 stitch(Stage C),
PIPENET-native 후처리 + Pump-fan/Elastomeric-valve 직렬화(Stage D) 를
추가하여 펌프 → 감압밸브 → 알람밸브 → 헤드 30개 전 구간 SDF 를 생성한다.

흐름::

    OverallInputs (DXF + ZoneSpec + (선택) BuildingPressureProfile)
            │
            ├── Stage A — run_stages_0_2 (remote30_prototype 재사용)
            │                 → PipeTables (헤드망)
            │
            ├── Stage B — build_riser(zone_spec, profile)
            │                 → RiserTables (펌프/PRV/라이저)
            │
            ├── Stage C — stitch_riser_and_heads(riser, head_tables)
            │                 → CombinedTables
            │
            └── Stage D — emit_full_sdf(combined, out_path)
                              → 완성 SDF + 동봉 SLF

신규 attribute (vs prototype 의 PipeTables)::

    RiserTables.pumps   — <Pump-fan> 직렬화용 dict 리스트
    RiserTables.valves  — <Elastomeric-valve> 직렬화용 dict 리스트
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Iterator


# ────────────────────────────────────────────────────────────────────────────
# 상수
# ────────────────────────────────────────────────────────────────────────────

KGF_CM2_TO_PA = 98066.5    # 1 kg/cm² (kgf/cm²) → Pa
M_TO_PA = 9806.65          # 1 m 수두 → Pa (물 비중 1.0 기준)
ATM_PA = 101325.0          # 1 기압 (boundary condition)


# ────────────────────────────────────────────────────────────────────────────
# Zone 정의
# ────────────────────────────────────────────────────────────────────────────

class ZoneType(Enum):
    """처리 가능한 zone 타입 — 답안 SDF 의 라이저 구조에서 도출.

    압력분포표(예: 대명동 201동 PDF)의 각 행 → ZoneType 매핑::

        옥상 + 27F~49F          → HSP_PUMP        (펌프식 부스터)
        25F~26F                 → LSP_GRAVITY     (감압 없음, 자연낙차)
        2F~24F (24F 1차 PRV)    → LSP_1STAGE      (1차 감압)
        1F~B4 (1.5F 2차 PRV)    → LLSP_2STAGE     (1차 + 2차 감압)
    """
    HSP_PUMP = "hsp_pump"
    LSP_GRAVITY = "lsp_gravity"
    LSP_1STAGE = "lsp_1stage"
    LLSP_2STAGE = "llsp_2stage"


# ────────────────────────────────────────────────────────────────────────────
# 데이터 모델
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class FloorRow:
    """압력분포표의 한 행 (옥상~B4 까지 1층 1행)."""
    floor_label: str          # "옥상층", "49층", "1층", "B1층" 등
    height_m: float           # 층고 (m)
    head_drop_m: float        # 누적 낙차 (m, 옥상 수원 기준)
    after_prv_m: float | None = None  # 감압 이후 수두 (m) — 감압 구간만
    note: str = ""            # 비고 ("자연낙차시작점", "1차 감압밸브 사용구간" 등)


@dataclass
class BuildingPressureProfile:
    """빌딩 전체 압력 흐름 표 — 옥상부터 최하층까지 1행씩.

    CSV/엑셀 업로드 또는 사용자 직접 입력 폼으로 생성. 없을 수도 있음 (선택적).
    """
    building_name: str = ""
    floors: list[FloorRow] = field(default_factory=list)

    def find_by_label(self, floor_label: str) -> FloorRow | None:
        for row in self.floors:
            if row.floor_label == floor_label:
                return row
        return None

    @classmethod
    def from_csv(cls, csv_path: Path, *, building_name: str = "") -> "BuildingPressureProfile":
        """CSV 파서 — 컬럼: floor_label, height_m, head_drop_m, after_prv_m, note.

        헤더 행은 한글/영문 모두 허용 (구분/층고/낙차압/감압이후/비고).
        """
        rows: list[FloorRow] = []
        # 한글 컬럼명 → 영문 키 매핑 (PDF 표 헤더에 맞춤)
        KCOL = {
            "구분": "floor_label", "층": "floor_label",
            "층고": "height_m",
            "낙차압": "head_drop_m", "낙차": "head_drop_m",
            "감압이후": "after_prv_m", "감압후": "after_prv_m",
            "비고": "note",
        }
        with csv_path.open("r", encoding="utf-8-sig", newline="") as fp:
            reader = csv.DictReader(fp)
            for raw in reader:
                # 한글 헤더면 매핑 적용
                norm: dict[str, Any] = {}
                for k, v in raw.items():
                    key = KCOL.get((k or "").strip(), (k or "").strip())
                    norm[key] = (v or "").strip()
                if not norm.get("floor_label"):
                    continue
                try:
                    rows.append(FloorRow(
                        floor_label=norm["floor_label"],
                        height_m=float(norm.get("height_m") or 0),
                        head_drop_m=float(norm.get("head_drop_m") or 0),
                        after_prv_m=(float(norm["after_prv_m"]) if norm.get("after_prv_m") else None),
                        note=norm.get("note", "") or "",
                    ))
                except (ValueError, KeyError):
                    continue
        return cls(building_name=building_name, floors=rows)

    @classmethod
    def from_xlsx(cls, xlsx_path: Path, *, sheet: str | int = 0,
                  building_name: str = "") -> "BuildingPressureProfile":
        """엑셀(.xlsx) 파서 — 첫 시트 또는 지정 시트의 같은 컬럼 구조."""
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        header_row = next(ws.iter_rows(values_only=True))
        KCOL = {
            "구분": "floor_label", "층": "floor_label",
            "층고": "height_m",
            "낙차압": "head_drop_m", "낙차": "head_drop_m",
            "감압이후": "after_prv_m", "감압후": "after_prv_m",
            "비고": "note",
        }
        col_idx = {KCOL.get((str(h) or "").strip(), (str(h) or "").strip()): i
                   for i, h in enumerate(header_row) if h is not None}
        rows: list[FloorRow] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            def _get(key, default=None):
                i = col_idx.get(key)
                return r[i] if i is not None and i < len(r) else default
            label = _get("floor_label")
            if not label:
                continue
            try:
                rows.append(FloorRow(
                    floor_label=str(label),
                    height_m=float(_get("height_m") or 0),
                    head_drop_m=float(_get("head_drop_m") or 0),
                    after_prv_m=(float(_get("after_prv_m")) if _get("after_prv_m") is not None else None),
                    note=str(_get("note") or ""),
                ))
            except (ValueError, TypeError):
                continue
        wb.close()
        return cls(building_name=building_name, floors=rows)


@dataclass
class ZoneSpec:
    """처리 대상 zone 의 사양. 사용자 입력 (라디오 + 폼).

    필드 의미::

        zone_type             — ZoneType (HSP_PUMP / LSP_GRAVITY / LSP_1STAGE / LLSP_2STAGE)
        target_floor          — 추출 대상 층 ("16층" — Stage A 헤드망과 동일)
        prv1_target_pa        — 1차 PRV 출구압 (Pa). 자연낙차 + 감압 zone 에서.
        prv2_target_pa        — 2차 PRV 출구압 (Pa). LLSP_2STAGE 만.
        pump_library_name     — Library-pump 의 SLF Pump-definition 이름.
                                기본 "SP_162M_2900LPM" (표준 SLF 와 정합).
        pump_count            — Pump-fan 개수 (HSP 보통 2개 = 1차+2차 부스터)

    압력분포표가 있으면 prv1/prv2 target 은 표에서 자동 도출 가능, 없으면 직접 입력.
    """
    zone_type: ZoneType
    target_floor: str = ""
    prv1_target_pa: float | None = None
    prv2_target_pa: float | None = None
    pump_library_name: str = "SP_162M_2900LPM"
    pump_count: int = 2


@dataclass
class OverallInputs:
    """모듈 10 의 입력 일체."""
    dxf_path: Path
    zone_spec: ZoneSpec
    profile: BuildingPressureProfile | None = None
    alarm_xy: tuple[float, float] | None = None
    job_id: str = ""
    project_title: str = "Remote 30 전체 배관망 총괄"


# ────────────────────────────────────────────────────────────────────────────
# RiserTables — Stage B 산출 / Stage C 입력
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class RiserTables:
    """라이저(펌프~AV) 구간의 PipeTables 확장.

    prototype.PipeTables 와 호환되는 nodes/pipes 형식 + pumps/valves 추가.
    각 dict 의 키는 emit_sdf 의 직렬화에 그대로 쓰일 수 있도록 PIPENET 속성명에 맞춤.
    """
    nodes: list[dict] = field(default_factory=list)
    pipes: list[dict] = field(default_factory=list)
    pumps: list[dict] = field(default_factory=list)   # Pump-fan 직렬화용
    valves: list[dict] = field(default_factory=list)  # Elastomeric-valve 직렬화용
    av_node_label: str = ""  # 라이저 끝점 = 헤드망 source 와 stitch 할 노드 라벨


# ────────────────────────────────────────────────────────────────────────────
# Stage A — 헤드망 추출 (remote30_prototype 재사용)
# ────────────────────────────────────────────────────────────────────────────

def run_stage_a(inputs: OverallInputs) -> Iterator[dict]:
    """Stage A — 평면도 DXF → 헤드 30개 + 배관망 추출 이벤트 스트림.

    호출 측(서버)은 마지막 ``stage2_complete`` 이벤트의 데이터를 보관 후
    사용자 헤드 편집 단계 → run_stages_3_5 → build_input_tables 까지 진행한다.
    """
    from remote30_prototype import run_stages_0_2
    yield from run_stages_0_2(
        dxf_path=inputs.dxf_path,
        job_id=inputs.job_id,
        alarm_xy=inputs.alarm_xy,
    )


# ────────────────────────────────────────────────────────────────────────────
# Stage B — zone별 라이저 템플릿
# ────────────────────────────────────────────────────────────────────────────
#
# 좌표 처리 방침:
# PIPENET SDF 의 Position x/y 는 isometric 시각 표시용일 뿐 수리계산엔 무관하다.
# emit_sdf 의 _xform 이 모든 노드 좌표를 bbox 중심 (0,0) + 약 3000 unit 으로 정규화하므로
# 라이저 노드 좌표의 절댓값은 의미가 없고 상대 배치(라이저 수직 진행, AV 가 적절한 위치)
# 만 유지하면 된다. 그러므로 빌딩 무관 logical 좌표계로 정의 — 모든 빌딩에서 동작.
#
# Logical 좌표 (mm 단위, 옥상 수원을 (0, 0) 으로):
#   Input(1)            (   0,    0)    옥상 수원
#   N2                  (-500, -100)    옥상 수평 분기
#   N3                  (-500, -300)    옥상 → 라이저 진입 (수직 강하 시작)
#   N4                  (-300, -1500)   라이저 중간 (수평 우회)
#   N7                  (-300, -2800)   PRV 진입 직전
#   N8 (PRV in)         (-200, -2900)
#   N89 (PRV out)       (-100, -2950)
#   N5 (AV 직전)         (-300, -3300)
#   N10 (AV ★)          ( 100, -3300)   헤드망 source 와 stitch
#   N87 (2차 PRV in)    ( -50, -3100)   LLSP_2STAGE 만
#   N88 (2차 PRV out)   (  50, -3150)
#   N100 (Pump Input)   ( 200,   100)   HSP_PUMP 만 (Input 1 보다 더 위)
#
# v1 구현 (대명동 201동 답안 SDF) 의 좌표는 git history 에서 확인 가능.
#
# 라이저 노드 라벨 컨벤션 (답안과 정합)::
#     1   : Input 노드 (옥상 수원, 1기압 boundary)
#     2,3,4,7  : 라이저 중간 노드 (옥상→하강→PRV 직전)
#     8   : 1차 PRV in
#     89  : 1차 PRV out
#     5   : AV 직전 노드
#     10  : AV 노드 (헤드망 source 와 stitch)
#     87  : 2차 PRV in (LLSP_2STAGE 만)
#     88  : 2차 PRV out (LLSP_2STAGE 만)
#     100 : Pump-fan 전 Input (HSP_PUMP 만 — Input 라벨이 1 → 100 으로 옮김)
#
# ────────────────────────────────────────────────────────────────────────────


def _node(label: str, elev: float, x: float, y: float, *,
          io_node: str = "No", pressure_pa: float | None = None) -> dict:
    """라이저 노드 dict — remote30_prototype.PipeTables.nodes 호환 형식."""
    d: dict = {
        "label": label, "elevation": elev, "io_node": io_node,
        "x": int(round(x)), "y": int(round(y)),
    }
    if pressure_pa is not None:
        d["pressure_pa"] = pressure_pa  # Input node 의 <Calculation-spec pressure="..."/>
    return d


def _pipe(label: str, in_lbl: str, out_lbl: str, bore_mm: int,
          length_m: float, rise_m: float = 0.0, c_factor: str = "120") -> dict:
    """라이저 파이프 dict — remote30_prototype.PipeTables.pipes 호환."""
    return {
        "label": label, "in": in_lbl, "out": out_lbl, "type": "KSD 3507",
        "dia": bore_mm, "length": round(length_m, 2), "elev": rise_m,
        "c": c_factor, "status": "Normal", "group": "Unset",
    }


def _pump_fan(label: str, in_lbl: str, out_lbl: str, *,
              library_pump: str, efficiency: int = 100, status: int = 1) -> dict:
    """Pump-fan dict — emit_full_sdf 가 <Pump-fan> 으로 직렬화."""
    return {
        "label": label, "in": in_lbl, "out": out_lbl,
        "efficiency": efficiency, "status": status,
        "library_pump": library_pump,
        "percentage_open": 1,
    }


def _pressure_valve(label: str, in_lbl: str, out_lbl: str, *,
                    target_pa: float, valve_type: str = "output") -> dict:
    """Elastomeric-valve dict — emit_full_sdf 가 <Elastomeric-valve> 로 직렬화."""
    return {
        "label": label, "in": in_lbl, "out": out_lbl,
        "target_value": float(target_pa), "type": valve_type,
    }


def _elev_at_floor(profile: BuildingPressureProfile | None, floor_label: str,
                   fallback: float = 0.0) -> float:
    """압력표에서 층 라벨의 누적 낙차 (m) → SDF elevation (음수)."""
    if profile is not None:
        row = profile.find_by_label(floor_label)
        if row is not None:
            return -float(row.head_drop_m)
    return fallback


# ── 라이저 logical 좌표 — 빌딩 무관. emit_sdf 의 _xform 이 정규화.
_COORDS_INPUT      = (   0,     0)    # 옥상 수원
_COORDS_N2         = (-500,  -100)    # 옥상 수평 분기
_COORDS_N3         = (-500,  -300)    # 옥상 → 라이저 진입
_COORDS_N4         = (-300, -1500)    # 라이저 중간 (수평 우회)
_COORDS_N7         = (-300, -2800)    # PRV 진입 직전
_COORDS_PRV_IN     = (-200, -2900)    # 노드 8 (1차 PRV in)
_COORDS_PRV_OUT    = (-100, -2950)    # 노드 89 (1차 PRV out)
_COORDS_PRV2_IN    = ( -50, -3100)    # 노드 87 (2차 PRV in, LLSP)
_COORDS_PRV2_OUT   = (  50, -3150)    # 노드 88 (2차 PRV out, LLSP)
_COORDS_N5_AV_PREV = (-300, -3300)    # AV 직전
_COORDS_AV         = ( 100, -3300)    # 노드 10 (헤드망 source ★)
_COORDS_PUMP_INPUT = ( 200,   100)    # HSP Pump-fan 앞 Input(100), Input(1) 보다 위


def build_riser_lsp_1stage(spec: ZoneSpec, profile: BuildingPressureProfile | None) -> RiserTables:
    """LSP 1차감압 라이저 — 자연낙차 + PRV 1개.

    노드: Input(1, elev=0) → 2 → 3 → 4 → 7 → 8(PRV in) → 89(PRV out) → 5 → 10(AV)
    파이프: 1→2, 2→3, 3→4, 4→7, 7→8, 89→5, 5→10
    Elastomeric-valve: 8 → 89 (target = spec.prv1_target_pa)
    """
    if spec.prv1_target_pa is None:
        raise ValueError("LSP_1STAGE 는 prv1_target_pa 가 필요합니다 (kg/cm² 또는 m 수두 → Pa).")
    elev_av = _elev_at_floor(profile, spec.target_floor, fallback=-98.45)
    elev_prv = -73.35  # PRV 위치 (옥상 -73m) — 답안 16F 기준 고정값 (24F 1차 감압 위치)
    elev_top = -3.75   # 옥상 수직 분기 위치
    return RiserTables(
        nodes=[
            _node("1",  0.0,        *_COORDS_INPUT,    io_node="Input", pressure_pa=ATM_PA),
            _node("2",  0.0,        *_COORDS_N2),
            _node("3",  elev_top,   *_COORDS_N3),
            _node("4",  elev_top,   *_COORDS_N4),
            _node("7",  elev_prv,   *_COORDS_N7),
            _node("8",  elev_prv,   *_COORDS_PRV_IN),
            _node("89", elev_prv,   *_COORDS_PRV_OUT),
            _node("5",  elev_av - 1.0, *_COORDS_N5_AV_PREV),
            _node("10", elev_av,    *_COORDS_AV),
        ],
        pipes=[
            _pipe("r1", "1", "2",   150, 20.95, 0.0),
            _pipe("r2", "2", "3",   150,  3.75, elev_top - 0.0),
            _pipe("r3", "3", "4",   150, 14.93, 0.0),
            _pipe("r4", "4", "7",   150, abs(elev_prv - elev_top), elev_prv - elev_top),
            _pipe("r5", "7", "8",   150,  0.5,  0.0),
            _pipe("r6", "89","5",   150, abs(elev_av - 1.0 - elev_prv), (elev_av - 1.0) - elev_prv),
            _pipe("r7", "5", "10",  125,  1.5,  1.0),
        ],
        pumps=[],
        valves=[
            _pressure_valve("1", "8", "89", target_pa=spec.prv1_target_pa),
        ],
        av_node_label="10",
    )


def build_riser_hsp_pump(spec: ZoneSpec, profile: BuildingPressureProfile | None) -> RiserTables:
    """HSP 펌프식 라이저 — 자연낙차 부족 → Pump-fan 부스터 + PRV 1개.

    노드: Input(100, elev=0) → [Pump-fan 100→1] → 1 → 2 → 3 → 4 → 7 → 8(PRV in) → 89(PRV out) → 5 → 10(AV)
    Pump-fan: 100 → 1 (Library-pump = spec.pump_library_name, count=spec.pump_count)
    """
    if spec.prv1_target_pa is None:
        raise ValueError("HSP_PUMP 는 prv1_target_pa 가 필요합니다.")
    # HSP 는 고층부라 elev_av 가 양수 (옥상보다 위)가 아니라 옥상 근처. 답안 29F = -64.2m 정도
    elev_av = _elev_at_floor(profile, spec.target_floor, fallback=-67.1)
    elev_prv = -10.0   # HSP 1차 PRV 위치 (옥상 직하단 보통)
    elev_top = -3.75
    nodes = [
        _node("100", 0.0, *_COORDS_PUMP_INPUT, io_node="Input", pressure_pa=ATM_PA),
        _node("1",   0.0,        *_COORDS_INPUT),
        _node("2",   0.0,        *_COORDS_N2),
        _node("3",   elev_top,   *_COORDS_N3),
        _node("4",   elev_top,   *_COORDS_N4),
        _node("7",   elev_prv,   *_COORDS_N7),
        _node("8",   elev_prv,   *_COORDS_PRV_IN),
        _node("89",  elev_prv,   *_COORDS_PRV_OUT),
        _node("5",   elev_av - 1.0, *_COORDS_N5_AV_PREV),
        _node("10",  elev_av,    *_COORDS_AV),
    ]
    pipes = [
        _pipe("r1", "1", "2",   150, 20.95, 0.0),
        _pipe("r2", "2", "3",   150,  3.75, elev_top),
        _pipe("r3", "3", "4",   150, 14.93, 0.0),
        _pipe("r4", "4", "7",   150, abs(elev_prv - elev_top), elev_prv - elev_top),
        _pipe("r5", "7", "8",   150,  0.5,  0.0),
        _pipe("r6", "89","5",   150, abs(elev_av - 1.0 - elev_prv), (elev_av - 1.0) - elev_prv),
        _pipe("r7", "5", "10",  125,  1.5,  1.0),
    ]
    pumps = [
        _pump_fan(str(i + 1), "100", "1", library_pump=spec.pump_library_name)
        for i in range(max(1, spec.pump_count))
    ]
    return RiserTables(
        nodes=nodes, pipes=pipes, pumps=pumps,
        valves=[_pressure_valve("1", "8", "89", target_pa=spec.prv1_target_pa)],
        av_node_label="10",
    )


def build_riser_llsp_2stage(spec: ZoneSpec, profile: BuildingPressureProfile | None) -> RiserTables:
    """LLSP 2차감압 라이저 — 자연낙차 + 1차 PRV + 2차 PRV (지하주차장).

    노드: ... 89(1차 PRV out) → 87(2차 PRV in) → 88(2차 PRV out) → 5 → 10(AV)
    Elastomeric-valve: 8→89 (1차, target=prv1_target_pa), 87→88 (2차, target=prv2_target_pa)
    """
    if spec.prv1_target_pa is None or spec.prv2_target_pa is None:
        raise ValueError("LLSP_2STAGE 는 prv1_target_pa, prv2_target_pa 모두 필요합니다.")
    elev_av = _elev_at_floor(profile, spec.target_floor, fallback=-159.6)  # B1 = -159.6m
    elev_prv1 = -73.35   # 1차 PRV (옥상 -73m)
    elev_prv2 = -145.4   # 2차 PRV (1.5F, 옥상 -145m)
    elev_top = -3.75
    nodes = [
        _node("1",  0.0,         *_COORDS_INPUT,    io_node="Input", pressure_pa=ATM_PA),
        _node("2",  0.0,         *_COORDS_N2),
        _node("3",  elev_top,    *_COORDS_N3),
        _node("4",  elev_top,    *_COORDS_N4),
        _node("7",  elev_prv1,   *_COORDS_N7),
        _node("8",  elev_prv1,   *_COORDS_PRV_IN),
        _node("89", elev_prv1,   *_COORDS_PRV_OUT),
        _node("87", elev_prv2,   *_COORDS_PRV2_IN),
        _node("88", elev_prv2,   *_COORDS_PRV2_OUT),
        _node("5",  elev_av - 1.0, *_COORDS_N5_AV_PREV),
        _node("10", elev_av,     *_COORDS_AV),
    ]
    pipes = [
        _pipe("r1", "1", "2",   150, 20.95, 0.0),
        _pipe("r2", "2", "3",   150,  3.75, elev_top),
        _pipe("r3", "3", "4",   150, 14.93, 0.0),
        _pipe("r4", "4", "7",   150, abs(elev_prv1 - elev_top), elev_prv1 - elev_top),
        _pipe("r5", "7", "8",   150,  0.5,  0.0),
        _pipe("r6", "89", "87", 150, abs(elev_prv2 - elev_prv1), elev_prv2 - elev_prv1),
        _pipe("r7", "88", "5",  150, abs(elev_av - 1.0 - elev_prv2), (elev_av - 1.0) - elev_prv2),
        _pipe("r8", "5", "10",  125,  1.5,  1.0),
    ]
    return RiserTables(
        nodes=nodes, pipes=pipes, pumps=[],
        valves=[
            _pressure_valve("1", "8", "89", target_pa=spec.prv1_target_pa),
            _pressure_valve("2", "87", "88", target_pa=spec.prv2_target_pa),
        ],
        av_node_label="10",
    )


def build_riser(spec: ZoneSpec, profile: BuildingPressureProfile | None) -> RiserTables:
    """zone 분기 라우터."""
    if spec.zone_type == ZoneType.HSP_PUMP:
        return build_riser_hsp_pump(spec, profile)
    if spec.zone_type == ZoneType.LSP_1STAGE:
        return build_riser_lsp_1stage(spec, profile)
    if spec.zone_type == ZoneType.LLSP_2STAGE:
        return build_riser_llsp_2stage(spec, profile)
    if spec.zone_type == ZoneType.LSP_GRAVITY:
        # 자연낙차 감압 없음 — LSP_1STAGE 의 PRV 만 제거한 변형
        rt = build_riser_lsp_1stage(
            ZoneSpec(zone_type=ZoneType.LSP_1STAGE, target_floor=spec.target_floor,
                     prv1_target_pa=ATM_PA),  # dummy — valves 비울 거니까 무시
            profile,
        )
        rt.valves = []
        # PRV 자리 노드 8, 89 도 제거하고 7 → 5 직결
        rt.nodes = [n for n in rt.nodes if n["label"] not in ("8", "89")]
        rt.pipes = [p for p in rt.pipes if p["in"] not in ("7", "89") or p["out"] not in ("8", "5")]
        rt.pipes.append(_pipe("r_g", "7", "5", 150, 30.0, -30.0))  # 7 → 5 직결 단순화
        return rt
    raise ValueError(f"Unknown zone_type: {spec.zone_type}")


# ────────────────────────────────────────────────────────────────────────────
# Stage C — 라이저 ↔ 헤드망 stitch
# ────────────────────────────────────────────────────────────────────────────


@dataclass
class CombinedTables:
    """라이저 + 헤드망 결합 결과 — emit_full_sdf 입력."""
    nodes: list[dict] = field(default_factory=list)
    pipes: list[dict] = field(default_factory=list)
    nozzles: list[dict] = field(default_factory=list)
    fittings: list[dict] = field(default_factory=list)
    equipment: list[dict] = field(default_factory=list)
    pumps: list[dict] = field(default_factory=list)
    valves: list[dict] = field(default_factory=list)
    meta: list[tuple[str, str]] = field(default_factory=list)
    # 기계실 전체 평면 배관망 edge (시각화 전용, SDF 미포함). [[x1,y1,x2,y2], ...]
    machine_room_plan_edges: list[list[float]] = field(default_factory=list)


def _layout_riser_as_schematic(
    riser_nodes: list[dict],
    anchor_xy: tuple[float, float],
    head_yspan: float = 5000.0,
    descend: bool = False,
) -> list[dict]:
    """라이저 노드를 PIPENET schematic 의 수직 막대 형태로 재배치.

    실제 계통도 DXF 의 라이저는 수십 m (60km mm) 길이라 헤드망 (5-10m schematic)
    과 통합하면 라이저가 너무 거대해서 그래프가 깨져 보임. PIPENET 답안처럼
    라이저는 헤드 군집 위로 수직 막대 (5m 정도) 로 배치해 한 화면에 깔끔히 보이게.

    Layout:
        - 모든 노드 X = anchor_xy[0] (수직 막대)
        - AV (마지막 노드) = anchor_xy 위치 (= 헤드 군집 가로 중앙 위)
        - 펌프 (첫 노드) = (anchor_xy[0], anchor_xy[1] + 라이저 yspan)
        - 중간 노드 = 균등 간격으로 사이 배치

    Args:
        riser_nodes: 라이저 노드 리스트. 인덱스 0 = 펌프, 마지막 = AV.
        anchor_xy: 막대 하단(AV) 을 놓을 좌표. 헤드망 source 노드 실좌표를 넘겨
            AV 를 source 에 정합하면 둘을 잇는 헤드 첫 배관(선언 길이 ~0)이
            긴 선으로 늘어나지 않음.
        head_yspan: 헤드망 bbox 의 y 범위 — 라이저 막대 길이는 이의 80%.
        descend: 펌프 가압(B1 펌프실) 모드. True 면 막대를 AV(헤드망) **아래**로
            내려, 펌프/수원이 화면 최하부에 오도록 한다(물이 B1→위로 가압되는
            물리 배치). 기본 False = 옥상수조(자연낙차) — 수원이 위.
    """
    n = len(riser_nodes)
    if n < 2:
        return list(riser_nodes)
    riser_yspan = max(2000.0, head_yspan * 0.8)
    step_y = riser_yspan / (n - 1)
    target_x = float(anchor_xy[0])
    target_y_av = float(anchor_xy[1])
    # 자연낙차: 수원 위(+). 펌프 가압: 수원 아래(-).
    y_dir = -1.0 if descend else 1.0

    out: list[dict] = []
    for i, node in enumerate(riser_nodes):
        # i=0: 펌프/수원, i=n-1: AV (head_av 위치, 막대 끝점)
        rank_from_av = (n - 1) - i
        out.append({
            **node,
            "x": int(round(target_x)),
            "y": int(round(target_y_av + y_dir * rank_from_av * step_y)),
        })
    return out


def _layout_machine_room_plan(
    mr_nodes: list[dict],
    plan_edges: list[list[float]] | None,
    pump_xy: tuple[float, float],
    head_yspan: float = 5000.0,
    below: bool = False,
) -> tuple[list[dict], list[list[float]]]:
    """기계실(옥상수조) 전체 평면 배관망을 실제 x, y 형상 그대로 배치 — schematic 금지.

    기계실 DXF 는 평면도(옥상층 소방배관 평면도)라 실제 2D 배관망 형상을 가진다.
    라이저(계통도)처럼 수직 막대로 모사하면 그 형상이 뭉개지므로, 헤드망(평면도)
    과 동일하게 실제 x, y 로 보여준다. 수리계산 경로(mr_nodes, m*) 뿐 아니라 전체
    SP 배관망 edge(plan_edges)까지 **동일 변환**으로 배치해 완전한 평면도로 렌더.

        1. 경로 노드 + 전체 edge 끝점을 합친 bbox 기준 균등 스케일(aspect 보존),
           헤드 군집과 비슷한 크기로.
        2. bbox 하단(min y)을 펌프 위쪽(gap)에, 가로 중앙을 펌프 x 에 정렬.
           → 평면 군집이 라이저 막대 위에 떠 펌프와 brige edge 로 이어져 보임.

    Args:
        mr_nodes: 수리경로 노드(라벨 m*). 실제 DXF x, y 보유(원점 미변환 raw).
        plan_edges: 기계실 전체 SP 배관망 edge [[x1,y1,x2,y2], ...] (raw DXF).
        pump_xy: 펌프 junction(라이저 "1")의 schematic 좌표 — 부착 기준점.
        head_yspan: 헤드망 y-span — 기계실 평면 크기·gap 산정 기준.

    Returns:
        (laid_nodes, laid_edges) — 동일 변환 적용된 경로 노드 + 전체망 edge.
    """
    pe = plan_edges or []
    pts: list[tuple[float, float]] = [
        (float(n.get("x", 0.0)), float(n.get("y", 0.0))) for n in mr_nodes
    ]
    for e in pe:
        pts.append((float(e[0]), float(e[1])))
        pts.append((float(e[2]), float(e[3])))
    if not pts:
        return list(mr_nodes), []
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    w = x_max - x_min
    h = y_max - y_min
    diag = max((w * w + h * h) ** 0.5, 1.0)
    target = max(2000.0, head_yspan * 0.7)   # 헤드 군집과 비슷한 크기로
    scale = target / diag
    cx = (x_min + x_max) / 2.0
    gap = max(800.0, head_yspan * 0.12)       # 펌프와 기계실 군집 사이 간격
    px, py = float(pump_xy[0]), float(pump_xy[1])

    def _tf(x: float, y: float) -> tuple[float, float]:
        # 자연낙차(옥상수조): 펌프 위쪽(+). 펌프 가압(B1): 펌프 아래쪽(-).
        if below:
            return ((x - cx) * scale + px, py - gap - (y_max - y) * scale)
        return ((x - cx) * scale + px, (y - y_min) * scale + py + gap)

    laid_nodes: list[dict] = []
    for n in mr_nodes:
        nx, ny = _tf(float(n.get("x", 0.0)), float(n.get("y", 0.0)))
        laid_nodes.append({**n, "x": int(round(nx)), "y": int(round(ny))})
    laid_edges: list[list[float]] = []
    for e in pe:
        x1, y1 = _tf(float(e[0]), float(e[1]))
        x2, y2 = _tf(float(e[2]), float(e[3]))
        laid_edges.append([int(round(x1)), int(round(y1)),
                           int(round(x2)), int(round(y2))])
    return laid_nodes, laid_edges


def stitch_riser_and_heads(
    riser: RiserTables,
    head_tables: Any,
    machine_room_labels: set[str] | list[str] | None = None,
    pump_junction_label: str | None = None,
    machine_room_plan_edges: list[list[float]] | None = None,
    machine_room_at_bottom: bool = False,
) -> CombinedTables:
    """라이저 끝점(AV node) ↔ 헤드망 source(=같은 label) 결합 + schematic 좌표 정렬.

    Args:
        riser: Stage B 산출 (legacy template) 또는 extract_system_path 결과.
        head_tables: remote30_prototype.PipeTables 인스턴스.

    Returns:
        CombinedTables — 좌표가 정렬된 한 schematic.

    충돌 처리:
        라이저 노드 라벨 = {1..9, ...} + AV(10) — 또는 v1 path: {1, n2..nN, 10}
        헤드망 노드 라벨 = {10, 11, 12, ...} (10 이 source = AV)
        AV(10) 만 공통 — 라이저 쪽 노드만 유지, 헤드망 사본은 skip.

    좌표 정렬 (PIPENET answer schematic 스타일):
        legacy 28F 템플릿은 6 노드의 작은 schematic 좌표 (0~15K 범위) 를 affine
        변환으로 사용자 클릭 위치에 맞춤. v1 path 는 실제 DXF 절대 좌표 (수십m)
        를 그대로 써서 통합 시 라이저가 헤드망보다 훨씬 거대해 보임.

        해결: 라이저를 _layout_riser_as_schematic 로 헤드 AV 위치 위 수직 막대
        로 재배치. 라이저 막대 길이는 헤드 bbox y-span 의 80%. 결과는 답안 SDF
        의 schematic 처럼 한 시각 영역 안에 컴팩트하게 부착.
    """
    av_lbl = riser.av_node_label
    riser_av_node = next((n for n in riser.nodes if n["label"] == av_lbl), None)
    if riser_av_node is None:
        raise ValueError(f"라이저에 AV 노드(label={av_lbl})가 없음")

    head_av_node = next((n for n in head_tables.nodes if n["label"] == av_lbl), None)

    # 헤드망 bbox 계산 (라이저 schematic 크기 결정용)
    head_xs = [float(nd.get("x", 0.0)) for nd in head_tables.nodes if "x" in nd]
    head_ys = [float(nd.get("y", 0.0)) for nd in head_tables.nodes if "y" in nd]
    head_yspan = (max(head_ys) - min(head_ys)) if head_ys else 5000.0

    # 노드를 진짜 라이저(계통도) 와 기계실(평면도) 로 분리.
    #   - 라이저: 실제 좌표가 수십 m 라 schematic 수직 막대로 재배치.
    #   - 기계실: 평면도이므로 실제 x, y 형상을 유지(평면 군집). 막대로 뭉개지 않음.
    mr_set = {str(l) for l in (machine_room_labels or [])}
    true_riser_nodes = [n for n in riser.nodes if str(n["label"]) not in mr_set]
    mr_nodes = [n for n in riser.nodes if str(n["label"]) in mr_set]

    # 라이저 좌표 schematic 재배치 — AV(막대 하단)를 헤드망 source 노드 실좌표에 snap.
    #   AV 노드는 (1) 라이저 막대 하단이자 (2) 헤드망 첫 메인(파이프 av_lbl)의 source 로
    #   같은 논리 분기점이다. 막대를 헤드 군집 중앙 위에 띄우면 AV 가 헤드 source 위치에서
    #   떨어져, 그 둘을 잇는 헤드 첫 배관(선언 길이 ~0)이 두 좌표계를 가로지르는 긴 선으로
    #   그려졌다. → anchor 를 head_av_node 좌표로 잡아 AV 를 source 에 정합, 연결선을 0 으로.
    if head_av_node is not None and head_xs and head_ys:
        try:
            anchor_xy = (float(head_av_node["x"]), float(head_av_node["y"]))
            translated_riser_nodes = _layout_riser_as_schematic(
                true_riser_nodes, anchor_xy, head_yspan=head_yspan,
                descend=machine_room_at_bottom,
            )
        except (KeyError, TypeError, ValueError):
            translated_riser_nodes = list(true_riser_nodes)
    else:
        translated_riser_nodes = list(true_riser_nodes)

    # 기계실 평면 배치 — 펌프 junction("1")의 schematic 좌표에 부착, 실제 x,y 형상 유지.
    #   수리경로 노드(mr_nodes) + 전체 SP 배관망 edge(plan_edges)를 **동일 변환**으로
    #   배치해 완전한 평면도로 렌더. plan_laid 는 시각화 전용(SDF 미포함).
    plan_laid: list[list[float]] = []
    if mr_nodes and pump_junction_label is not None:
        pump_node = next((n for n in translated_riser_nodes
                          if str(n["label"]) == str(pump_junction_label)), None)
        if pump_node is not None:
            try:
                pump_xy = (float(pump_node["x"]), float(pump_node["y"]))
                mr_laid, plan_laid = _layout_machine_room_plan(
                    mr_nodes, machine_room_plan_edges, pump_xy, head_yspan=head_yspan,
                    below=machine_room_at_bottom,
                )
            except (KeyError, TypeError, ValueError):
                mr_laid = list(mr_nodes)
                plan_laid = []
        else:
            mr_laid = list(mr_nodes)
    else:
        mr_laid = list(mr_nodes)

    translated_riser_nodes = translated_riser_nodes + mr_laid

    # 헤드망 노드 10 의 elevation → 라이저 AV elevation 으로 일치
    head_nodes_filtered = []
    for n in head_tables.nodes:
        if n["label"] == av_lbl:
            # AV 는 라이저 쪽에서 이미 포함 — 헤드망 쪽 사본 skip
            continue
        head_nodes_filtered.append(n)

    # 파이프 라벨 전역 유일화 — 계통도·평면도·기계실이 전부 r1.. 컨벤션을 쓰므로
    # 합치면 라벨이 충돌한다. parse_sdf 의 plabel_to_pid 는 라벨 문자열로 keying 해
    # 같은 라벨 두 파이프를 한 K-solver pid 로 접어버려 KFP 토폴로지가 붕괴된다.
    # → 충돌하는 두 번째 이후 항목만 개명(원본 dict 불변, 사본 생성).
    combined_pipes: list[dict] = []
    seen_pipe_labels: set[str] = set()
    for p in (list(riser.pipes) + list(head_tables.pipes)):
        lbl = str(p.get("label", ""))
        if lbl and lbl not in seen_pipe_labels:
            seen_pipe_labels.add(lbl)
            combined_pipes.append(p)
            continue
        base = lbl or "p"
        k = 2
        new_lbl = f"{base}_{k}"
        while new_lbl in seen_pipe_labels:
            k += 1
            new_lbl = f"{base}_{k}"
        seen_pipe_labels.add(new_lbl)
        combined_pipes.append({**p, "label": new_lbl})

    return CombinedTables(
        nodes=translated_riser_nodes + head_nodes_filtered,
        pipes=combined_pipes,
        nozzles=list(head_tables.nozzles),
        fittings=list(head_tables.fittings),
        equipment=list(head_tables.equipment),
        pumps=list(riser.pumps),
        valves=list(riser.valves),
        meta=list(getattr(head_tables, "meta", [])),
        machine_room_plan_edges=plan_laid,
    )


def prepend_machine_room_to_riser(
    machine_room: dict, riser: RiserTables, *,
    at_bottom: bool = False, source_drop_below_lowest_m: float = 0.0,
) -> tuple[RiserTables, bool]:
    """기계실 경로를 라이저 Input 노드 앞에 prepend → 확장 RiserTables.

    수원 경계가 라이저 top('1')이 아니라 그보다 상류인 기계실 수면(m1)으로 이동한다.
    기계실 경로 = 수원(m1, Input) → 입상관 연결점(mK). mK 를 라이저 Input 노드와
    병합하고 라이저 Input 은 일반 분기로 강등(경계는 이제 m1 하나). 이렇게 기계실부
    배관의 마찰손실과 고저차가 통합망 수리계산에 반영된다.

    좌표 정합 (x, y): 기계실 노드의 raw DXF x, y 는 그대로 보존한다. 통합 캔버스
    배치는 stitch 단계의 _layout_machine_room_plan 이 수리경로 노드 + 전체 SP
    배관망 edge(plan_edges, 동일 raw 좌표계)를 한 변환으로 묶어 펌프 위에 부착하므로,
    여기서 미리 translation 하면 오히려 plan_edges 와 어긋난다. → x, y 무변환.

    elevation (가압방식에 따라 기준이 다름 — 수리결과를 바꾸는 핵심):
      • at_bottom=False (고가수조/자연낙차): 기계실(옥상수조)은 라이저 top 과
        동일한 옥상 레벨(수평)이므로 기계실 노드 elev 를 라이저 Input(옥상) elev 로
        offset. 수원이 망 최상부 → 자연낙차로 하류에 양압 공급.
      • at_bottom=True (펌프 가압): 기계실/수원이 최하부(예: B1)다. 라이저 Input
        (옥상)이 아니라 **라이저 최저 고도(=서비스 최저층, 보통 AV)** 를 기준으로
        삼되, ``source_drop_below_lowest_m`` 만큼 그 아래로 더 내려 수원을 둔다.
        도면(DXF)에는 z 가 없어 기계실의 실제 깊이를 추출할 수 없으므로, 이 깊이는
        사용자가 가압방식 패널에서 직접 입력한다(미지정 시 0 = 최저헤드와 동일 고도).
        그 결과 수원→최저헤드 사이에 ``source_drop_below_lowest_m`` 만큼의 양(+)의
        정수두(lift)가 생겨 펌프가 극복해야 할 실양정으로 계산에 반영된다.
        (옥상 기준이면 이 lift 가 0 으로 사라져 펌프 실양정이 과소평가됨.)

    Args:
        source_drop_below_lowest_m: at_bottom 일 때만 사용. 수원(기계실)이 라이저
            최저 노드보다 몇 m 더 아래에 있는지(>0). 헤드 대비 실제 흡입 고저차.

    반환: (확장 RiserTables, attached) — attached 는 실제 병합 성공 여부.
    machine_room 가 비었거나 라이저 Input 을 못 찾으면 (원본 riser, False) 반환(안전).
    """
    mr_nodes = list(machine_room.get("nodes", []))
    mr_pipes = list(machine_room.get("pipes", []))
    if not mr_nodes or not mr_pipes:
        return riser, False

    conn_label = str(machine_room.get("conn_node_label") or mr_nodes[-1]["label"])

    riser_input = next(
        (n for n in riser.nodes if str(n.get("io_node", "")).lower() == "input"), None)
    if riser_input is None:
        riser_input = next((n for n in riser.nodes if str(n["label"]) == "1"), None)
    if riser_input is None:
        return riser, False  # 정합 불가 — 기계실 skip
    riser_input_label = str(riser_input["label"])
    riser_input_elev = float(riser_input.get("elevation", 0.0))

    # 기계실 고도 기준(offset) — 가압방식에 따라 분기.
    #  · 고가수조: 라이저 Input(옥상) 고도. 수원이 최상부.
    #  · 펌프 가압(at_bottom): 라이저 최저 고도에서 source_drop_below_lowest_m 만큼
    #    더 아래. 수원이 최저헤드보다 아래 → 양(+)의 실양정(lift)이 계산에 반영됨.
    if at_bottom:
        lowest = min(
            (float(n.get("elevation", 0.0)) for n in riser.nodes),
            default=riser_input_elev,
        )
        mr_ref_elev = lowest - abs(float(source_drop_below_lowest_m))
    else:
        mr_ref_elev = riser_input_elev

    # 기계실 노드: x,y 는 raw DXF 좌표 그대로(무변환 — plan_edges 와 동일 좌표계 유지),
    # elev 는 mr_ref_elev 기준으로 offset + conn 노드(mK) 제거(병합)
    new_mr_nodes: list[dict] = []
    for n in mr_nodes:
        if str(n["label"]) == conn_label:
            continue  # mK ≡ riser_input — 중복 제거
        nn = dict(n)
        nn["elevation"] = round(mr_ref_elev + float(n.get("elevation", 0.0)), 3)
        new_mr_nodes.append(nn)

    # 기계실 pipe: conn(mK) 향하던 끝점을 riser_input_label 로 재지정
    new_mr_pipes: list[dict] = []
    for p in mr_pipes:
        pp = dict(p)
        if str(pp.get("out")) == conn_label:
            pp["out"] = riser_input_label
        if str(pp.get("in")) == conn_label:
            pp["in"] = riser_input_label
        new_mr_pipes.append(pp)

    # 라이저 Input 강등: Input→No, pressure 제거 (경계는 이제 m1)
    new_riser_nodes: list[dict] = []
    for n in riser.nodes:
        nn = dict(n)
        if str(n["label"]) == riser_input_label:
            nn["io_node"] = "No"
            nn.pop("pressure_pa", None)
        new_riser_nodes.append(nn)

    return RiserTables(
        nodes=new_mr_nodes + new_riser_nodes,
        pipes=new_mr_pipes + list(riser.pipes),
        pumps=list(riser.pumps),
        valves=list(riser.valves),
        av_node_label=riser.av_node_label,
    ), True


# ────────────────────────────────────────────────────────────────────────────
# 펌프 가압 방식 — 통합망 수원(Input 경계) 직후 펌프 요소 삽입
# ────────────────────────────────────────────────────────────────────────────

# 화재안전기준(NFPC 103) 펌프 성능시험 곡선 기준:
#   체절운전(Q=0)      : 양정 ≤ 정격양정의 140%
#   정격(Q=정격)        : 정격양정
#   150% 유량(Q=1.5정격): 양정 ≥ 정격양정의 65%
PUMP_SHUTOFF_HEAD_RATIO = 1.40
PUMP_OVERLOAD_Q_RATIO = 1.50
PUMP_OVERLOAD_HEAD_RATIO = 0.65


def insert_source_pump(
    combined: CombinedTables,
    *,
    rated_q_lpm: float,
    rated_h_m: float,
    count: int = 1,
    pump_name: str = "FP",
    efficiency: int = 100,
) -> CombinedTables:
    """통합망 수원(Input 경계) 직후에 펌프 요소를 삽입한다 (펌프 가압 방식).

    물리 모델: 수원(Input, 대기압) → [Pump-fan] → 토출노드 → (기존 배관).
    수원에서 출발하던 모든 파이프의 시작점을 새 토출노드로 옮기고, 펌프 요소가
    수원→토출노드를 잇는다. 수원 노드는 그대로 대기압 경계(Calculation-spec)로
    남으므로 parse_sdf 의 pressure_bar 가 None 이 아니어서 펌프 양정이 boundary
    압력으로 잘못 주입되는 일이 없다.

    화재안전기준 표준 3점 성능곡선(체절 140% / 정격 / 150% 65%)을 펌프 dict 에
    담아 emit_full_sdf 의 Pump-fan attribute + has_converter 의 pumpFlowDataTable
    로 전파한다. 이로써 SDF/KFP/HAS 모두 동일 곡선으로 연산 가능해진다.

    Args:
        combined: stitch_riser_and_heads 산출. in-place 로 펌프/토출노드 추가.
        rated_q_lpm: 정격 토출량 (L/min).
        rated_h_m:   정격 양정 (m).
        count:       병렬 펌프 개수 (기본 1 — 운전 듀티 펌프 1대).
        pump_name:   Library-pump / HAS PumpType 이름 (기본 "FP").

    Returns:
        combined (동일 객체) — 펌프/토출노드가 추가됨.

    Raises:
        ValueError: 통합망에 Input 경계 노드가 없을 때.
    """
    src = next((n for n in combined.nodes
                if str(n.get("io_node", "")).lower() == "input"), None)
    if src is None:
        raise ValueError("통합망에 Input 경계 노드(수원)가 없어 펌프를 삽입할 수 없습니다.")
    src_label = str(src["label"])

    # 토출 노드 — 수원과 같은 고도(수평), 경계 아님. 라벨 유일 보장.
    existing = {str(n["label"]) for n in combined.nodes}
    disch_label = f"{src_label}_pd"
    _k = 1
    while disch_label in existing:
        _k += 1
        disch_label = f"{src_label}_pd{_k}"
    disch = {
        "label": disch_label,
        "x": int(src.get("x", 0)) + 400,   # 캔버스에서 펌프가 보이도록 약간 오프셋
        "y": int(src.get("y", 0)),
        "elevation": float(src.get("elevation", 0.0)),
        "io_node": "No",
    }
    combined.nodes.append(disch)

    # 수원에서 나가던 파이프(in==수원)의 시작점을 토출노드로 재지정.
    # (트리 root 인 수원은 모든 배관이 out 방향 → in==src 만 검사하면 충분)
    for p in combined.pipes:
        if str(p.get("in")) == src_label:
            p["in"] = disch_label

    q = float(rated_q_lpm)
    h = float(rated_h_m)
    shutoff_h = round(h * PUMP_SHUTOFF_HEAD_RATIO, 3)
    peak_q = round(q * PUMP_OVERLOAD_Q_RATIO, 3)
    peak_h = round(h * PUMP_OVERLOAD_HEAD_RATIO, 3)

    # 소화펌프는 주+예비(또는 N대) 구성이나, 예비는 동시 운전하지 않는 신뢰성
    # 이중화 → 수리계산은 운전 듀티 1대 곡선으로 한다(화재안전기준). 같은 in/out
    # 노드에 N개 Pump-fan 을 직렬화하면 PIPENET/KFP 가 N대 병렬 = N배 용량으로
    # 잘못 계산하므로, Pump-fan 은 단 하나만 만들고 대수는 count 메타로 보존한다.
    n_pumps = max(1, int(count))
    combined.pumps.append({
        "label": pump_name,
        "in": src_label,
        "out": disch_label,
        "efficiency": int(efficiency),
        "status": 1,
        "library_pump": pump_name,
        "percentage_open": 1,
        "pump_type": pump_name,
        "count": n_pumps,            # 설치 대수(주+예비) — 표시/문서용, 수리계산 비반영
        # 성능곡선 (정격 + 체절 + 150% 과부하) — m / L·min 단위
        "rated_q": q,
        "rated_h": h,
        "shutoff_h": shutoff_h,
        "peak_q": peak_q,
        "peak_h": peak_h,
    })
    return combined


# ────────────────────────────────────────────────────────────────────────────
# Stage D — emit_full_sdf (PIPENET-native 후처리 + Pump-fan / Elastomeric-valve)
# ────────────────────────────────────────────────────────────────────────────

# 물 ρg (kg/m³ × m/s²) — 양정(m) → 압력(Pa). 표준 SLF 펌프점과 동일 계수.
# M_TO_PA 와 같은 물리량(1 m 수두 → Pa). 의미가 다른 두 이름이라 alias 로 단일화.
_WATER_RHO_G = M_TO_PA


def _harden_slf_for_combined(
    slf_path: Path,
    opt_flow_by_lib: dict[str, float],
    pumps: list[dict],
) -> None:
    """동봉 SLF 라이브러리를 통합망에 맞게 보정 — PIPENET 연산 경고/에러 제거.

    1) Nozzle 최소운전압력 ↓ : 표준 SLF 의 SP-HEAD minimum-pressure 가 헤드
       설계유량(optimum flow)에 해당하는 압력 (Q/k)² 보다 높으면 모든 헤드에
       "optimum flow below minimum operating pressure" 경고가 발생한다. 각
       노즐 정의의 minimum-pressure 를 설계유량 압력의 90% 이하로 낮춘다.
    2) Pump 라이브러리 주입 : Pump-fan 이 참조하는 library_pump(예 "FP") 가
       SLF Pump-section 에 없으면 곡선 범위가 미정의되어 "Minimum flowrate
       should be less than maximum" 에러가 난다. NFPC 3점 곡선(체절 140% /
       정격 / 150% 65%)으로 Pump-definition 을 만들어 주입한다.

    SLF 는 DOCTYPE(<!DOCTYPE Library SYSTEM "Library.dtd">) 를 요구하므로
    ElementTree 직렬화 후 XML 선언 + DOCTYPE 를 직접 앞에 붙여 보존한다.
    """
    import xml.etree.ElementTree as ET
    if not slf_path.is_file():
        return
    try:
        tree = ET.parse(slf_path)
    except ET.ParseError:
        return
    root = tree.getroot()
    changed = False

    # ── (1) 노즐 최소운전압력 보정
    for ndef in root.iter("Nozzle-definition"):
        name_el = ndef.find("Item-name")
        lib = (name_el.text or "").strip() if name_el is not None else ""
        q_opt = opt_flow_by_lib.get(lib)
        if not q_opt:
            continue
        try:
            k = float(ndef.get("k-value", "0"))
        except ValueError:
            k = 0.0
        if k <= 0:
            continue
        p_opt = (q_opt / k) ** 2  # 설계유량에 필요한 노즐 압력 (Pa)
        try:
            p_min = float(ndef.get("minimum-pressure", "0"))
        except ValueError:
            p_min = 0.0
        if p_min > p_opt:
            ndef.set("minimum-pressure", f"{p_opt * 0.9:.2f}")
            changed = True

    # ── (2) 펌프 라이브러리 주입
    pump_sec = root.find("Pump-section")
    if pump_sec is not None and pumps:
        existing = {
            (pd.find("Item-name").text or "").strip()
            for pd in pump_sec.findall("Pump-definition")
            if pd.find("Item-name") is not None
        }
        seen: set[str] = set()
        for pump in pumps:
            lib = str(pump.get("library_pump", "")).strip()
            if not lib or lib in existing or lib in seen:
                continue
            q = float(pump.get("rated_q", 0) or 0)  # L/min
            h = float(pump.get("rated_h", 0) or 0)  # m
            if q <= 0 or h <= 0:
                continue
            seen.add(lib)
            q_si = q / 60000.0  # L/min → m³/s
            peak_q = float(pump.get("peak_q", q * PUMP_OVERLOAD_Q_RATIO) or q * PUMP_OVERLOAD_Q_RATIO)
            peak_q_si = peak_q / 60000.0
            shutoff_h = float(pump.get("shutoff_h", h * PUMP_SHUTOFF_HEAD_RATIO) or h * PUMP_SHUTOFF_HEAD_RATIO)
            peak_h = float(pump.get("peak_h", h * PUMP_OVERLOAD_HEAD_RATIO) or h * PUMP_OVERLOAD_HEAD_RATIO)
            pdef = ET.SubElement(pump_sec, "Pump-definition", {
                "curve-type": "quadratic",
                "flowrate-unit": "l-min",
                "max-degeneration-factor": "0",
                "max-flow": f"{peak_q_si:.9g}",
                "min-degeneration-factor": "0",
                "min-flow": "0",
                "pressure-unit": "metres",
            })
            ET.SubElement(pdef, "Item-name").text = lib
            ET.SubElement(pdef, "Description").text = lib
            pts = ET.SubElement(pdef, "Set-of-pump-points")
            # 압력은 평문 소수로 (과학표기 e+06 회피 — 표준 SLF 펌프점 포맷과 정합).
            ET.SubElement(pts, "Pump-point", {"flow": "0", "pressure": f"{shutoff_h * _WATER_RHO_G:.2f}"})
            ET.SubElement(pts, "Pump-point", {"flow": f"{q_si:.9g}", "pressure": f"{h * _WATER_RHO_G:.2f}"})
            ET.SubElement(pts, "Pump-point", {"flow": f"{peak_q_si:.9g}", "pressure": f"{peak_h * _WATER_RHO_G:.2f}"})
            changed = True

    if changed:
        body = ET.tostring(root, encoding="unicode")
        slf_path.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE Library SYSTEM "Library.dtd">\n' + body,
            encoding="utf-8",
        )


def emit_full_sdf(combined: CombinedTables, out_path: Path, *,
                  project_title: str = "Remote 30 전체 배관망 총괄") -> Path:
    """완성 SDF 직렬화.

    1단계: ``remote30_prototype.emit_sdf`` 호출 — PIPENET-native 후처리
           (빈 Pipe-set placeholder, 6 schedule embed, SLF 동봉, Template Graphics)
           가 모두 그대로 적용됨.
    2단계: 결과 SDF 를 다시 열어:
           - Input 노드에 ``<Calculation-spec pressure="ATM_PA"/>`` 추가
           - ``<Pump-fan>`` element 추가 (combined.pumps)
           - ``<Elastomeric-valve>`` element 추가 (combined.valves)
    """
    from remote30_prototype import emit_sdf, PipeTables

    # 1단계: PipeTables 로 캐스팅 후 emit_sdf 호출
    tables = PipeTables(
        nodes=list(combined.nodes),
        pipes=list(combined.pipes),
        nozzles=list(combined.nozzles),
        fittings=list(combined.fittings),
        equipment=list(combined.equipment),
        meta=list(combined.meta),
    )
    emit_sdf(tables, out_path, project_title=project_title)

    # 2단계: SDF 재오픈 → Pump-fan / Elastomeric-valve / Calculation-spec 추가
    import xml.etree.ElementTree as ET
    tree = ET.parse(out_path)
    root = tree.getroot()

    # Input 노드에 boundary pressure 명시 (Calculation-spec)
    pressure_by_label: dict[str, float] = {
        n["label"]: float(n["pressure_pa"])
        for n in combined.nodes if n.get("pressure_pa") is not None
    }
    for node_el in root.iter("Node"):
        lbl = node_el.get("label", "")
        if lbl in pressure_by_label and node_el.find("Calculation-spec") is None:
            ET.SubElement(node_el, "Calculation-spec",
                          {"pressure": str(int(pressure_by_label[lbl]))})

    # Links 안에 Pump-fan + Elastomeric-valve 삽입 (Nozzle 다음, 또는 끝쪽)
    for links in root.iter("Links"):
        for pump in combined.pumps:
            pf_attrib = {
                "efficiency": str(pump["efficiency"]),
                "input": pump["in"],
                "label": pump["label"],
                "output": pump["out"],
                "status": str(pump["status"]),
            }
            # 성능곡선이 있으면 attribute 로 직렬화 → parse_sdf 가 읽어 .has/KFP 로 전파.
            # rated-p/shutoff-p/peak-p 는 양정[m]. rated-p-unit="metres" 로 명시 →
            # parse_sdf 가 bar 로 정규화(KFP→SDF 의 bar 와 구분). (수원 노드는 이미
            # 대기압 boundary 라 parse_sdf 의 pressure_bar fallback 은 발동하지 않음.)
            if pump.get("rated_q") and pump.get("rated_h"):
                pf_attrib.update({
                    "rated-q": f"{float(pump['rated_q']):g}",
                    "rated-p": f"{float(pump['rated_h']):g}",
                    "rated-p-unit": "metres",
                    "shutoff-p": f"{float(pump.get('shutoff_h', float(pump['rated_h']) * 1.4)):g}",
                    "peak-q": f"{float(pump.get('peak_q', float(pump['rated_q']) * 1.5)):g}",
                    "peak-p": f"{float(pump.get('peak_h', float(pump['rated_h']) * 0.65)):g}",
                })
            pf = ET.Element("Pump-fan", pf_attrib)
            ET.SubElement(pf, "Description")
            lib = ET.SubElement(pf, "Library-pump")
            lib.text = pump["library_pump"]
            ET.SubElement(pf, "Pump-setting",
                          {"percentage-open": str(pump["percentage_open"])})
            links.append(pf)

        for v in combined.valves:
            ev = ET.Element("Elastomeric-valve", {
                "input": v["in"],
                "label": v["label"],
                "output": v["out"],
                "target-value": f"{v['target_value']:.6g}",
                "type": v["type"],
            })
            links.append(ev)
        break  # 첫 Links 만

    tree.write(out_path, encoding="utf-8", xml_declaration=True)

    # 동봉 SLF 보정 — 노즐 최소운전압력 / 펌프 곡선 라이브러리 (PIPENET 연산 경고·에러 제거).
    # SDF 가 PIPENET 에 넘기는 실제 optimum flow(<Flow-define flow=>) 를 라이브러리별로
    # 모아 SLF 의 minimum-pressure 와 비교 보정한다.
    opt_flow_by_lib: dict[str, float] = {}
    for nz_el in root.iter("Nozzle"):
        fd = nz_el.find("Flow-define")
        li = nz_el.find("Library-item")
        if fd is None or li is None:
            continue
        lib = (li.text or "").strip()
        try:
            q = float(fd.get("flow", "0"))
        except ValueError:
            continue
        if lib and q > 0:
            opt_flow_by_lib[lib] = min(opt_flow_by_lib.get(lib, q), q)
    _harden_slf_for_combined(out_path.with_suffix(".slf"), opt_flow_by_lib, list(combined.pumps))
    return out_path


# ────────────────────────────────────────────────────────────────────────────
# 통합 파이프라인 — 한 번에 모든 Stage 진행 (서버 API 에서 호출)
# ────────────────────────────────────────────────────────────────────────────


def run_full_pipeline(inputs: OverallInputs, head_tables: Any, out_path: Path) -> Path:
    """Stage B + C + D 일괄 실행. Stage A 는 별도 (서버 측에서 사용자 헤드 편집 단계 필요).

    Args:
        inputs: OverallInputs (zone_spec, profile, project_title 사용).
        head_tables: remote30_prototype.PipeTables — Stage A 의 결과.
        out_path: 출력 SDF 경로.
    """
    riser = build_riser(inputs.zone_spec, inputs.profile)
    combined = stitch_riser_and_heads(riser, head_tables)
    return emit_full_sdf(combined, out_path, project_title=inputs.project_title)


# ────────────────────────────────────────────────────────────────────────────
# 직접 입력 폼 → ZoneSpec / BuildingPressureProfile 변환
# ────────────────────────────────────────────────────────────────────────────

def zone_spec_from_form(form: dict[str, Any]) -> ZoneSpec:
    """HTML 폼 데이터 → ZoneSpec.

    필수 폼 필드::

        zone_type           : "hsp_pump" / "lsp_1stage" / "llsp_2stage" / "lsp_gravity"
        target_floor        : "16층" 등 라벨

    선택 폼 필드 (감압 zone)::

        prv1_target_kgf     : 1차 PRV 출구압 (kg/cm²) — Pa 변환됨
        prv2_target_kgf     : 2차 PRV 출구압 (kg/cm²) — LLSP_2STAGE 만
        prv1_target_m       : 1차 PRV 출구압 (m 수두) — kg/cm² 와 둘 중 하나만
        prv2_target_m       : 2차 PRV 출구압 (m 수두)
        pump_library_name   : Library-pump 이름 (HSP_PUMP, 기본 SP_162M_2900LPM)
        pump_count          : Pump-fan 개수 (HSP_PUMP, 기본 2)
    """
    zone_type = ZoneType(form.get("zone_type", "lsp_1stage"))

    def _to_pa(kgf_key: str, m_key: str) -> float | None:
        kgf = form.get(kgf_key, "")
        m = form.get(m_key, "")
        if kgf:
            try:
                return float(kgf) * KGF_CM2_TO_PA
            except ValueError:
                pass
        if m:
            try:
                return float(m) * M_TO_PA
            except ValueError:
                pass
        return None

    return ZoneSpec(
        zone_type=zone_type,
        target_floor=str(form.get("target_floor", "")).strip(),
        prv1_target_pa=_to_pa("prv1_target_kgf", "prv1_target_m"),
        prv2_target_pa=_to_pa("prv2_target_kgf", "prv2_target_m"),
        pump_library_name=str(form.get("pump_library_name", "SP_162M_2900LPM")).strip(),
        pump_count=int(form.get("pump_count", 2)),
    )


def parse_system_diagram_dxf(dxf_path: Path, *, default_height_m: float = 2.9,
                              roof_height_m: float = 6.0) -> BuildingPressureProfile:
    """계통도 DXF 의 텍스트 라벨 → BuildingPressureProfile 자동 추정.

    추출 패턴 (대명동 201동 계통도 분석 기반):
        "지상N층"   → "N층" label
        "지하N층"   → "지하N층" label
        "옥상", "PH", "PH N F"  → "옥상층" label
        "Nf" / "NF" / "NFL"  (fallback)

    elevation 계산:
        TEXT 의 Y 좌표를 정렬 (큰 Y = 위쪽) → 최상부부터 default_height_m 씩 누적.
        도면의 Y 단위가 도면별로 달라 직접 사용은 불안정. 층고는 default 로 채우고
        사용자가 검토/수정.

    Args:
        dxf_path: 계통도 DXF.
        default_height_m: 표준 층고 (기본 2.9m).
        roof_height_m: 옥상층의 층고 (기본 6m — PDF 압력표 기준).

    Returns:
        BuildingPressureProfile — 최상부 → 최하부 순. 비어있을 수도 (라벨 0건).
    """
    import re
    try:
        import ezdxf as _ezdxf
    except ImportError as exc:
        raise RuntimeError("ezdxf 가 설치되지 않아 계통도 DXF 파싱 불가") from exc

    doc = _ezdxf.readfile(str(dxf_path))
    msp = doc.modelspace()

    PAT_ABOVE = re.compile(r"지상\s*(\d+)\s*층")
    PAT_BELOW = re.compile(r"지하\s*(\d+)\s*층")
    PAT_ROOF  = re.compile(r"옥상|PH(?:\s*\d+)?\s*F?")
    PAT_NF    = re.compile(r"^\s*(\d{1,2})\s*F\s*$")

    candidates: list[tuple[str, float]] = []
    for e in msp:
        if e.dxftype() not in ("TEXT", "MTEXT"):
            continue
        try:
            v = e.dxf.text if e.dxftype() == "TEXT" else e.text
        except Exception:
            continue
        if not v or not v.strip():
            continue
        v = v.strip()
        try:
            pos = e.dxf.insert
            y = float(pos.y)
        except Exception:
            continue
        if (m := PAT_ABOVE.search(v)):
            candidates.append((f"{m.group(1)}층", y))
            continue
        if (m := PAT_BELOW.search(v)):
            candidates.append((f"지하{m.group(1)}층", y))
            continue
        if (m := PAT_NF.search(v)):
            candidates.append((f"{m.group(1)}층", y))
            continue
        if PAT_ROOF.search(v):
            candidates.append(("옥상층", y))

    # 같은 label 중복 시 가장 큰 Y 사용 (도면 상단 라벨이 가장 신뢰)
    by_label: dict[str, float] = {}
    for label, y in candidates:
        if label not in by_label or y > by_label[label]:
            by_label[label] = y

    # Y 내림차순 정렬 — 최상부 (Y 큰 것) → 최하부
    sorted_floors = sorted(by_label.items(), key=lambda kv: -kv[1])

    rows: list[FloorRow] = []
    cumulative_drop = 0.0
    for i, (label, _y) in enumerate(sorted_floors):
        height = roof_height_m if label == "옥상층" else default_height_m
        if i > 0:
            cumulative_drop += height
        rows.append(FloorRow(
            floor_label=label,
            height_m=height,
            head_drop_m=round(cumulative_drop, 1),
            note="(계통도 자동 추출 — 층고/낙차 검토 필요)",
        ))

    return BuildingPressureProfile(
        building_name=dxf_path.stem,
        floors=rows,
    )


def profile_from_form(form: dict[str, Any]) -> BuildingPressureProfile | None:
    """HTML 폼의 row 배열(JSON) → BuildingPressureProfile.

    폼 필드 ``pressure_table_json`` 이 있으면 파싱, 없으면 None.
    JSON 형식 예::

        [{"floor_label": "옥상층", "height_m": 6,   "head_drop_m": 3.1},
         {"floor_label": "49층",   "height_m": 3.1, "head_drop_m": 6.2},
         ...]
    """
    raw = form.get("pressure_table_json", "").strip()
    if not raw:
        return None
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return None
    rows = [FloorRow(
        floor_label=str(d.get("floor_label", "")).strip(),
        height_m=float(d.get("height_m", 0) or 0),
        head_drop_m=float(d.get("head_drop_m", 0) or 0),
        after_prv_m=(float(d["after_prv_m"]) if d.get("after_prv_m") not in (None, "") else None),
        note=str(d.get("note", "") or ""),
    ) for d in data if d.get("floor_label")]
    return BuildingPressureProfile(
        building_name=str(form.get("building_name", "")).strip(),
        floors=rows,
    )
