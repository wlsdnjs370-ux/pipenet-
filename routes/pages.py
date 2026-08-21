# -*- coding: utf-8 -*-
"""페이지·리포트·워크벤치 등 잔여(misc) 도메인 라우트.

`대조 서버.py` 에서 `register(app, ...)` 로 등록. 이 도메인 전용 헬퍼는
주입 이름(_sdf_*·_bbox 등)을 클로저로 참조하므로 register 안에 중첩한다.
공유 헬퍼·전역은 main 유지·주입. 라우트 본문·엔드포인트명 원본 그대로.
matplotlib Agg 백엔드는 main 이 프로세스 초기에 설정(그 뒤 이 모듈 import).
"""
from __future__ import annotations

import html as html_lib
import json
import math
import subprocess
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from pathlib import Path

import matplotlib.pyplot as plt
from flask import (jsonify, make_response, redirect, render_template,
                   request, send_file)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from cad_engine import locate_oda_exe
from core.upload_names import sanitize_upload_name
from cad_match import _cad_layer_weight
from pipenet_validator import PipenetGuideValidator

# 설계자동화 자식 서버는 로그인 게이트가 없다. 0.0.0.0 으로 띄우면 fncadnet.com
# 호스트의 7870 이 무인증으로 열린다 → 루프백 고정. 원격 접근이 필요해지면
# 리다이렉트 대신 로그인 게이트 뒤 리버스 프록시로 붙여야 한다.
DESIGN_AUTOMATION_BIND_HOST = "127.0.0.1"

# 모듈 E — CAD 프로젝트 편집기(Import.zip). PySide6 데스크톱 앱이라 브라우저
# 카드 안에 렌더할 수 없다. 카드 클릭 시 서버가 이 앱을 하위 프로세스로 띄우고,
# Qt 창은 서버(=이 PC)에서 열린다. 웹서버가 아니므로 포트 리다이렉트가 아니라
# 프로세스 실행으로 붙인다. 개발 Flask 단일 워커라 프로세스 핸들을 보관해
# 중복 실행을 막는다.
CAD_EDITOR_ROOT = Path(__file__).resolve().parent.parent / "cad_project_editor"
CAD_EDITOR_MAIN = CAD_EDITOR_ROOT / "main.py"
_cad_editor_proc: dict = {"handle": None}


def register(app, *, _analyze_sdf_sprinkler_network, DESIGN_AUTOMATION_PID_PATH, DESIGN_AUTOMATION_PORT, DESIGN_AUTOMATION_ROOT, DESIGN_AUTOMATION_SERVER_PATH, DESIGN_AUTOMATION_STDERR_PATH, DESIGN_AUTOMATION_STDOUT_PATH, EXPORT_SCHEMA, REMOTE30_OUTPUT_DIR, UPDATE_HISTORY_PATH, UPLOAD_DIR, _approx_arc_points, _bbox, _ensure_design_automation_static_layout, _entity_preview_row, _fig_to_data_url, _is_local_port_open, _load_cad_sdf_learning_profile, _mark_similar_cad_pipe_entities, _norm_point, _normalize_layer_name, _point_on_polyline, _save_upload, _sdf_av_node, _sdf_bore_reductions, _sdf_branch_nodes, _sdf_build_adjacency, _sdf_dijkstra, _sdf_farthest_heads, _sdf_fitting_stats, _sdf_graph_pipes, _sdf_length_checks, _sdf_parse_nodes, _sdf_parse_nozzles, _sdf_parse_pipes_equipment, _sdf_vertical_pipes, _to_float, _write_cad_sdf_learning_profile):

    def _start_design_automation_server() -> None:
        if _is_local_port_open(DESIGN_AUTOMATION_PORT):
            return
        _ensure_design_automation_static_layout()
        if not DESIGN_AUTOMATION_SERVER_PATH.exists():
            raise FileNotFoundError(f"Design automation server.py not found: {DESIGN_AUTOMATION_SERVER_PATH}")

        with DESIGN_AUTOMATION_STDOUT_PATH.open("ab") as stdout_fp, DESIGN_AUTOMATION_STDERR_PATH.open("ab") as stderr_fp:
            process = subprocess.Popen(
                [
                    sys.executable,
                    str(DESIGN_AUTOMATION_SERVER_PATH),
                    "--host",
                    DESIGN_AUTOMATION_BIND_HOST,
                    "--port",
                    str(DESIGN_AUTOMATION_PORT),
                ],
                cwd=str(DESIGN_AUTOMATION_ROOT),
                stdout=stdout_fp,
                stderr=stderr_fp,
            )
        DESIGN_AUTOMATION_PID_PATH.write_text(str(process.pid), encoding="utf-8")

        deadline = time.time() + 15
        while time.time() < deadline:
            if _is_local_port_open(DESIGN_AUTOMATION_PORT):
                return
            time.sleep(0.5)
        raise RuntimeError(
            f"Design automation server did not start on port {DESIGN_AUTOMATION_PORT}. "
            f"Check {DESIGN_AUTOMATION_STDERR_PATH.name}."
        )

    def _printable_report_text(path: Path) -> str:
        return PipenetGuideValidator(report_path=path)._read_report_text(path)

    def _print_report_url(path: Path, copies: int = 2) -> str:
        return f"/print-report/{path.name}?copies={copies}"

    def _load_update_history() -> dict:
        if not UPDATE_HISTORY_PATH.exists():
            return {
                "title": "업데이트 기록",
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "items": [],
            }
        # utf-8-sig — 이 파일은 Windows 편집기로 손대는 일이 잦아 BOM 이 붙는다.
        # 순수 utf-8 로 읽으면 JSONDecodeError 로 업데이트 기록 API 가 통째 500.
        with UPDATE_HISTORY_PATH.open("r", encoding="utf-8-sig") as fp:
            payload = json.load(fp)
        # 이 파일은 손으로 관리돼 스냅샷 dict 를 리스트에 덧붙인 형태로도 쌓인다.
        # 프론트는 단일 dict 만 받으므로 items 를 합쳐 한 장으로 접는다.
        if isinstance(payload, list):
            merged: list = []
            for snap in payload:
                if isinstance(snap, dict):
                    merged.extend(snap.get("items") or [])
            stamps = [str(s.get("updated_at")) for s in payload
                      if isinstance(s, dict) and s.get("updated_at")]
            payload = {"items": merged}
            if stamps:
                payload["updated_at"] = max(stamps)
        payload.setdefault("title", "업데이트 기록")
        payload.setdefault("updated_at", datetime.now().strftime("%Y-%m-%d %H:%M"))
        payload.setdefault("items", [])
        payload["items"] = sorted(
            payload["items"],
            key=lambda item: str(item.get("timestamp") or item.get("date") or ""),
            reverse=True,
        )
        return payload

    def _build_visualizations(validation: dict, report_path: Path, sdf_path: Path | None) -> list[dict]:
        tables = validation.get("tables") or {}
        visuals: list[dict] = []

        # 1) Pipe velocity vs limit
        pipe_rows = tables.get("pipes") or []
        pipe_labels: list[str] = []
        velocities: list[float] = []
        limits: list[float] = []
        for r in pipe_rows:
            label = str(r.get("label", ""))
            vel = _to_float(r.get("velocity_mps"), 0.0)
            pipe_labels.append(label)
            velocities.append(vel)
            limits.append(_to_float(r.get("velocity_limit_mps"), 0.0))
        if pipe_labels:
            fig, ax = plt.subplots(figsize=(10, 3.8))
            x = list(range(len(pipe_labels)))
            ax.plot(x, velocities, marker="o", linewidth=1.6, color="#1d4ed8", label="Velocity")
            ax.plot(x, limits, linestyle="--", linewidth=1.2, color="#dc2626", label="Limit")
            ax.set_title("Pipe Velocity vs Limit")
            ax.set_xlabel("Pipe Label")
            ax.set_ylabel("m/s")
            if len(pipe_labels) <= 30:
                ax.set_xticks(x, pipe_labels, rotation=0)
            else:
                step = max(1, len(pipe_labels) // 20)
                ticks = x[::step]
                ax.set_xticks(ticks, [pipe_labels[i] for i in ticks], rotation=0)
            ax.grid(alpha=0.25)
            ax.legend()
            visuals.append(
                {
                    "title": "Pipe Velocity Check",
                    "description": "Compares each pipe velocity with topology-based branch/other limits from the validator.",
                    "image_data_url": _fig_to_data_url(fig),
                }
            )

        # 2) Nozzle pressure-flow scatter
        noz_rows = tables.get("nozzles") or []
        pressures = [_to_float(r.get("inlet_pressure_kgf_cm2"), 0.0) for r in noz_rows]
        flows = [_to_float(r.get("actual_flow_lpm"), 0.0) for r in noz_rows]
        if pressures and flows:
            fig, ax = plt.subplots(figsize=(6.2, 4.2))
            colors = ["#dc2626" if _to_float(r.get("actual_flow_lpm"), 0.0) < 80.0 else "#16a34a" for r in noz_rows]
            ax.scatter(pressures, flows, c=colors, alpha=0.85)
            ax.axhline(80.0, color="#dc2626", linestyle="--", linewidth=1.2, label="80 L/min")
            ax.axvline(1.0, color="#f59e0b", linestyle="--", linewidth=1.2, label="1.0 kg/cm^2")
            ax.set_title("Nozzle Pressure-Flow Distribution")
            ax.set_xlabel("Inlet Pressure (kg/cm^2)")
            ax.set_ylabel("Actual Flow (L/min)")
            ax.grid(alpha=0.25)
            ax.legend(loc="best")
            visuals.append(
                {
                    "title": "Nozzle Pressure-Flow",
                    "description": "Green points pass the flow threshold, red points are below the flow threshold.",
                    "image_data_url": _fig_to_data_url(fig),
                }
            )

        return visuals

    def _build_engineering_visualizations(validation: dict) -> list[dict]:
        tables = validation.get("tables") or {}
        pipe_rows = tables.get("pipes") or []
        ratio_rows: list[dict] = []

        for row in pipe_rows:
            label = row.get("label")
            friction_loss = _to_float(row.get("friction_loss"), 0.0)
            length_m = _to_float(row.get("base_length_m") or row.get("pipe_length_m"), 0.0)
            if label is None or length_m <= 0:
                continue
            ratio = friction_loss / length_m
            ratio_rows.append(
                {
                    "label": int(label),
                    "ratio": ratio,
                    "friction_loss": friction_loss,
                    "length_m": length_m,
                    "velocity_mps": _to_float(row.get("velocity_mps"), 0.0),
                    "velocity_limit_mps": _to_float(row.get("velocity_limit_mps"), 0.0),
                    "nominal_bore_mm": _to_float(row.get("nominal_bore_mm"), 0.0),
                    "fitting_eq_length_m": _to_float(row.get("fitting_eq_length_m"), 0.0),
                    "special_eq_length_m": _to_float(row.get("special_eq_length_m"), 0.0),
                    "total_length_m": _to_float(row.get("total_length_m"), length_m),
                    "engineering_flag": bool(row.get("engineering_flag")),
                }
            )

        if not ratio_rows:
            return []

        ratio_rows.sort(key=lambda item: item["label"])
        labels = [str(item["label"]) for item in ratio_rows]
        values = [item["ratio"] for item in ratio_rows]
        colors = ["#2563eb" if item["engineering_flag"] else "#9ca3af" for item in ratio_rows]
        threshold = 1.0
        lengths = sorted(item["length_m"] for item in ratio_rows)
        median_length = lengths[len(lengths) // 2] if lengths else 0.0
        max_value = max(values) if values else threshold
        y_max = max(max_value * 1.12, threshold * 1.8)
        spike_points: list[dict] = []

        for idx, item in enumerate(ratio_rows):
            if idx == 0:
                continue
            previous = ratio_rows[idx - 1]
            prev_ratio = previous["ratio"]
            ratio = item["ratio"]
            delta = ratio - prev_ratio
            change_rate = delta / max(prev_ratio, 1e-9)
            spike_delta_limit = max(threshold, abs(prev_ratio) * 0.75)
            if ratio <= threshold or delta <= spike_delta_limit:
                continue

            eq_length = item["fitting_eq_length_m"] + item["special_eq_length_m"]
            eq_share = eq_length / max(item["total_length_m"], 1e-9)
            velocity_limit = item["velocity_limit_mps"]
            velocity_ratio = item["velocity_mps"] / velocity_limit if velocity_limit > 0 else 0.0
            causes: list[str] = []
            actions: list[str] = []

            if median_length > 0 and item["length_m"] >= median_length * 1.5:
                causes.append("긴 배관이라 총 마찰손실이 커질 수 있는 구간입니다.")
                actions.append("배관 경로를 단순화하거나 우회 길이를 줄여 실제 배관길이를 단축하는 방안을 검토하세요.")
            if median_length > 0 and item["length_m"] <= median_length * 0.6 and ratio > threshold:
                causes.append("짧은 배관인데 m당 마찰손실이 높아 손실이 국부적으로 집중된 구간입니다.")
                actions.append("해당 짧은 구간의 급격한 방향 전환, 국부 피팅, 특수설비 연결부를 우선 점검하세요.")
            if velocity_ratio >= 0.85:
                causes.append("유속이 적용 기준에 근접하여 마찰손실 증가에 크게 기여할 수 있습니다.")
                actions.append("구경 상향 또는 유량 분산으로 유속을 낮추는 대안을 검토하세요.")
            if eq_share >= 0.35:
                causes.append("피팅/특수설비 등가길이 비중이 커서 배관 자체 길이보다 부속 손실 영향이 큽니다.")
                actions.append("엘보/티/밸브/후렉시블 배관 수량을 줄이거나 손실이 작은 부속으로 변경하는 방안을 검토하세요.")
            if item["nominal_bore_mm"] <= 50 and item["velocity_mps"] >= 5.0:
                causes.append("소구경 배관에서 비교적 높은 유속이 발생하여 손실 집중 가능성이 있습니다.")
                actions.append("50A 이하 구간은 가지배관 기준 6m/s에 근접하는지 확인하고, 필요 시 한 단계 큰 구경을 검토하세요.")
            if not causes:
                causes.append("직전 배관 대비 m당 마찰손실 증가율이 커서 국부 조건 변화가 의심됩니다.")
                actions.append("해당 배관 전후의 구경 변화, 유속 변화, 피팅 수량, 특수설비 연결 여부를 함께 확인하세요.")

            spike_points.append(
                {
                    "label": item["label"],
                    "previous_label": previous["label"],
                    "ratio": ratio,
                    "previous_ratio": prev_ratio,
                    "delta": delta,
                    "change_rate_percent": change_rate * 100.0,
                    "data_index": idx,
                    "left_percent": 50.0,
                    "top_percent": 50.0,
                    "cards": {
                        "criteria": [
                            "직전 배관 대비 m당 마찰손실 변화율이 큰 구간을 급증 후보로 표시합니다.",
                            f"급증 조건: 현재 비율 > {threshold:.3f} kg/cm^2/m AND 증가량 > max({threshold:.3f}, 직전 비율 x 75%)",
                        ],
                        "formula": [
                            "m당 마찰손실 = FLOW IN PIPES Frict. Loss / PIPE CONFIGURATION Length",
                            "증가량 = 현재 m당 마찰손실 - 직전 배관 m당 마찰손실",
                            "변화율 = 증가량 / max(직전 배관 m당 마찰손실, 1e-9)",
                        ],
                        "values": [
                            f"Pipe {item['label']}: {item['friction_loss']:.4f} / {item['length_m']:.3f} = {ratio:.4f} kg/cm^2/m",
                            f"Previous Pipe {previous['label']}: {previous['friction_loss']:.4f} / {previous['length_m']:.3f} = {prev_ratio:.4f} kg/cm^2/m",
                            f"증가량 = {delta:.4f} kg/cm^2/m, 변화율 = {change_rate * 100.0:.1f}%",
                            f"구경 = {item['nominal_bore_mm']:.0f}A, 유속 = {item['velocity_mps']:.3f} m/s, 피팅+특수설비 등가길이 비중 = {eq_share * 100.0:.1f}%",
                        ],
                        "conclusion": [
                            *causes,
                            *actions,
                        ],
                    },
                }
            )

        fig, ax = plt.subplots(figsize=(11, 4.2))
        x = list(range(len(labels)))
        ax.bar(x, values, color=colors, width=0.78)
        ax.axhline(threshold, color="#dc2626", linestyle="--", linewidth=1.2, label="Threshold 1.00")
        if spike_points:
            spike_x = [labels.index(str(point["label"])) for point in spike_points if str(point["label"]) in labels]
            spike_y = [point["ratio"] for point in spike_points if str(point["label"]) in labels]
            ax.scatter(
                spike_x,
                spike_y,
                marker="v",
                s=26,
                color="#dc2626",
                edgecolor="#7f1d1d",
                linewidth=0.6,
                zorder=5,
                label="Sharp Increase",
            )
        ax.set_title("Friction Loss Ratio by Pipe")
        ax.set_xlabel("Pipe Label")
        ax.set_ylabel("Friction Loss / Length (kg/cm^2/m)")
        ax.set_ylim(0, y_max)
        if len(labels) <= 35:
            ax.set_xticks(x, labels, rotation=0)
        else:
            step = max(1, len(labels) // 24)
            ticks = x[::step]
            ax.set_xticks(ticks, [labels[i] for i in ticks], rotation=0)
        ax.grid(axis="y", alpha=0.25)
        ax.legend(loc="upper right")
        fig.canvas.draw()
        fig_w, fig_h = fig.canvas.get_width_height()
        if fig_w > 0 and fig_h > 0:
            for point in spike_points:
                px, py = ax.transData.transform((point["data_index"], point["ratio"]))
                point["left_percent"] = max(0.0, min(100.0, (px / fig_w) * 100.0))
                point["top_percent"] = max(0.0, min(100.0, ((fig_h - py) / fig_h) * 100.0))

        return [
            {
                "title": "Friction Loss Ratio by Pipe",
                "description": "Shows friction_loss / base_length_m for each pipe. Red markers indicate sharp increases from the previous pipe.",
                "image_data_url": _fig_to_data_url(fig, tight=False),
                "spike_points": spike_points,
            }
        ]

    def _build_sdf_graph(sdf_path: Path | None, tables: dict | None) -> dict:
        if sdf_path is None or not sdf_path.exists():
            return {}

        tables = tables or {}
        pipe_table = {int(r.get("label")): r for r in tables.get("pipes", []) if str(r.get("label", "")).isdigit()}
        nozzle_table = {int(r.get("label")): r for r in tables.get("nozzles", []) if str(r.get("label", "")).isdigit()}
        equipment_table = {int(r.get("label")): r for r in tables.get("equipment", []) if str(r.get("label", "")).isdigit()}
        valve_table = {int(r.get("label")): r for r in tables.get("valves", []) if str(r.get("label", "")).isdigit()}

        root = ET.parse(sdf_path).getroot()

        node_pos: dict[str, tuple[float, float]] = {}
        for node in root.findall(".//Node"):
            label = node.attrib.get("label")
            pos = node.find("Position")
            if not label or pos is None:
                continue
            try:
                x = float(pos.attrib.get("x", "0"))
                y = float(pos.attrib.get("y", "0"))
            except ValueError:
                continue
            node_pos[label] = (x, y)

        pipes: list[dict] = []
        pipe_paths: dict[int, list[tuple[float, float]]] = {}
        for pipe in root.findall(".//Pipe"):
            label_raw = pipe.attrib.get("label", "")
            if not label_raw.isdigit():
                continue
            label = int(label_raw)
            input_node = pipe.attrib.get("input", "")
            output_node = pipe.attrib.get("output", "")

            path: list[tuple[float, float]] = []
            if input_node in node_pos:
                path.append(node_pos[input_node])
            waypoints = pipe.find("Waypoints")
            if waypoints is not None:
                for wp in waypoints.findall("Position"):
                    try:
                        path.append((float(wp.attrib.get("x", "0")), float(wp.attrib.get("y", "0"))))
                    except ValueError:
                        continue
            if output_node in node_pos:
                path.append(node_pos[output_node])
            if len(path) < 2:
                continue

            pipe_paths[label] = path
            trow = pipe_table.get(label, {})
            status = "fail" if trow.get("highlight") else "pass"
            pipes.append(
                {
                    "label": label,
                    "input_node": input_node,
                    "output_node": output_node,
                    "path": [[x, y] for x, y in path],
                    "status": status,
                }
            )

        nozzles: list[dict] = []
        for nozzle in root.findall(".//Nozzle"):
            label_raw = nozzle.attrib.get("label", "")
            input_node = nozzle.attrib.get("input", "")
            if not label_raw.isdigit() or input_node not in node_pos:
                continue
            label = int(label_raw)
            x, y = node_pos[input_node]
            trow = nozzle_table.get(label, {})
            status = "fail" if trow.get("highlight") else "pass"
            nozzles.append({"label": label, "input_node": input_node, "x": x, "y": y, "status": status})

        equipment: list[dict] = []
        equipment_pos_by_label: dict[int, tuple[float, float]] = {}
        for eq in root.findall(".//Equipment"):
            label_raw = eq.attrib.get("label", "")
            if not label_raw.isdigit():
                continue
            label = int(label_raw)
            rel = float(eq.attrib.get("rel-position", "0.5"))
            desc = eq.attrib.get("description", "")
            table_row = equipment_table.get(label, {})
            pipe_label = table_row.get("pipe_label")
            if isinstance(pipe_label, str) and pipe_label.isdigit():
                pipe_label = int(pipe_label)
            if not isinstance(pipe_label, int):
                continue
            path = pipe_paths.get(pipe_label)
            if not path:
                continue
            p = _point_on_polyline(path, rel)
            if p is None:
                continue
            x, y = p
            equipment_pos_by_label[label] = (x, y)
            status = "warn" if table_row.get("warn") else ("fail" if table_row.get("highlight") else "pass")
            equipment.append(
                {
                    "label": label,
                    "description": desc,
                    "pipe_label": pipe_label,
                    "x": x,
                    "y": y,
                    "status": status,
                }
            )

        valves: list[dict] = []
        for label, row in valve_table.items():
            pos = equipment_pos_by_label.get(label)
            if pos is None:
                continue
            x, y = pos
            status = "fail" if row.get("highlight") else "pass"
            valves.append({"label": label, "x": x, "y": y, "status": status})

        return {
            "nodes": [{"id": nid, "x": x, "y": y} for nid, (x, y) in node_pos.items()],
            "pipes": pipes,
            "nozzles": nozzles,
            "equipment": equipment,
            "valves": valves,
        }

    def _build_cad_sdf_learning_profile(cad: dict, sdf: dict, source_sdf: Path | None = None, source_cad: Path | None = None) -> dict:
        entities = cad.get("drawing_entities") or []
        layer_stats: dict[str, dict] = {}
        for ent in entities:
            layer = ent.get("layer") or "0"
            stat = layer_stats.setdefault(layer, {"entity_count": 0, "total_length": 0.0, "similar_count": 0})
            stat["entity_count"] += 1
            stat["total_length"] += float(ent.get("draw_length") or 0.0)
            if ent.get("similar_to_sdf"):
                stat["similar_count"] += 1

        positive_layers: set[str] = set()
        suppressed_layers: set[str] = set()
        for layer, stat in layer_stats.items():
            norm = _normalize_layer_name(layer)
            has_fire_keyword = any(keyword in norm for keyword in ["SP", "소화", "배관", "후렉", "SPRINKLER", "FIRE"])
            entity_count = int(stat.get("entity_count") or 0)
            similar_count = int(stat.get("similar_count") or 0)
            if has_fire_keyword:
                positive_layers.add(layer)
            elif entity_count >= 1000 and similar_count >= 10:
                # Large generic architectural layers can accidentally resemble SDF
                # line geometry after normalization. Treat them as background noise.
                suppressed_layers.add(layer)
            elif norm in {"0", "L1", "L2", "L3", "L4"} and not has_fire_keyword:
                suppressed_layers.add(layer)

        profile = {
            "version": 1,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_sdf": str(source_sdf) if source_sdf else "",
            "source_cad": str(source_cad) if source_cad else "",
            "method": "sample_pair_layer_weighting",
            "positive_keywords": ["SP", "소화", "배관", "후렉", "SPRINKLER", "FIRE"],
            "positive_layers": sorted(positive_layers),
            "suppressed_layers": sorted(suppressed_layers),
            "layer_stats": {
                layer: {
                    "entity_count": int(stat["entity_count"]),
                    "total_length": round(float(stat["total_length"]), 3),
                    "similar_count": int(stat["similar_count"]),
                }
                for layer, stat in sorted(layer_stats.items())
            },
            "notes": [
                "H-100 단위세대 소방평면도_도면하나.dxf와 201동 3F SDF 샘플에서 추출한 CAD-SDF 대조 가중치입니다.",
                "소방/SP/배관/후렉 계열 레이어를 우선 배관 후보로 보고, L 계열 대량 건축선은 배경 잡음으로 낮게 평가합니다.",
                "일반 AI 모델 학습이 아니라 샘플 기반 휴리스틱 학습 프로필입니다. 여러 라벨링 샘플이 누적되면 세그멘테이션 모델 학습으로 확장할 수 있습니다.",
            ],
        }
        _write_cad_sdf_learning_profile(profile)
        return profile

    def _extract_cad_head_candidates(cad_path: Path) -> dict:
        # Lightweight DXF scan for the 7th module. It extracts enough geometry for a
        # quick drawing preview without invoking the heavier CAD graph engine.
        learning_profile = _load_cad_sdf_learning_profile()
        try:
            raw = cad_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        except UnicodeDecodeError:
            raw = cad_path.read_text(encoding="cp949", errors="ignore").splitlines()

        pairs: list[tuple[str, str]] = []
        for i in range(0, len(raw) - 1, 2):
            pairs.append((raw[i].strip(), raw[i + 1].strip()))

        entities: list[dict] = []
        in_entities = False
        current: dict | None = None

        def is_triangle_polyline(ent: dict) -> bool:
            if ent.get("type") not in {"LWPOLYLINE", "POLYLINE"} or not ent.get("closed"):
                return False
            points = ent.get("points") or []
            unique: list[tuple[float, float]] = []
            for point in points:
                try:
                    pt = (float(point[0]), float(point[1]))
                except Exception:
                    continue
                if not any(math.hypot(pt[0] - old[0], pt[1] - old[1]) < 1e-6 for old in unique):
                    unique.append(pt)
            if len(unique) != 3:
                return False
            xs = [pt[0] for pt in unique]
            ys = [pt[1] for pt in unique]
            w = max(xs) - min(xs)
            h = max(ys) - min(ys)
            diag = math.hypot(w, h)
            if diag <= 0 or diag > 1200 or min(w, h) <= 0:
                return False
            if max(w, h) / max(min(w, h), 1e-9) > 2.2:
                return False
            area = abs(
                unique[0][0] * (unique[1][1] - unique[2][1])
                + unique[1][0] * (unique[2][1] - unique[0][1])
                + unique[2][0] * (unique[0][1] - unique[1][1])
            ) / 2
            return area > 20 and area / max(w * h, 1e-9) > 0.18

        def flush() -> None:
            nonlocal current
            if not current:
                return
            etype = current.get("type")
            if etype in {"CIRCLE", "INSERT"} and current.get("x") is not None and current.get("y") is not None:
                entities.append(current)
            elif etype == "LINE" and current.get("x") is not None and current.get("y") is not None and current.get("x2") is not None and current.get("y2") is not None:
                entities.append(current)
            elif etype == "ARC" and current.get("x") is not None and current.get("y") is not None and current.get("radius") is not None:
                current["points"] = _approx_arc_points(current)
                if len(current["points"]) >= 2:
                    entities.append(current)
            elif etype in {"LWPOLYLINE", "POLYLINE"} and len(current.get("points") or []) >= 2:
                entities.append(current)
            current = None

        for code, value in pairs:
            if code == "0" and value == "SECTION":
                continue
            if code == "2" and value == "ENTITIES":
                in_entities = True
                continue
            if code == "0" and value == "ENDSEC":
                flush()
                in_entities = False
                continue
            if not in_entities:
                continue
            if code == "0":
                flush()
                if value in {"CIRCLE", "INSERT"}:
                    current = {"type": value, "layer": "0", "x": None, "y": None, "radius": 0.0}
                elif value == "ARC":
                    current = {"type": value, "layer": "0", "x": None, "y": None, "radius": 0.0, "start_angle": 0.0, "end_angle": 0.0}
                elif value == "LINE":
                    current = {"type": value, "layer": "0", "x": None, "y": None, "x2": None, "y2": None}
                elif value in {"LWPOLYLINE", "POLYLINE"}:
                    current = {"type": value, "layer": "0", "points": [], "_pending_x": None}
                continue
            if current is None:
                continue
            if code == "8":
                current["layer"] = value
            elif code == "10":
                if current.get("type") in {"LWPOLYLINE", "POLYLINE"}:
                    current["_pending_x"] = _to_float(value)
                else:
                    current["x"] = _to_float(value)
            elif code == "20":
                if current.get("type") in {"LWPOLYLINE", "POLYLINE"}:
                    px = current.pop("_pending_x", None)
                    if px is not None:
                        current.setdefault("points", []).append([px, _to_float(value)])
                else:
                    current["y"] = _to_float(value)
            elif code == "11":
                current["x2"] = _to_float(value)
            elif code == "21":
                current["y2"] = _to_float(value)
            elif code == "40":
                current["radius"] = _to_float(value)
            elif code == "50":
                current["start_angle"] = _to_float(value)
            elif code == "51":
                current["end_angle"] = _to_float(value)
            elif code == "70" and current.get("type") in {"LWPOLYLINE", "POLYLINE"}:
                current["closed"] = bool(int(_to_float(value)) & 1)
            elif code == "2" and current.get("type") == "INSERT":
                current["block"] = value
        flush()

        circles = [e for e in entities if e.get("type") == "CIRCLE"]
        inserts = [e for e in entities if e.get("type") == "INSERT"]
        triangles = [e for e in entities if is_triangle_polyline(e)]
        line_entities = [e for e in entities if e.get("type") in {"LINE", "LWPOLYLINE", "POLYLINE", "ARC"}]
        source = [*circles, *triangles] if len(circles) + len(triangles) >= 3 else [*inserts, *triangles]
        candidates: list[dict] = []
        for idx, ent in enumerate(source, start=1):
            if ent.get("type") in {"LWPOLYLINE", "POLYLINE"}:
                xs = [p[0] for p in ent.get("points") or []]
                ys = [p[1] for p in ent.get("points") or []]
                x = sum(xs) / len(xs) if xs else 0.0
                y = sum(ys) / len(ys) if ys else 0.0
            else:
                x = _to_float(ent.get("x"))
                y = _to_float(ent.get("y"))
            candidates.append(
                {
                    "label": str(idx),
                    "entity_id": str(idx),
                    "type": ent.get("type", ""),
                    "layer": ent.get("layer", ""),
                    "x": x,
                    "y": y,
                    "radius": _to_float(ent.get("radius")),
                }
            )

        if len(candidates) > 300:
            # Keep the comparison responsive. The user should filter the DXF to sprinkler/head layers for precision.
            candidates = candidates[:300]

        drawing_entities = []
        for idx, ent in enumerate(line_entities, start=1):
            row = _entity_preview_row(ent, idx)
            if row:
                row["layer_weight"] = _cad_layer_weight(row.get("layer"), learning_profile)
                drawing_entities.append(row)

        # Full DXF drawings can contain 100k+ entities. Learned fire/SP layers are
        # retained first; within the same weight keep longer pipe-like geometry first.
        drawing_entities.sort(key=lambda item: (item.get("layer_weight", 0.0), item.get("draw_length", 0.0)), reverse=True)
        drawing_entities_for_preview = drawing_entities[:20000]

        preview_points: list[dict] = []
        for ent in drawing_entities_for_preview:
            for x, y in ent.get("points") or []:
                preview_points.append({"x": x, "y": y})
        preview_points.extend(candidates)

        return {
            "filename": cad_path.name,
            "bounds": _bbox(preview_points) or _bbox(candidates) or {},
            "layers": sorted({e.get("layer", "") for e in entities if e.get("layer")}),
            "network_layers": sorted(learning_profile.get("positive_layers", [])),
            "learned_profile_applied": bool(learning_profile),
            "learned_profile_updated_at": learning_profile.get("updated_at"),
            "learned_positive_layers": learning_profile.get("positive_layers", []),
            "learned_suppressed_layers": learning_profile.get("suppressed_layers", []),
            "raw_circle_count": len(circles),
            "raw_triangle_count": len(triangles),
            "raw_insert_count": len(inserts),
            "raw_line_count": len(line_entities),
            "drawing_entity_count": len(drawing_entities),
            "drawing_entity_returned_count": len(drawing_entities_for_preview),
            "candidate_count": len(candidates),
            "candidates": candidates[:500],
            "drawing_entities": drawing_entities_for_preview,
        }

    def _compare_cad_heads_to_sdf(cad: dict, sdf: dict) -> dict:
        sdf_heads = sdf.get("farthest_heads") or sdf.get("nozzles") or []
        cad_heads = cad.get("candidates") or []
        sdf_box = _bbox(sdf_heads)
        cad_box = _bbox(cad_heads)
        if not sdf_box or not cad_box or not sdf_heads or not cad_heads:
            return {
                "status": "REVIEW",
                "message": "CAD 헤드 후보 또는 SDF 헤드 좌표가 부족하여 자동 비교를 보류했습니다.",
                "matches": [],
                "unmatched_sdf": sdf_heads,
                "unmatched_cad": cad_heads,
                "mismatch_count": 0,
            }

        unused = set(range(len(cad_heads)))
        matches = []
        for sdf_head in sdf_heads:
            sx, sy = _norm_point(sdf_head, sdf_box)
            best_idx = None
            best_dist = float("inf")
            for idx in unused:
                cx, cy = _norm_point(cad_heads[idx], cad_box)
                d = math.hypot(sx - cx, sy - cy)
                if d < best_dist:
                    best_idx = idx
                    best_dist = d
            if best_idx is None:
                continue
            unused.remove(best_idx)
            cad_head = cad_heads[best_idx]
            matches.append(
                {
                    "sdf_head": sdf_head.get("label"),
                    "sdf_node": sdf_head.get("input_node"),
                    "sdf_x": sdf_head.get("x"),
                    "sdf_y": sdf_head.get("y"),
                    "cad_candidate": cad_head.get("label"),
                    "cad_layer": cad_head.get("layer"),
                    "cad_type": cad_head.get("type"),
                    "cad_x": cad_head.get("x"),
                    "cad_y": cad_head.get("y"),
                    "normalized_error": round(best_dist, 4),
                    "status": "FAIL" if best_dist > 0.08 else "PASS",
                    "reason": "정규화 좌표 오차가 0.08을 초과합니다." if best_dist > 0.08 else "정규화 좌표 기준 최근접 매칭 허용범위 이내입니다.",
                }
            )

        pipe_shape_match = _mark_similar_cad_pipe_entities(cad, sdf)
        mismatch_count = sum(1 for m in matches if m["status"] == "FAIL")
        return {
            "status": "FAIL" if mismatch_count else "PASS",
            "message": f"SDF 최원단 헤드 {len(sdf_heads)}개와 CAD 헤드 후보 {len(cad_heads)}개를 정규화 좌표 기준으로 비교했습니다.",
            "matches": matches,
            "unmatched_sdf": sdf_heads[len(matches):],
            "unmatched_cad": [cad_heads[i] for i in sorted(unused)],
            "mismatch_count": mismatch_count,
            "pipe_shape_match": pipe_shape_match,
        }

    def _apply_sheet_style(ws):
        thin = Side(style="thin", color="C9CDD3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F3F4F6")
        title_fill = PatternFill("solid", fgColor="111827")
        fail_fill = PatternFill("solid", fgColor="FEE2E2")
        warn_fill = PatternFill("solid", fgColor="FEF3C7")
        eng_fill = PatternFill("solid", fgColor="DBEAFE")
        econ_fill = PatternFill("solid", fgColor="DCFCE7")

        max_col = ws.max_column
        max_row = ws.max_row

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(color="FFFFFF", bold=True, size=12)
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, max_col + 1):
            c = ws.cell(row=3, column=col)
            c.fill = header_fill
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        for row in range(4, max_row + 1):
            if max_col >= 4:
                flag_fail = ws.cell(row=row, column=max_col - 3).value == "Y"
                flag_warn = ws.cell(row=row, column=max_col - 2).value == "Y"
                flag_eng = ws.cell(row=row, column=max_col - 1).value == "Y"
                flag_econ = ws.cell(row=row, column=max_col).value == "Y"
            else:
                flag_fail = flag_warn = flag_eng = flag_econ = False

            row_fill = None
            if flag_fail:
                row_fill = fail_fill
            elif flag_warn:
                row_fill = warn_fill
            elif flag_eng:
                row_fill = eng_fill
            elif flag_econ:
                row_fill = econ_fill

            for col in range(1, max_col + 1):
                c = ws.cell(row=row, column=col)
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
                if row_fill:
                    c.fill = row_fill

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{ws.cell(row=3, column=max_col).column_letter}{max_row}"

        for col in range(1, max_col + 1):
            width = 12
            for row in range(1, max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                width = max(width, min(len(str(val)) + 2, 40))
            ws.column_dimensions[get_column_letter(col)].width = width

    @app.errorhandler(Exception)
    def _api_safe_errorhandler(exc):
        # Flask 의 HTTPException (400, 404 등) 은 그대로 전달
        from werkzeug.exceptions import HTTPException
        if request.path.startswith("/api/"):
            import traceback as _tb
            if isinstance(exc, HTTPException):
                return jsonify({
                    "ok": False,
                    "message": exc.description or str(exc),
                    "status": exc.code,
                }), exc.code
            from routes import traceback_for_client
            app.logger.error("Unhandled (%s): %s\n%s", request.path, exc, _tb.format_exc())
            body = {
                "ok": False,
                "message": f"서버 오류: {type(exc).__name__}: {str(exc)[:300]}",
            }
            tb = traceback_for_client()
            if tb:
                body["traceback"] = tb
            return jsonify(body), 500
        # /api/ 가 아니면 Flask 기본 처리 (HTML 페이지 OK)
        if isinstance(exc, HTTPException):
            return exc
        raise exc

    @app.get("/")
    def index():
        return render_template("index.html")

    @app.get("/sprinkler-pipeline")
    def sprinkler_pipeline():
        response = make_response(render_template("sprinkler_pipeline.html"))
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    @app.get("/design-automation-module-8")
    def design_automation_module_8():
        try:
            _start_design_automation_server()
        except Exception as exc:
            return (
                "설계자동화 인터페이스 서버를 시작하지 못했습니다. "
                f"원인: {html_lib.escape(str(exc))}",
                500,
            )
        return redirect(f"http://{DESIGN_AUTOMATION_BIND_HOST}:{DESIGN_AUTOMATION_PORT}/", code=302)

    @app.get("/module-e-cad-editor")
    def module_e_cad_editor():
        if not CAD_EDITOR_MAIN.exists():
            return (
                "CAD 프로젝트 편집기 프로그램을 찾을 수 없습니다: "
                f"{html_lib.escape(str(CAD_EDITOR_MAIN))}",
                500,
            )
        proc = _cad_editor_proc.get("handle")
        already_running = proc is not None and proc.poll() is None
        launch_error = ""
        if not already_running:
            try:
                _cad_editor_proc["handle"] = subprocess.Popen(
                    [sys.executable, str(CAD_EDITOR_MAIN)],
                    cwd=str(CAD_EDITOR_ROOT),
                )
            except Exception as exc:
                launch_error = str(exc)

        if launch_error:
            status_line = (
                "<p class=\"err\">편집기를 실행하지 못했습니다: "
                f"{html_lib.escape(launch_error)}</p>"
            )
        elif already_running:
            status_line = "<p class=\"ok\">CAD 프로젝트 편집기가 이미 실행 중입니다. 서버 화면에서 창을 확인하세요.</p>"
        else:
            status_line = "<p class=\"ok\">CAD 프로젝트 편집기를 실행했습니다. 서버(이 PC) 화면에 창이 열립니다.</p>"

        page = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <title>Module E · CAD 프로젝트 편집기</title>
  <style>
    body {{ margin:0; min-height:100vh; display:flex; align-items:center; justify-content:center;
           background:#0a1228; color:#e5e7eb; font-family:"Malgun Gothic","맑은 고딕",sans-serif; }}
    .box {{ max-width:560px; padding:36px 40px; background:#111827; border:1px solid #1f2937;
            border-radius:16px; box-shadow:0 18px 60px rgba(0,0,0,.45); }}
    .chip {{ display:inline-block; font-size:12px; letter-spacing:.14em; font-weight:800; color:#93c5fd;
             border:1px solid #1d4ed8; border-radius:999px; padding:5px 12px; margin-bottom:14px; }}
    h1 {{ margin:0 0 10px; font-size:22px; }}
    p {{ margin:8px 0; line-height:1.6; font-size:14px; color:#cbd5e1; }}
    .ok {{ color:#86efac; font-weight:700; }}
    .err {{ color:#fca5a5; font-weight:700; }}
    .note {{ font-size:12.5px; color:#94a3b8; }}
    a.btn {{ display:inline-block; margin-top:18px; padding:10px 18px; background:#1d4ed8; color:#fff;
             font-weight:800; text-decoration:none; border-radius:10px; }}
  </style>
</head>
<body>
  <div class="box">
    <span class="chip">MODULE E</span>
    <h1>CAD 프로젝트 편집기</h1>
    {status_line}
    <p class="note">DXF를 불러와 헤드를 매핑하고, 편집한 결과를 K-solver .kfp로 저장하는 데스크톱 프로그램입니다. 인터넷 연결 없이 동작합니다.</p>
    <p class="note">데스크톱 프로그램이라 브라우저 안에는 표시되지 않고, 서버를 실행 중인 PC 화면에 창으로 열립니다.</p>
    <a class="btn" href="/" onclick="window.close();return false;">이 창 닫기</a>
  </div>
</body>
</html>"""
        response = make_response(page)
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return response

    @app.get("/print-report/<path:filename>")
    def print_report(filename: str):
        safe_name = Path(filename).name
        report_path = UPLOAD_DIR / safe_name
        if not report_path.exists() or report_path.suffix.lower() not in {".docx", ".pdf"}:
            return "출력할 결과서 파일을 찾을 수 없습니다.", 404
        try:
            copies = max(1, min(int(request.args.get("copies", "2")), 10))
        except ValueError:
            copies = 2
        try:
            text = _printable_report_text(report_path)
        except Exception as exc:
            return f"결과서 내용을 읽을 수 없습니다: {html_lib.escape(str(exc))}", 500

        body = "\n".join(html_lib.escape(line.rstrip()) for line in text.splitlines())
        title = html_lib.escape(report_path.name)
        copy_blocks = []
        for idx in range(copies):
            page_break = " page-break-before: always;" if idx else ""
            copy_blocks.append(
                f"""
                <section class="print-copy" style="{page_break}">
                  <header class="print-head">
                    <div>
                      <p>PIPENET REPORT PRINT</p>
                      <h1>{title}</h1>
                    </div>
                    <strong>{idx + 1}/{copies}부</strong>
                  </header>
                  <pre>{body}</pre>
                </section>
                """
            )
        html = f"""<!doctype html>
    <html lang="ko">
    <head>
      <meta charset="utf-8">
      <title>{title} 출력</title>
      <style>
        @page {{ size: A4 portrait; margin: 12mm; }}
        * {{ box-sizing: border-box; }}
        body {{ margin: 0; background: #f3f4f6; color: #111827; font-family: "Malgun Gothic", "맑은 고딕", sans-serif; }}
        .print-toolbar {{ position: sticky; top: 0; z-index: 5; display: flex; justify-content: space-between; align-items: center; gap: 12px; padding: 14px 18px; border-bottom: 1px solid #d1d5db; background: #fff; }}
        .print-toolbar strong {{ font-size: 14px; }}
        .print-toolbar button {{ border: 1px solid #111827; background: #111827; color: #fff; padding: 10px 16px; font-weight: 800; cursor: pointer; }}
        .print-copy {{ width: 210mm; min-height: 297mm; margin: 16px auto; padding: 14mm; background: #fff; border: 1px solid #d1d5db; }}
        .print-head {{ display: flex; justify-content: space-between; align-items: flex-start; gap: 12px; margin-bottom: 12px; padding-bottom: 10px; border-bottom: 2px solid #111827; }}
        .print-head p {{ margin: 0 0 4px; font-size: 10px; letter-spacing: .12em; color: #6b7280; }}
        .print-head h1 {{ margin: 0; font-size: 18px; line-height: 1.35; }}
        .print-head strong {{ border: 1px solid #111827; padding: 6px 10px; font-size: 12px; white-space: nowrap; }}
        pre {{ margin: 0; white-space: pre-wrap; word-break: break-word; font-family: "Consolas", "D2Coding", "Malgun Gothic", monospace; font-size: 9.5pt; line-height: 1.38; }}
        @media print {{
          body {{ background: #fff; }}
          .print-toolbar {{ display: none; }}
          .print-copy {{ width: auto; min-height: auto; margin: 0; padding: 0; border: 0; }}
        }}
      </style>
    </head>
    <body>
      <div class="print-toolbar">
        <strong>결과서 전체 내용 출력 - {copies}부</strong>
        <button type="button" onclick="window.print()">프린트 실행</button>
      </div>
      {''.join(copy_blocks)}
      <script>window.addEventListener('load', () => setTimeout(() => window.print(), 350));</script>
    </body>
    </html>"""
        response = make_response(html)
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return response

    @app.get("/api/update-history")
    def update_history():
        return jsonify({"ok": True, "history": _load_update_history()})

    @app.post("/api/sdf-sprinkler-analysis")
    def sdf_sprinkler_analysis():
        try:
            sdf_path = _save_upload("sdf_file", {".sdf"}, required=True)
            analysis = _analyze_sdf_sprinkler_network(sdf_path)
            cad_analysis = None
            comparison = None
            cad_path = _save_upload("cad_file", {".dxf", ".dwg"}, required=False)
            if cad_path is not None:
                cad_analysis = _extract_cad_head_candidates(cad_path)
                comparison = _compare_cad_heads_to_sdf(cad_analysis, analysis)
                if not cad_analysis.get("learned_profile_applied"):
                    profile = _build_cad_sdf_learning_profile(cad_analysis, analysis, sdf_path, cad_path)
                    for ent in cad_analysis.get("drawing_entities") or []:
                        ent["layer_weight"] = _cad_layer_weight(ent.get("layer"), profile)
                        ent.pop("similar_to_sdf", None)
                        ent.pop("similarity_score", None)
                    cad_analysis.update(
                        {
                            "network_layers": sorted(profile.get("positive_layers", [])),
                            "learned_profile_applied": True,
                            "learned_profile_updated_at": profile.get("updated_at"),
                            "learned_positive_layers": profile.get("positive_layers", []),
                            "learned_suppressed_layers": profile.get("suppressed_layers", []),
                        }
                    )
                    comparison = _compare_cad_heads_to_sdf(cad_analysis, analysis)
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 400
        except Exception as exc:
            return jsonify({"ok": False, "message": f"SDF 분석 중 오류가 발생했습니다: {exc}"}), 500

        return jsonify(
            {
                "ok": True,
                "message": "SDF 스프링클러 배관 분석이 완료되었습니다.",
                "analysis": analysis,
                "cad_analysis": cad_analysis,
                "comparison": comparison,
            }
        )

    @app.post("/api/validate")
    def validate_files():
        try:
            report_path = _save_upload("report_file", {".docx", ".pdf"}, required=True)
            sdf_path = _save_upload("sdf_file", {".sdf"}, required=False)
            validation = PipenetGuideValidator(report_path=report_path, sdf_path=sdf_path).validate()
            sdf_graph = _build_sdf_graph(sdf_path, validation.get("tables"))
            visualizations = _build_visualizations(validation, report_path, sdf_path)
            engineering_visualizations = _build_engineering_visualizations(validation)
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 400
        except Exception as exc:
            # 어떤 예외도 잡아 JSON 으로 반환 — 절대 HTML 500 페이지로 빠지지 않게.
            import traceback as _tb
            return jsonify({
                "ok": False,
                "message": f"검증 중 오류가 발생했습니다: {type(exc).__name__}: {str(exc)[:300]}",
                "traceback": _tb.format_exc()[-2000:],
            }), 500

        return jsonify(
            {
                "ok": True,
                "message": "검증이 완료되었습니다.",
                "filename": validation["report_name"],
                "print_url": _print_report_url(report_path, copies=2),
                "sdf_filename": validation["sdf_name"],
                "summary": validation["summary"],
                "results": validation["results"],
                "insights": {
                    **validation["insights"],
                    "engineering_visualizations": engineering_visualizations,
                },
                "rules": validation.get("rules", {}),
                "stats": validation["stats"],
                "visualizations": visualizations,
                "tables": validation["tables"],
                "sdf_graph": sdf_graph,
                "report": validation["report"],
            }
        )

    @app.post("/api/export-xlsx")
    def export_xlsx():
        try:
            payload = request.get_json(force=True)
            tables = payload.get("tables") or {}
            report_name = payload.get("report_name") or "pipenet_result"
        except Exception:
            return jsonify({"ok": False, "message": "엑셀 내보내기 요청 형식이 올바르지 않습니다."}), 400

        wb = Workbook()
        wb.remove(wb.active)

        for key, meta in EXPORT_SCHEMA.items():
            ws = wb.create_sheet(title=meta["sheet"])
            rows = tables.get(key, [])

            ws.cell(row=1, column=1, value=f"{meta['sheet']} 결과 데이터")
            ws.cell(row=2, column=1, value="빨강=기준 위반, 노랑=확인 필요, 파랑=공학 후보, 초록=경제성 후보")

            headers = [label for _, label in meta["columns"]] + ["기준위반", "확인필요", "공학후보", "경제후보"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=3, column=col, value=header)

            for idx, row in enumerate(rows, start=4):
                values = [row.get(k, "") for k, _ in meta["columns"]]
                values += [
                    "Y" if row.get("highlight") else "N",
                    "Y" if row.get("warn") else "N",
                    "Y" if row.get("engineering_flag") else "N",
                    "Y" if row.get("economy_flag") else "N",
                ]
                for col, v in enumerate(values, start=1):
                    ws.cell(row=idx, column=col, value=v)

            _apply_sheet_style(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = Path(report_name).stem
        out_name = f"{safe_name}_결과테이블_{timestamp}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/remote30-workbench")
    def remote30_workbench():
        response = make_response(render_template("remote30_workbench.html"))
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    @app.get("/remote30-workbench-gnn")
    def remote30_workbench_gnn():
        """Remote 30 워크벤치 GNN 버전 - DXF→SDF ML 파이프라인 charter 기반"""
        response = make_response(render_template("remote30_workbench_gnn.html"))
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    @app.post("/api/remote30/extract")
    def remote30_extract():
        # 1) dxf_token 우선 - inspect 단계에서 저장된 파일을 재사용 (재업로드 불필요)
        dxf_token = request.form.get("dxf_token", "").strip()
        dxf_path = None
        if dxf_token:
            safe_token = sanitize_upload_name(dxf_token)
            if safe_token and safe_token == dxf_token:
                candidate = UPLOAD_DIR / safe_token
                if candidate.exists() and candidate.suffix.lower() == ".dxf":
                    dxf_path = candidate
        if dxf_path is None:
            try:
                dxf_path = _save_upload("dxf_file", {".dxf", ".dwg"}, required=True)
            except ValueError as exc:
                return jsonify({"ok": False, "message": str(exc)}), 400

        auto_detect = str(request.form.get("auto_detect_alarm", "true")).lower() in {"1", "true", "yes", "on"}
        alarm_xy = None
        if not auto_detect:
            ax = request.form.get("alarm_x", "").strip()
            ay = request.form.get("alarm_y", "").strip()
            if ax == "" or ay == "":
                return jsonify({"ok": False, "message": "수동 모드에서는 알람밸브 X, Y 좌표가 모두 필요합니다."}), 400
            try:
                alarm_xy = (float(ax), float(ay))
            except ValueError:
                return jsonify({"ok": False, "message": "알람밸브 좌표는 숫자여야 합니다."}), 400

        overrides = {}
        for key in ("pipe_layer_keywords", "head_layer_keywords", "text_layer_keywords", "arch_layer_keywords", "alarm_valve_keywords", "exclude_layer_keywords"):
            raw = request.form.get(key, "").strip()
            if raw:
                overrides[key] = [s.strip() for s in raw.split(",") if s.strip()]
        for key in ("snap_tol", "graph_closure_tol", "head_to_pipe_tol", "diameter_text_search_radius", "cad_unit_to_m", "c_factor",
                    "elevation_alarm_m", "elevation_head_m", "k_factor", "design_flow_per_head_lpm", "fallback_dia_mm"):
            raw = request.form.get(key, "").strip()
            if raw:
                try:
                    overrides[key] = float(raw)
                except ValueError:
                    return jsonify({"ok": False, "message": f"`{key}` 는 숫자여야 합니다."}), 400
        raw_count = request.form.get("remote_head_count", "").strip()
        if raw_count:
            try:
                overrides["remote_head_count"] = int(raw_count)
            except ValueError:
                return jsonify({"ok": False, "message": "`remote_head_count` 는 정수여야 합니다."}), 400
        remote_mode = request.form.get("remote_mode", "").strip().lower()
        if remote_mode in {"length", "hydraulic"}:
        # r30_inspect 도메인 라우트 → routes/r30_inspect.py (register 로 등록)
            overrides["remote_mode"] = remote_mode
        emit_sdf_raw = request.form.get("emit_sdf", "").strip().lower()
        if emit_sdf_raw:
            overrides["emit_sdf"] = emit_sdf_raw in {"1", "true", "yes", "on"}
        emit_csv_raw = request.form.get("emit_csv", "").strip().lower()
        if emit_csv_raw:
            overrides["emit_csv"] = emit_csv_raw in {"1", "true", "yes", "on"}
        # zone_bbox: "x_min,y_min,x_max,y_max" 4-tuple
        zone_raw = request.form.get("zone_bbox", "").strip()
        if zone_raw:
            try:
                parts = [float(x) for x in zone_raw.split(",")]
                if len(parts) == 4:
                    x_min, y_min, x_max, y_max = parts
                    if x_min < x_max and y_min < y_max:
                        overrides["zone_bbox"] = (x_min, y_min, x_max, y_max)
            except ValueError:
                return jsonify({"ok": False, "message": "zone_bbox 는 'x_min,y_min,x_max,y_max' 형식의 숫자여야 합니다."}), 400

        try:
            from sprinkler_remote30_extractor import run_remote30_extraction
        except ImportError as exc:
            return jsonify({"ok": False, "message": f"Remote30 모듈을 불러오지 못했습니다: {exc}"}), 500

        # 워크벤치에서 사용자가 확정한 헤드 / 추가한 배관 (JSON 배열)
        override_heads = None
        override_heads_raw = request.form.get("override_heads", "").strip()
        if override_heads_raw:
            try:
                override_heads = json.loads(override_heads_raw)
                if not isinstance(override_heads, list):
                    return jsonify({"ok": False, "message": "override_heads 는 배열이어야 합니다."}), 400
            except json.JSONDecodeError as exc:
                return jsonify({"ok": False, "message": f"override_heads JSON 파싱 실패: {exc}"}), 400
        override_pipes = None
        override_pipes_raw = request.form.get("override_pipes", "").strip()
        if override_pipes_raw:
            try:
                override_pipes = json.loads(override_pipes_raw)
                if not isinstance(override_pipes, list):
                    return jsonify({"ok": False, "message": "override_pipes 는 배열이어야 합니다."}), 400
            except json.JSONDecodeError as exc:
                return jsonify({"ok": False, "message": f"override_pipes JSON 파싱 실패: {exc}"}), 400

        try:
            result = run_remote30_extraction(
                dxf_path=dxf_path,
                alarm_xy=alarm_xy,
                out_dir=REMOTE30_OUTPUT_DIR,
                overrides=overrides or None,
                override_heads=override_heads,
                override_pipes=override_pipes,
            )
        except Exception as exc:
            return jsonify({"ok": False, "message": f"Remote30 추출 중 오류: {exc}"}), 500

        payload = {
            "ok": True,
            "run_id": result["run_id"],
            "alarm_xy": result["alarm_xy"],
            "alarm_node_xy": result.get("alarm_node_xy"),
            "alarm_source": result["alarm_source"],
            "remote_mode": result.get("remote_mode"),
            "counts": result["counts"],
            "summary": result["summary"],
            "warnings": result["warnings"],
            "selected_heads_xy": result.get("selected_heads_xy", []),
            "path_edges_xy": result.get("path_edges_xy", []),
            "sdf_tables": result.get("sdf_tables"),
            "png_url": f"/api/remote30/result/{result['run_id']}/png" if result.get("png_path") else None,
            "xlsx_url": f"/api/remote30/result/{result['run_id']}/xlsx" if result.get("xlsx_path") else None,
            "sdf_url": f"/api/remote30/result/{result['run_id']}/sdf" if result.get("sdf_path") else None,
            "csv_url": f"/api/remote30/result/{result['run_id']}/csv_zip" if result.get("csv_paths") else None,
        }
        return jsonify(payload)

    @app.post("/api/remote30/ml-detect")
    def remote30_ml_detect():
        """DXF → YOLO 헤드 검출. Layer 기반 결과와 비교용으로 워크벤치 캔버스에 표시.
        Input: dxf_token 또는 dxf_file (multipart)
        Output: { ok, ml_heads: [{x, y, conf}], counts: {detected}, source }
        """
        # 1) DXF 경로 확보 (토큰 우선)
        dxf_token = request.form.get("dxf_token", "").strip()
        dxf_path = None
        if dxf_token:
            safe_token = sanitize_upload_name(dxf_token)
            if safe_token and safe_token == dxf_token:
                cand = UPLOAD_DIR / safe_token
                if cand.exists() and cand.suffix.lower() == ".dxf":
                    dxf_path = cand
        if dxf_path is None:
            try:
                dxf_path = _save_upload("dxf_file", {".dxf", ".dwg"}, required=True)
            except ValueError as exc:
                return jsonify({"ok": False, "message": str(exc)}), 400

        # 2) DXF parsing → head_detector 호환 entity 포맷
        try:
            import ezdxf
            from sprinkler_remote30_extractor import Remote30Settings, layer_match
        except ImportError as exc:
            return jsonify({"ok": False, "message": f"의존성 누락: {exc}"}), 500

        try:
            doc = ezdxf.readfile(str(dxf_path))
            msp = doc.modelspace()
        except Exception as exc:
            return jsonify({"ok": False, "message": f"DXF 파싱 실패: {exc}"}), 500

        settings = Remote30Settings()
        entities = []
        bounds = [float("inf"), float("inf"), float("-inf"), float("-inf")]
        visible_layers = set()

        def _upd(x, y):
            if x < bounds[0]: bounds[0] = x
            if y < bounds[1]: bounds[1] = y
            if x > bounds[2]: bounds[2] = x
            if y > bounds[3]: bounds[3] = y

        for e in msp:
            layer = e.dxf.layer if hasattr(e.dxf, "layer") else ""
            if layer_match(layer, settings.arch_layer_keywords):
                continue
            if layer_match(layer, settings.exclude_layer_keywords):
                continue
            etype = e.dxftype()
            try:
                if etype == "LINE":
                    x1, y1 = float(e.dxf.start.x), float(e.dxf.start.y)
                    x2, y2 = float(e.dxf.end.x), float(e.dxf.end.y)
                    entities.append({"type": "LINE", "layer": layer, "start": {"x": x1, "y": y1}, "end": {"x": x2, "y": y2}})
                    _upd(x1, y1); _upd(x2, y2)
                    visible_layers.add(layer)
                elif etype == "LWPOLYLINE":
                    pts = [{"x": float(p[0]), "y": float(p[1])} for p in e.get_points()]
                    if pts:
                        for p in pts: _upd(p["x"], p["y"])
                        entities.append({"type": "LWPOLYLINE", "layer": layer, "points": pts, "closed": bool(e.closed) if hasattr(e, "closed") else False})
                        visible_layers.add(layer)
                elif etype == "ARC":
                    cx, cy = float(e.dxf.center.x), float(e.dxf.center.y)
                    r = float(e.dxf.radius)
                    entities.append({
                        "type": "ARC", "layer": layer,
                        "center": {"x": cx, "y": cy}, "radius": r,
                        "startAngle": float(e.dxf.start_angle), "endAngle": float(e.dxf.end_angle),
                    })
                    _upd(cx - r, cy - r); _upd(cx + r, cy + r)
                    visible_layers.add(layer)
                elif etype == "CIRCLE":
                    cx, cy = float(e.dxf.center.x), float(e.dxf.center.y)
                    r = float(e.dxf.radius)
                    entities.append({"type": "CIRCLE", "layer": layer, "center": {"x": cx, "y": cy}, "radius": r})
                    _upd(cx - r, cy - r); _upd(cx + r, cy + r)
                    visible_layers.add(layer)
            except Exception:
                continue

        if bounds[0] == float("inf"):
            return jsonify({"ok": False, "message": "추출할 entity 가 없습니다."}), 400

        rect = {"minX": bounds[0], "minY": bounds[1], "maxX": bounds[2], "maxY": bounds[3]}
        counts_meta = {"entities_rendered": len(entities)}

        # 사용자 지정 검출 범위 (zone_bbox) — 있으면 entity + bounds 그 안으로 제한
        zone_raw = request.form.get("zone_bbox", "").strip()
        if zone_raw:
            try:
                parts = [float(x) for x in zone_raw.split(",")]
                if len(parts) == 4:
                    zx_min, zy_min, zx_max, zy_max = parts
                    if zx_min < zx_max and zy_min < zy_max:
                        def _in_zone(x, y):
                            return zx_min <= x <= zx_max and zy_min <= y <= zy_max
                        def _ent_in_zone(e):
                            t = e.get("type")
                            if t == "LINE":
                                return _in_zone(e["start"]["x"], e["start"]["y"]) or _in_zone(e["end"]["x"], e["end"]["y"])
                            if t == "LWPOLYLINE":
                                return any(_in_zone(p["x"], p["y"]) for p in e.get("points", []))
                            if t in ("CIRCLE", "ARC"):
                                return _in_zone(e["center"]["x"], e["center"]["y"])
                            return False
                        entities = [e for e in entities if _ent_in_zone(e)]
                        rect = {"minX": zx_min, "minY": zy_min, "maxX": zx_max, "maxY": zy_max}
                        if not entities:
                            return jsonify({"ok": False, "message": "검출 범위 안에 entity 가 없습니다. zone 을 더 크게 잡아주세요."}), 400
                        counts_meta["zone_applied"] = [zx_min, zy_min, zx_max, zy_max]
            except ValueError:
                pass

        # 타일 옵션
        try:
            tile_grid = int(request.form.get("tile_grid", "2") or "2")
        except ValueError:
            tile_grid = 2
        try:
            tile_px = int(request.form.get("tile_px", "1280") or "1280")
        except ValueError:
            tile_px = 1280
        try:
            conf_thr = float(request.form.get("conf", "0.18") or "0.18")
        except ValueError:
            conf_thr = 0.18

        # 학습된 sprinkler_yolo 모델 우선 사용. 없으면 triangle_head_yolo 로 fallback.
        try:
            from remote30_ml import resolve_sprinkler_model_path
            model_path = resolve_sprinkler_model_path()
            if model_path is None:
                return jsonify({"ok": False, "message": "YOLO 모델 가중치를 찾을 수 없습니다. models/sprinkler_yolo/weights/best.pt 또는 triangle_head_yolo 확인."}), 500
        except Exception as exc:
            return jsonify({"ok": False, "message": f"모델 경로 결정 오류: {exc}"}), 500

        # 검출 방식 선택: yolo | color | layer | layer_yolo
        method = (request.form.get("method") or "color").lower()

        ml_heads = []
        ml_alarms = []
        counts_meta["method"] = method
        counts_meta["entities_rendered"] = len(entities)  # zone 필터 후 재계산

        try:
            from remote30_ml import detect_heads_with_tiles, detect_by_color_on_dxf, detect_heads_by_layer_insert
            from sprinkler_remote30_extractor import layer_match
            import ezdxf

            if method == "yolo":
                # sprinkler_yolo 모델 클래스: 0 head_yellow_circle, 1 head_red_triangle,
                # 2 head_red_dot, 3 alarm_valve. triangle_head_yolo (fallback) 은 단일 class 0.
        # r30_combined 도메인 라우트 → routes/r30_combined.py (register 로 등록)
                sprinkler_class_names = ["head_yellow_circle", "head_red_triangle", "head_red_dot", "alarm_valve"]
                tile_result = detect_heads_with_tiles(
                    entities=entities, rect=rect, visible_layers=visible_layers,
                    model_path=model_path,
                    tile_grid=tile_grid, tile_px=tile_px, conf=conf_thr,
                    class_names=sprinkler_class_names,
                )
                for box in tile_result["boxes"]:
                    cls = box.get("cls", 0)
                    cx = (box["minX"] + box["maxX"]) / 2.0
                    cy = (box["minY"] + box["maxY"]) / 2.0
                    obj = {
                        "x": cx, "y": cy,
                        "cls": cls,
                        "cls_name": sprinkler_class_names[cls] if 0 <= cls < len(sprinkler_class_names) else f"cls{cls}",
                        "conf": box.get("conf"),
                        "bbox": [box["minX"], box["minY"], box["maxX"], box["maxY"]],
                    }
                    if cls == 3:
                        ml_alarms.append(obj)
                    else:
                        ml_heads.append(obj)
                counts_meta.update({"tiles": tile_result["tiles"], "raw": tile_result["raw_detections"], "model_path": str(model_path)})

            elif method == "color":
                # DXF 의 헤드 layer 를 컬러로 렌더 → HSV 마스크 → contour 검출. layer 분류 의존.
                color_result = detect_by_color_on_dxf(
                    entities=entities, rect=rect, visible_layers=visible_layers,
                    tile_grid=tile_grid, tile_px=max(tile_px, 1600),
                )
                for box in color_result["boxes"]:
                    ml_heads.append({
                        "x": (box["minX"] + box["maxX"]) / 2.0,
                        "y": (box["minY"] + box["maxY"]) / 2.0,
                        "bbox": [box["minX"], box["minY"], box["maxX"], box["maxY"]],
                    })
                counts_meta.update({"tiles": color_result["tiles"], "raw": color_result["raw_detections"]})

            elif method == "layer":
                # DXF ground truth — HEAD layer 의 INSERT/CIRCLE 직접 추출. 가장 정확.
                # 위에서 이미 로드한 msp 재사용 (대형 도면 재파싱 30초+ 회피)
                settings2 = Remote30Settings()
                heads_layer = detect_heads_by_layer_insert(msp=msp, settings=settings2, layer_match_fn=layer_match)
                # zone 적용
                if "zone_applied" in counts_meta:
                    zx_min, zy_min, zx_max, zy_max = counts_meta["zone_applied"]
                    heads_layer = [h for h in heads_layer if zx_min <= h["x"] <= zx_max and zy_min <= h["y"] <= zy_max]
                for h in heads_layer:
                    ml_heads.append({"x": h["x"], "y": h["y"], "source": h["source"], "origin": "layer"})

            elif method == "layer_yolo":
                # Layer 먼저 (DXF ground truth) → YOLO 로 layer 가 놓친 추가 후보 보강
                # 위에서 이미 로드한 msp 재사용 (대형 도면 재파싱 30초+ 회피)
                settings2 = Remote30Settings()
                heads_layer = detect_heads_by_layer_insert(msp=msp, settings=settings2, layer_match_fn=layer_match)
                # zone 적용
                if "zone_applied" in counts_meta:
                    zx_min, zy_min, zx_max, zy_max = counts_meta["zone_applied"]
                    heads_layer = [h for h in heads_layer if zx_min <= h["x"] <= zx_max and zy_min <= h["y"] <= zy_max]
                layer_pts = [(h["x"], h["y"]) for h in heads_layer]
                for h in heads_layer:
                    ml_heads.append({"x": h["x"], "y": h["y"], "source": h["source"], "origin": "layer"})

                # YOLO 보강
                sprinkler_class_names = ["head_yellow_circle", "head_red_triangle", "head_red_dot", "alarm_valve"]
                try:
                    tile_result = detect_heads_with_tiles(
                        entities=entities, rect=rect, visible_layers=visible_layers,
                        model_path=model_path,
                        tile_grid=tile_grid, tile_px=tile_px, conf=conf_thr,
                        class_names=sprinkler_class_names,
                    )
                except Exception as exc:
                    tile_result = {"boxes": [], "tiles": 0, "raw_detections": 0, "image_count": 0}
                    counts_meta["yolo_error"] = str(exc)

                # 중복 제거 거리 (CAD 단위. 도면 단위 mm 기준 1500mm = 1.5m. 더 보수적으로 500.)
                dedup_radius = float(request.form.get("dedup_radius") or "1500")
                dedup_sq = dedup_radius ** 2
                yolo_only = 0
                for box in tile_result.get("boxes", []):
                    cls = box.get("cls", 0)
                    cx = (box["minX"] + box["maxX"]) / 2.0
                    cy = (box["minY"] + box["maxY"]) / 2.0
                    # 알람밸브는 별도 카운트
                    if cls == 3:
                        ml_alarms.append({
                            "x": cx, "y": cy, "cls": 3, "cls_name": "alarm_valve",
                            "conf": box.get("conf"), "origin": "yolo",
                        })
                        continue
                    # layer 결과와 중복 체크
                    is_dup = False
                    for (lx, ly) in layer_pts:
                        dx = cx - lx; dy = cy - ly
                        if dx*dx + dy*dy <= dedup_sq:
                            is_dup = True; break
                    if not is_dup:
                        ml_heads.append({
                            "x": cx, "y": cy,
                            "cls": cls,
                            "cls_name": sprinkler_class_names[cls] if 0 <= cls < len(sprinkler_class_names) else f"cls{cls}",
                            "conf": box.get("conf"),
                            "origin": "yolo_only",
                        })
                        yolo_only += 1

                counts_meta.update({
                    "layer_count": len(heads_layer),
                    "yolo_total": len(tile_result.get("boxes", [])),
                    "yolo_only_after_dedup": yolo_only,
                    "dedup_radius": dedup_radius,
                    "model_path": str(model_path),
                })

            else:
                return jsonify({"ok": False, "message": f"method '{method}' 지원 안 함. layer|color|yolo|layer_yolo 중 하나."}), 400

        except Exception as exc:
            return jsonify({"ok": False, "message": f"{method} 검출 오류: {exc}"}), 500

        return jsonify({
            "ok": True,
            "ml_heads": ml_heads,
            "ml_alarms": ml_alarms,
            "counts": {
                "detected": len(ml_heads),
                "alarm_detected": len(ml_alarms),
                **counts_meta,
            },
            "tile_config": {"tile_grid": tile_grid, "tile_px": tile_px},
            "source": method,
        })

    @app.post("/api/remote30/auto_process")
    def remote30_auto_process():
        """End-to-end 자동 파이프라인:
           DXF 업로드 → layer 기반 헤드/배관/관경 자동 분류
                     → (필요시 자동 zone 추천)
                     → 자동 누락 헤드 연결 (extractor 의 closure)
                     → hydraulic Remote 30 추출
                     → PNG + Excel + SDF + CSV 4종 모두 출력
        """
        try:
            dxf_path = _save_upload("dxf_file", {".dxf", ".dwg"}, required=True)
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 400

        try:
            from sprinkler_remote30_extractor import run_remote30_extraction
        except ImportError as exc:
            return jsonify({"ok": False, "message": f"extractor import 실패: {exc}"}), 500

        # 기본 설정: hydraulic remote + SDF + CSV 모두 생성
        overrides = {
            "remote_mode": "hydraulic",
            "emit_sdf": True,
            "emit_csv": True,
            "elevation_alarm_m": float(request.form.get("elevation_alarm_m") or "1.0"),
            "elevation_head_m":  float(request.form.get("elevation_head_m")  or "2.8"),
            "k_factor":          float(request.form.get("k_factor")          or "80"),
            "design_flow_per_head_lpm": float(request.form.get("design_flow_per_head_lpm") or "80"),
            "remote_head_count": int(float(request.form.get("remote_head_count") or "30")),
            # 자연낙차 가정 (옥상 수원). 답안지 RV03_NEW 기준 약 137m. 0 = 비활성
            "natural_fall_height_m": float(request.form.get("natural_fall_height_m") or "0"),
        }
        # 사용자 알람밸브 좌표 (있으면)
        alarm_xy = None
        ax = (request.form.get("alarm_x") or "").strip()
        ay = (request.form.get("alarm_y") or "").strip()
        if ax and ay:
            try:
                alarm_xy = (float(ax), float(ay))
            except ValueError:
                alarm_xy = None
        # zone (옵션)
        zone_raw = (request.form.get("zone_bbox") or "").strip()
        if zone_raw:
            try:
                parts = [float(x) for x in zone_raw.split(",")]
                if len(parts) == 4 and parts[0] < parts[2] and parts[1] < parts[3]:
                    overrides["zone_bbox"] = tuple(parts)
            except ValueError:
                pass

        try:
            result = run_remote30_extraction(
                dxf_path=dxf_path,
                alarm_xy=alarm_xy,
                out_dir=REMOTE30_OUTPUT_DIR,
                overrides=overrides,
            )
        except Exception as exc:
            return jsonify({"ok": False, "message": f"자동 처리 오류: {exc}"}), 500

        payload = {
            "ok": True,
            "dxf_filename": dxf_path.name,
            "run_id": result["run_id"],
            "alarm_xy": result["alarm_xy"],
            "alarm_source": result["alarm_source"],
            "remote_mode": result.get("remote_mode"),
            "counts": result["counts"],
            "summary": result["summary"],
            "warnings": result["warnings"],
            "selected_heads_xy": result.get("selected_heads_xy", []),
            "path_edges_xy": result.get("path_edges_xy", []),
            "sdf_tables": result.get("sdf_tables"),
            "png_url": f"/api/remote30/result/{result['run_id']}/png" if result.get("png_path") else None,
            "xlsx_url": f"/api/remote30/result/{result['run_id']}/xlsx" if result.get("xlsx_path") else None,
            "sdf_url": f"/api/remote30/result/{result['run_id']}/sdf" if result.get("sdf_path") else None,
            "csv_url": f"/api/remote30/result/{result['run_id']}/csv_zip" if result.get("csv_paths") else None,
        }
        return jsonify(payload)

    @app.post("/api/remote30/sdf-from-tables")
    def remote30_sdf_from_tables():
        """편집된 PIPENET tables(JSON) → SDF XML 응답. 인라인 편집된 결과를 즉시 다운로드 가능."""
        try:
            payload = request.get_json(force=True)
        except Exception:
            return jsonify({"ok": False, "message": "JSON body 가 필요합니다."}), 400

        tables = payload.get("tables") if isinstance(payload, dict) else None
        if not isinstance(tables, dict):
            return jsonify({"ok": False, "message": "tables 객체가 필요합니다."}), 400
        for key in ("nodes", "pipes", "nozzles"):
            if not isinstance(tables.get(key), list):
                return jsonify({"ok": False, "message": f"tables.{key} 배열이 필요합니다."}), 400
        tables.setdefault("valves", [])

        try:
            from sprinkler_remote30_extractor import build_sdf_xml, Remote30Settings
        except ImportError as exc:
            return jsonify({"ok": False, "message": f"모듈 import 실패: {exc}"}), 500

        settings = Remote30Settings()
        overrides = payload.get("settings") if isinstance(payload, dict) else None
        if isinstance(overrides, dict):
            for k, v in overrides.items():
                if hasattr(settings, k) and v is not None:
                    try:
                        setattr(settings, k, v)
                    except Exception:
                        pass
        title = str(payload.get("title", "Remote 30 Auto-Extracted"))[:80]

        try:
            xml_text = build_sdf_xml(tables, settings, title=title)
        except Exception as exc:
            return jsonify({"ok": False, "message": f"SDF 생성 오류: {exc}"}), 500

        # ASCII 파일명 정책: PIPENET 은 SDF/SLF 경로·파일명을 CP949(ANSI)로 읽으므로
        # UTF-8 한글 파일명은 깨진다. 사용자 지정 이름에서 비-ASCII 를 제거(secure_filename)
        # 하고, 남는 stem 이 없으면 기본값으로 폴백한다. (.sdf 확장자 보장)
        _safe = secure_filename(str(payload.get("filename") or ""))
        _stem = Path(_safe).stem
        if not _stem:
            _stem = "remote30_edited"
        download_name = f"{_stem[:76]}.sdf"
        return send_file(
            BytesIO(xml_text.encode("utf-8")),
            mimetype="application/xml",
            as_attachment=True,
            download_name=download_name,
        )

    @app.get("/api/remote30/result/<run_id>/<kind>")
    def remote30_result(run_id: str, kind: str):
        safe_run_id = secure_filename(run_id)
        if not safe_run_id or safe_run_id != run_id:
            return "잘못된 run_id 입니다.", 400
        if kind == "csv_zip":
            # CSV 4개 파일을 zip 으로 묶어 반환
            csv_dir = REMOTE30_OUTPUT_DIR / f"remote30_{safe_run_id}_csv"
            if not csv_dir.exists():
                return "CSV 결과 폴더를 찾을 수 없습니다.", 404
            import zipfile
            buf = BytesIO()
            with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for p in sorted(csv_dir.glob("*.csv")):
                    zf.write(p, arcname=p.name)
            buf.seek(0)
            return send_file(
                buf,
                mimetype="application/zip",
                as_attachment=True,
                download_name=f"remote30_{safe_run_id}_csv.zip",
            )

        suffix = {"png": ".png", "xlsx": ".xlsx", "sdf": ".sdf"}.get(kind)
        if suffix is None:
            return "지원하지 않는 결과 종류입니다.", 400
        target = REMOTE30_OUTPUT_DIR / f"remote30_{safe_run_id}{suffix}"
        if not target.exists():
            return "결과 파일을 찾을 수 없습니다.", 404
        mimetypes = {
            "png": "image/png",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "sdf": "application/xml",
        }
        return send_file(
            target,
            mimetype=mimetypes[kind],
            as_attachment=(kind != "png"),
            download_name=target.name,
        )

    @app.post("/api/remote30/export_cad")
    def remote30_export_cad():
        """레이어 정리 결과를 CAD 포맷(.dxf/.dwg)으로 재출력.

        워크벤치(건축 레이어 정리)에서 사용자가 체크로 남긴 레이어만 담아 원본 DXF 를
        필터링해 내보낸다. inspect 단계의 dxf_token 으로 원본을 재사용(재업로드 불필요).

        form:
            dxf_token       inspect 가 발급한 토큰 (필수)
            visible_layers  남길 레이어 이름 JSON 배열 (또는 콤마구분). 미지정 시 전체 유지.
            format          "dxf"(기본) | "dwg"  — dwg 는 ODA File Converter 필요
            filename        다운로드 파일명 stem (선택)
        """
        import ezdxf  # noqa: PLC0415

        dxf_token = request.form.get("dxf_token", "").strip()
        dxf_path = None
        if dxf_token:
            safe_token = sanitize_upload_name(dxf_token)
            if safe_token and safe_token == dxf_token:
                candidate = UPLOAD_DIR / safe_token
                if candidate.exists() and candidate.suffix.lower() == ".dxf":
                    dxf_path = candidate
        if dxf_path is None:
            try:
                dxf_path = _save_upload("dxf_file", {".dxf", ".dwg"}, required=True)
            except ValueError as exc:
                return jsonify({"ok": False, "message": str(exc)}), 400

        raw_layers = request.form.get("visible_layers", "").strip()
        kept: set[str] | None = None
        if raw_layers:
            try:
                parsed = json.loads(raw_layers)
                if isinstance(parsed, list):
                    kept = {str(x) for x in parsed}
            except (ValueError, TypeError):
                kept = {s.strip() for s in raw_layers.split(",") if s.strip()}

        fmt = (request.form.get("format", "dxf") or "dxf").strip().lower()
        if fmt not in {"dxf", "dwg"}:
            fmt = "dxf"

        try:
            doc = ezdxf.readfile(str(dxf_path))
        except Exception as exc:  # noqa: BLE001
            return jsonify({"ok": False, "message": f"원본 DXF 읽기 실패: {exc}"}), 400

        msp = doc.modelspace()
        removed = 0
        if kept is not None:
            to_delete = [e for e in msp if str(getattr(e.dxf, "layer", "0")) not in kept]
            for e in to_delete:
                try:
                    msp.delete_entity(e)
                    removed += 1
                except Exception:  # noqa: BLE001
                    pass
            # 더 이상 참조되지 않는 레이어 정의 정리 (실패는 무시 — 블록 참조 등)
            used = {str(getattr(e.dxf, "layer", "0")) for e in msp}
            for layer in list(doc.layers):
                name = str(layer.dxf.name)
                if name in ("0", "Defpoints") or name in used or name in (kept or set()):
                    continue
                try:
                    doc.layers.remove(name)
                except Exception:  # noqa: BLE001
                    pass

        stem = secure_filename(Path(request.form.get("filename", "") or dxf_path.stem).stem) or "cleaned"
        stem = f"{stem[:76]}_cleaned"
        out_path = REMOTE30_OUTPUT_DIR / f"{stem}.dxf"
        try:
            doc.saveas(str(out_path))
        except Exception as exc:  # noqa: BLE001
            return jsonify({"ok": False, "message": f"DXF 저장 실패: {exc}"}), 500

        if fmt == "dwg":
            from ezdxf.addons import odafc  # noqa: PLC0415
            exe = locate_oda_exe()
            if exe:
                try:
                    ezdxf.options.set("odafc-addon", "win_exec_path", exe)
                except Exception:  # noqa: BLE001
                    pass
            if not odafc.is_installed():
                return jsonify({
                    "ok": False,
                    "message": "DWG 출력에는 ODA File Converter(무료)가 필요합니다. "
                               "미설치 상태이니 DXF 로 내려받거나 ODA File Converter 설치 후 다시 시도하세요.",
                }), 400
            dwg_path = REMOTE30_OUTPUT_DIR / f"{stem}.dwg"
            try:
                odafc.export_dwg(doc, str(dwg_path), replace=True)
            except Exception as exc:  # noqa: BLE001
                return jsonify({"ok": False, "message": f"DWG 변환 실패: {exc}"}), 500
            resp = send_file(dwg_path, mimetype="image/vnd.dwg",
                             as_attachment=True, download_name=dwg_path.name)
        else:
            resp = send_file(out_path, mimetype="image/vnd.dxf",
                             as_attachment=True, download_name=out_path.name)
        resp.headers["X-Removed-Entities"] = str(removed)
        resp.headers["X-Kept-Layers"] = str(len(kept) if kept is not None else "all")
        return resp
