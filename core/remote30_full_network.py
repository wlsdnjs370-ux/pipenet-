"""Remote 30 전체 배관망 총괄 모듈 (10번 모듈 백엔드).

기존 ``remote30_prototype`` 의 헤드망 추출 로직(Stage A)을 재사용하고,
zone별 라이저 템플릿(Stage B), 라이저↔헤드망 stitch(Stage C),
PIPENET-native 후처리 + Pump-fan/Elastomeric-valve 직렬화(Stage D) 를
추가하여 펌프 → 감압밸브 → 알람밸브 → 헤드 30개 전 구간 SDF 를 생성한다.

흐름::

    OverallInputs (DXF + ZoneSpec + (선택) BuildingPressureProfile)
            │
            ├── Stage A — run_stages_0_2 (remote30_prototype 재사용)
            │                 → PipeTables (헤드망)
            │
            ├── Stage B — build_riser(zone_spec, profile)
            │                 → RiserTables (펌프/PRV/라이저)
            │
            ├── Stage C — stitch_riser_and_heads(riser, head_tables)
            │                 → CombinedTables
            │
            └── Stage D — emit_full_sdf(combined, out_path)
                              → 완성 SDF + 동봉 SLF

신규 attribute (vs prototype 의 PipeTables)::

    RiserTables.pumps   — <Pump-fan> 직렬화용 dict 리스트
    RiserTables.valves  — <Elastomeric-valve> 직렬화용 dict 리스트
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, Iterator

from remote30_constants import (DEFAULT_PUMP_LIBRARY_NAME, FX_DEFAULT_PROFILE,
                                FX_SPEC_PROFILES, REDUCER_SNAP_TO_TEE_MM,
                                RISER_PRV_APPROACH_M, RISER_ROOF_RUN_AFTER_DROP_M,
                                RISER_ROOF_RUN_TO_RISER_M, TEE_BRANCH_ABOVE_SLAB_M,
                                TEE_TO_ALARM_VALVE_RISE_M, TEE_TO_ALARM_VALVE_RUN_M,
                                TOP_FLOOR_EXTRA_HEIGHT_M)


# ────────────────────────────────────────────────────────────────────────────
# 상수
# ────────────────────────────────────────────────────────────────────────────

KGF_CM2_TO_PA = 98066.5    # 1 kg/cm² (kgf/cm²) → Pa
M_TO_PA = 9806.65          # 1 m 수두 → Pa (물 비중 1.0 기준)
ATM_PA = 101325.0          # 1 기압 (boundary condition)

# 호칭경 사다리 (KS/JIS A 계열, mm) — 오름차순. bore 정규화·한 치수 승급의 기준.
PIPE_BORE_LADDER_MM = [25, 32, 40, 50, 65, 80, 100, 125, 150, 200]


def _snap_bore_to_ladder(bore_mm: float) -> int:
    """임의 내경(mm)을 사다리의 가장 가까운(단, 미만이면 올림) 호칭경으로 스냅.

    수리 여유를 위해 사다리 값 미만은 바로 위 호칭으로 올린다 (내림 금지).
    사다리 최대(200)를 넘으면 200 으로 클램프.
    """
    b = float(bore_mm)
    for cand in PIPE_BORE_LADDER_MM:
        if b <= cand + 1e-6:
            return cand
    return PIPE_BORE_LADDER_MM[-1]


def _bump_one_size(bore_mm: int) -> int:
    """사다리에서 한 치수 위 호칭경 반환 (최대치는 유지)."""
    ladder = PIPE_BORE_LADDER_MM
    snapped = _snap_bore_to_ladder(bore_mm)
    idx = ladder.index(snapped)
    return ladder[min(idx + 1, len(ladder) - 1)]


def normalize_pipe_bores(
    nodes: list[dict],
    pipes: list[dict],
    *,
    bump_one_size: bool = False,
) -> int:
    """배관 내경을 트리 상류(입상관)→하류(가지) 단조 비증가로 정규화 (in-place).

    문제: 세그먼트별 최근접-관경 TEXT 매칭은 트리 위치를 모른 채 내경을 배정해,
    상류가 하류보다 얇아지는 '내경 꼬임'을 일으킨다.

    조치:
      1) source(io_node="Input")에서 hop 깊이를 BFS 로 계산해 각 파이프의
         상류(depth 작은 쪽)/하류(depth 큰 쪽) 끝점을 판별.
      2) 역방향 전파: 각 파이프 내경 = max(자기 내경, 자신이 먹이는 모든 하류
         파이프 내경) → 상류로 갈수록 굵어짐이 보장 (단조 비증가, outward).
         절대 얇게 줄이지 않음 (헤드 유량/압력 미달 방지와 일관).
      3) 사다리(PIPE_BORE_LADDER_MM)로 스냅.
      4) bump_one_size=True 면 전 구간 한 치수 승급 (build 시 1회만; rebuild 는 False).

    Returns:
        내경이 바뀐 파이프 수.
    """
    if not pipes:
        return 0

    # 인접: undirected (파이프는 양끝 label). depth 계산용.
    adj: dict[str, list[str]] = {}
    for p in pipes:
        a, b = str(p.get("in")), str(p.get("out"))
        adj.setdefault(a, []).append(b)
        adj.setdefault(b, []).append(a)

    # source 선정: io_node=="Input" 우선, 없으면 depth 계산 불가 → 첫 파이프 in.
    source = next((str(n.get("label")) for n in nodes
                   if str(n.get("io_node", "")).lower() == "input"), None)
    if source is None or source not in adj:
        source = str(pipes[0].get("in"))

    # BFS hop-depth (source=0).
    depth: dict[str, int] = {source: 0}
    queue = [source]
    while queue:
        cur = queue.pop(0)
        for nb in adj.get(cur, ()):
            if nb not in depth:
                depth[nb] = depth[cur] + 1
                queue.append(nb)

    def _pipe_depth(lbl: str) -> int:
        return depth.get(str(lbl), 10 ** 9)

    # 각 파이프의 상류(up)/하류(down) 끝점을 depth 로 판별.
    #   up = depth 작은 끝, down = depth 큰 끝.
    for p in pipes:
        a, b = str(p.get("in")), str(p.get("out"))
        if _pipe_depth(a) <= _pipe_depth(b):
            p["_up"], p["_down"] = a, b
        else:
            p["_up"], p["_down"] = b, a

    # down-node 를 상류 끝으로 갖는 파이프들 = 그 파이프가 먹이는 하류 파이프.
    feeds: dict[str, list[dict]] = {}
    for p in pipes:
        feeds.setdefault(p["_up"], []).append(p)

    # 파이프를 하류(깊은 곳)부터 처리하도록 down-depth 내림차순 정렬 후 전파.
    order = sorted(pipes, key=lambda p: _pipe_depth(p["_down"]), reverse=True)

    def _cur_bore(p: dict) -> float:
        try:
            return float(p.get("dia") or 0.0)
        except (TypeError, ValueError):
            return 0.0

    changed = 0
    for p in order:
        own = _cur_bore(p)
        child_max = 0.0
        for child in feeds.get(p["_down"], ()):  # p 의 하류 노드에서 뻗는 파이프들
            child_max = max(child_max, _cur_bore(child))
        raised = max(own, child_max)
        snapped = _snap_bore_to_ladder(raised) if raised > 0 else _snap_bore_to_ladder(own)
        if bump_one_size:
            snapped = _bump_one_size(snapped)
        new_val = int(snapped)
        old_val = int(round(own)) if own else None
        if new_val != old_val:
            changed += 1
        p["dia"] = new_val

    # 임시 키 제거.
    for p in pipes:
        p.pop("_up", None)
        p.pop("_down", None)

    return changed

# ────────────────────────────────────────────────────────────────────────────
# Zone 정의
# ────────────────────────────────────────────────────────────────────────────

class ZoneType(Enum):
    """처리 가능한 zone 타입 — 답안 SDF 의 라이저 구조에서 도출.

    압력분포표(예: 대명동 201동 PDF)의 각 행 → ZoneType 매핑::

        옥상 + 27F~49F          → HSP_PUMP        (펌프식 부스터)
        25F~26F                 → LSP_GRAVITY     (감압 없음, 자연낙차)
        2F~24F (24F 1차 PRV)    → LSP_1STAGE      (1차 감압)
        1F~B4 (1.5F 2차 PRV)    → LLSP_2STAGE     (1차 + 2차 감압)
    """
    HSP_PUMP = "hsp_pump"
    LSP_GRAVITY = "lsp_gravity"
    LSP_1STAGE = "lsp_1stage"
    LLSP_2STAGE = "llsp_2stage"


class MissingProjectInputError(ValueError):
    """수리 입력이 없어 라이저를 만들 수 없을 때. 메시지가 빠진 입력을 지목한다.

    ValueError 를 상속해 라우트의 기존 예외 처리에 그대로 걸린다.
    """


# ────────────────────────────────────────────────────────────────────────────
# 데이터 모델
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class FloorRow:
    """압력분포표의 한 행 (옥상~B4 까지 1층 1행).

    **표고 기준면 (작업지시서 4-2)** — 실측으로 확정된 것과 아닌 것:

    확정 — ``head_drop_m`` 은 수원에서 아래로 누적한 낙차이고, 행 사이의 증가분은
    **그 행 자신의 ``height_m``** 과 정확히 같다 (대명동 201동 배관압력흐름표
    57행 전량 검산). 첫 행만 예외로, 수원에서 옥상 측점까지의 낙차가 따로 주어진다
    (그 표에서는 3.1m — 옥상층 층고 6m 와 무관한 별개 값).

    **미확정 — 각 층에서 재는 지점이 층 바닥인지 알람밸브인지는 알 수 없다.**
    증가분이 층고와 같다는 사실은 "측점이 층마다 같은 상대위치에 있다" 까지만
    말해 주고, 그 상대위치가 무엇인지는 가리지 못한다. 같은 현장 수작업 모델의
    알람밸브 표고와 대조해도 차이가 계열마다 달라 하나의 기준면으로 안 모인다
    (자연낙차 LSP/MSP 6.35 일정, HSP 6.50~6.65, 지하 LOOP 3.70, LLSP 4.20~5.70).
    그래서 부호·오프셋을 맞추는 작업은 하지 않았다 — BLOCKED.md §26.

    현재 코드는 ``-head_drop_m`` 을 알람밸브 노드 표고로 그대로 쓴다. 이것은
    확정된 규약이 아니라 기존 동작이며, 기준면이 정해지면 바뀔 수 있다.
    """
    floor_label: str          # "옥상층", "49층", "1층", "B1층" 등
    height_m: float           # 층고 (m). 그 층 바닥~윗층 바닥.
    head_drop_m: float        # 누적 낙차 (m). 원점=수원, 층별 측점=미확정(위 참조).
    after_prv_m: float | None = None  # 감압 이후 수두 (m) — 감압 구간만
    note: str = ""            # 비고 ("자연낙차시작점", "1차 감압밸브 사용구간" 등)


def _to_int(value: Any, default: int) -> int:
    """폼/JSON 의 정수 필드 — 빈칸이나 숫자가 아니면 기본값."""
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _to_float(value: Any, default: float) -> float:
    """폼/JSON 의 실수 필드 — 빈칸이나 숫자가 아니면 기본값."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_opt_float(value: Any) -> float | None:
    """폼/JSON 의 선택 실수 필드 — 빈칸이나 숫자가 아니면 None (0 으로 때우지 않는다)."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _floor_row_from_mapping(raw: Any) -> FloorRow | None:
    """압력표 한 행(원시 매핑) → FloorRow. 못 읽으면 None.

    낙차압이 비어 있으면 0 으로 때우지 않고 행을 버린다 — 그 층의 표고를 물었을 때
    "0 이다" 라고 답하는 대신 "표에 없다" 로 걸리게 하려는 것.
    """
    if not isinstance(raw, dict):
        return None
    label = str(raw.get("floor_label") or "").strip()
    drop = raw.get("head_drop_m")
    if not label or drop in (None, ""):
        return None
    after = raw.get("after_prv_m")
    try:
        return FloorRow(
            floor_label=label,
            height_m=float(raw.get("height_m") or 0),
            head_drop_m=float(drop),
            after_prv_m=None if after in (None, "") else float(after),
            note=str(raw.get("note") or ""),
        )
    except (TypeError, ValueError):
        return None


@dataclass
class BuildingPressureProfile:
    """빌딩 전체 압력 흐름 표 — 옥상부터 최하층까지 1행씩.

    CSV/엑셀 업로드 또는 사용자 직접 입력 폼으로 생성. 없을 수도 있음 (선택적).
    """
    building_name: str = ""
    floors: list[FloorRow] = field(default_factory=list)

    def find_by_label(self, floor_label: str) -> FloorRow | None:
        for row in self.floors:
            if row.floor_label == floor_label:
                return row
        return None

    @classmethod
    def from_csv(cls, csv_path: Path, *, building_name: str = "") -> "BuildingPressureProfile":
        """CSV 파서 — 컬럼: floor_label, height_m, head_drop_m, after_prv_m, note.

        헤더 행은 한글/영문 모두 허용 (구분/층고/낙차압/감압이후/비고).
        낙차압을 어느 지점에서 재는지는 확정되지 않았다 — :class:`FloorRow` 참조.
        """
        rows: list[FloorRow] = []
        # 한글 컬럼명 → 영문 키 매핑 (PDF 표 헤더에 맞춤)
        KCOL = {
            "구분": "floor_label", "층": "floor_label",
            "층고": "height_m",
            "낙차압": "head_drop_m", "낙차": "head_drop_m",
            "감압이후": "after_prv_m", "감압후": "after_prv_m",
            "비고": "note",
        }
        with csv_path.open("r", encoding="utf-8-sig", newline="") as fp:
            reader = csv.DictReader(fp)
            for raw in reader:
                # 한글 헤더면 매핑 적용
                norm: dict[str, Any] = {}
                for k, v in raw.items():
                    key = KCOL.get((k or "").strip(), (k or "").strip())
                    norm[key] = (v or "").strip()
                row = _floor_row_from_mapping(norm)
                if row is not None:
                    rows.append(row)
        return cls(building_name=building_name, floors=rows)

    @classmethod
    def from_xlsx(cls, xlsx_path: Path, *, sheet: str | int = 0,
                  building_name: str = "") -> "BuildingPressureProfile":
        """엑셀(.xlsx) 파서 — 첫 시트 또는 지정 시트의 같은 컬럼 구조.

        낙차압을 어느 지점에서 재는지는 확정되지 않았다 — :class:`FloorRow` 참조.
        """
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        header_row = next(ws.iter_rows(values_only=True))
        KCOL = {
            "구분": "floor_label", "층": "floor_label",
            "층고": "height_m",
            "낙차압": "head_drop_m", "낙차": "head_drop_m",
            "감압이후": "after_prv_m", "감압후": "after_prv_m",
            "비고": "note",
        }
        col_idx = {KCOL.get((str(h) or "").strip(), (str(h) or "").strip()): i
                   for i, h in enumerate(header_row) if h is not None}
        rows: list[FloorRow] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            def _get(key, default=None):
                i = col_idx.get(key)
                return r[i] if i is not None and i < len(r) else default
            row = _floor_row_from_mapping({k: _get(k) for k in
                                           ("floor_label", "height_m", "head_drop_m",
                                            "after_prv_m", "note")})
            if row is not None:
                rows.append(row)
        wb.close()
        return cls(building_name=building_name, floors=rows)


@dataclass
class ZoneSpec:
    """처리 대상 zone 의 사양. 사용자 입력 (라디오 + 폼).

    필드 의미::

        zone_type             — ZoneType (HSP_PUMP / LSP_GRAVITY / LSP_1STAGE / LLSP_2STAGE)
        target_floor          — 추출 대상 층 ("16층" — Stage A 헤드망과 동일)
        prv1_target_pa        — 1차 PRV 출구압 (Pa). 자연낙차 + 감압 zone 에서.
        prv2_target_pa        — 2차 PRV 출구압 (Pa). LLSP_2STAGE 만.
        pump_library_name     — Library-pump 의 SLF Pump-definition 이름.
                                SLF 사본마다 펌프 이름이 다르므로 기본값이 실재하는지는
                                방출 시점에 대조한다 (DEFAULT_PUMP_LIBRARY_NAME 주석 참조).
        pump_count            — Pump-fan 개수 (HSP 보통 2개 = 1차+2차 부스터)
        pump_rated_q_lpm      — 정격유량 (L/min). 기본값 없음 — 모르면 None.
        pump_rated_h_m        — 정격양정 (m). 기본값 없음 — 모르면 None.

    정격유량/양정이 있으면 SLF 에 없는 펌프도 NFPC 3점 곡선으로 주입할 수 있다.
    없으면 주입도 못 하므로 미확정 경고로 올라간다 — 임의 곡선을 지어내지 않는다.

    압력분포표가 있으면 prv1/prv2 target 은 표에서 자동 도출 가능, 없으면 직접 입력.
    """
    zone_type: ZoneType
    target_floor: str = ""
    prv1_target_pa: float | None = None
    prv2_target_pa: float | None = None
    pump_library_name: str = DEFAULT_PUMP_LIBRARY_NAME
    pump_count: int = 2
    pump_rated_q_lpm: float | None = None
    pump_rated_h_m: float | None = None


class SourceTag(str, Enum):
    """값이 어디서 왔는지. 리포트 경고 블록의 판단 근거."""
    USER_CONFIRMED = "user_confirmed"      # 사람이 폼에 직접 넣고 확정
    DRAWING_ESTIMATED = "drawing_estimated"  # 도면/계통도에서 뽑아낸 추정치
    DEFAULT = "default"                    # 아무도 안 정해서 코드 기본값이 쓰인 것


# 경고 블록에 사람이 읽을 이름으로 찍는다. 여기 실린 필드만 미확정 추적 대상이다.
_CONTEXT_FIELD_LABELS: dict[str, str] = {
    "project_title": "프로젝트명",
    "zone_name": "존 이름",
    "building_name": "빌딩명",
    "floor_profile": "층별 압력분포표",
    "natural_fall_start_floor": "자연낙차 시작층",
    "machine_room_ceiling_m": "기계실 천장고",
    "roof_tank_water_level_m": "옥상수조 수위",
    "fx_profile_key": "신축배관 규격",
    "material_zones": "관종 구역",
    "ceiling_zones": "반자 구역",
    "roof_run_to_riser_m": "옥상 수평배관 길이 (수원→입상관)",
    "roof_run_after_drop_m": "옥상 수평배관 길이 (하강 후)",
    "roof_to_top_floor_drop_m": "옥탑→최상층 낙차",
    "tee_to_alarm_valve_m": "입상관 T분기→알람밸브 상승",
    "tee_branch_above_slab_m": "입상관 T분기 높이 (바닥 기준)",
    "top_floor_extra_height_m": "최상층 추가 층고 (단열재분)",
    "water_hammer_arrester": "수격방지기 유무",
}

# 받아서 서류에는 남기지만 아직 배관망을 바꾸지 않는 항목. 여기 적힌 필드를 채워도
# SDF 는 그대로다 — 채웠다는 이유로 "반영됐다" 고 읽히면 안 되므로 경고에 밝힌다.
_RECORD_ONLY_FIELDS: frozenset[str] = frozenset({
    "natural_fall_start_floor",   # 후보 계산·표시용. 라이저 분기는 zone_type 이 정한다.
    "machine_room_ceiling_m",     # 기계실 추출 경로(프로토타입) 전용 — 통합 산출에는 기계실이 없다.
    "roof_tank_water_level_m",
    "ceiling_zones",              # 상향/하향은 material_zones 의 구역 유형이 가른다.
    # 아래 셋은 표고 기준면이 확정돼야 표고에 반영할 수 있다 (BLOCKED.md §26).
    "tee_branch_above_slab_m",
    "top_floor_extra_height_m",
    "water_hammer_arrester",      # 부속 형상이 갈리지만 라이저에 그 부속이 아직 없다.
})


def _context_value_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (str, list, tuple, dict)):
        return len(value) == 0
    if isinstance(value, BuildingPressureProfile):
        return not value.floors
    return False


@dataclass
class ProjectContext:
    """한 과제의 수리계산 입력 일체 — 폼 딕셔너리를 대신하는 타입 있는 서류.

    ``sources`` 는 필드명 → SourceTag 문자열. 사용자가 실제로 채운 값만
    ``user_confirmed`` 가 되고 나머지는 ``default`` 로 남아 :meth:`unconfirmed`
    에 잡힌다. "값이 이렇다" 와 "아무도 안 정해서 이렇게 됐다" 를 구분하는 것이
    이 클래스의 존재 이유다.
    """
    zone_spec: ZoneSpec
    floor_profile: BuildingPressureProfile | None = None
    project_title: str = ""
    zone_name: str = ""
    building_name: str = ""
    natural_fall_start_floor: str | None = None
    machine_room_ceiling_m: float | None = None
    roof_tank_water_level_m: float | None = None
    fx_profile_key: str = FX_DEFAULT_PROFILE
    material_zones: list[dict] = field(default_factory=list)
    ceiling_zones: list[dict] = field(default_factory=list)
    # 라이저 형상 — 기본값은 대명동 201동 실측값이라 다른 현장에서는 전부 틀리다.
    # 확정 전에는 unconfirmed() 가 [미확정] 로 올린다 (remote30_constants 근거 주석).
    roof_run_to_riser_m: float = RISER_ROOF_RUN_TO_RISER_M
    roof_run_after_drop_m: float = RISER_ROOF_RUN_AFTER_DROP_M
    tee_to_alarm_valve_m: float = TEE_TO_ALARM_VALVE_RISE_M
    # 옥탑~최상층 낙차는 도면 규칙으로 못 구한다. None 이면 압력표 최상단 층고로 낸다.
    roof_to_top_floor_drop_m: float | None = None
    tee_branch_above_slab_m: float = TEE_BRANCH_ABOVE_SLAB_M
    top_floor_extra_height_m: float = TOP_FLOOR_EXTRA_HEIGHT_M
    water_hammer_arrester: bool = False
    sources: dict[str, str] = field(default_factory=dict)
    # 방출 시점에야 알 수 있는 미확정 — 입력이 아니라 산출물을 대조해 나온 것이라
    # 직렬화하지 않는다(to_dict 에 없음). 예: Pump-fan 이 SLF 에 없는 펌프를 참조.
    emit_findings: list[dict] = field(default_factory=list)

    # ── 출처 ────────────────────────────────────────────────────────────
    def tag(self, field_name: str) -> SourceTag:
        try:
            return SourceTag(self.sources.get(field_name, ""))
        except ValueError:
            return SourceTag.DEFAULT

    def unconfirmed(self) -> list[dict]:
        """사람이 확정하지 않은 항목 — 리포트 최상단 경고 블록의 재료."""
        out: list[dict] = []
        for name, label in _CONTEXT_FIELD_LABELS.items():
            tag = self.tag(name)
            if tag is SourceTag.USER_CONFIRMED:
                continue
            out.append({
                "field": name,
                "label": label,
                "tag": tag.value,
                "missing": _context_value_missing(getattr(self, name)),
                "record_only": name in _RECORD_ONLY_FIELDS,
            })
        return out + list(self.emit_findings)

    def note_emit_issue(self, field_name: str, label: str, state: str) -> None:
        """방출 대조에서 나온 미확정을 경고 블록에 올린다. 같은 항목은 한 번만."""
        item = {
            "field": field_name, "label": label, "state": state,
            "tag": SourceTag.DEFAULT.value, "missing": False, "record_only": False,
        }
        if item not in self.emit_findings:
            self.emit_findings.append(item)

    def warning_lines(self) -> list[str]:
        """`[미확정] <필드> — <한글이름> (<상태>)` 한 줄씩."""
        state_of = {
            SourceTag.DRAWING_ESTIMATED.value: "도면 추정",
            SourceTag.DEFAULT.value: "기본값",
        }
        lines = []
        for item in self.unconfirmed():
            state = item.get("state") or (
                "미입력" if item["missing"] else state_of.get(item["tag"], "기본값"))
            note = ", 기록 전용 — 현재 산출물 미반영" if item["record_only"] else ""
            lines.append(f"[미확정] {item['field']} — {item['label']} ({state}{note})")
        return lines

    # ── 파생값 ──────────────────────────────────────────────────────────
    def report_title(self) -> str:
        """리포트/SDF 제목. 사용자가 넣은 프로젝트명이 최우선."""
        for candidate in (self.project_title, self.zone_name):
            if candidate.strip():
                return candidate.strip()
        return (f"Remote 30 전체 — {self.zone_spec.zone_type.value} "
                f"{self.zone_spec.target_floor}").strip()

    def natural_fall_candidates(self) -> list[dict]:
        """자연낙차 시작층 후보 — 계산해서 보여만 주고 고르지는 않는다.

        규칙(지시서 T4): 낙차압이 헤드 최소 방수압(NFTC 2.2.1.11, 0.1 MPa)
        이상이 되는 최초의 층. 압력표가 없으면 후보도 없다 — 짐작하지 않는다.
        """
        if self.floor_profile is None or not self.floor_profile.floors:
            return []
        from nftc_rules import head_pressure_min_mpa
        required_m = head_pressure_min_mpa() * 1e6 / M_TO_PA
        return [{
            "floor_label": row.floor_label,
            "head_drop_m": row.head_drop_m,
            "required_m": round(required_m, 2),
            "ok": row.head_drop_m >= required_m,
        } for row in self.floor_profile.floors]

    def suggested_natural_fall_floor(self) -> str | None:
        """UI 가 커서만 올려둘 기본 후보. 자동 확정이 아니다."""
        for cand in self.natural_fall_candidates():
            if cand["ok"]:
                return cand["floor_label"]
        return None

    # ── 직렬화 (combined 경로의 선택 필드 project_context) ───────────────
    def to_dict(self) -> dict:
        return {
            "zone_type": self.zone_spec.zone_type.value,
            "target_floor": self.zone_spec.target_floor,
            "prv1_target_pa": self.zone_spec.prv1_target_pa,
            "prv2_target_pa": self.zone_spec.prv2_target_pa,
            "pump_library_name": self.zone_spec.pump_library_name,
            "pump_count": self.zone_spec.pump_count,
            "pump_rated_q_lpm": self.zone_spec.pump_rated_q_lpm,
            "pump_rated_h_m": self.zone_spec.pump_rated_h_m,
            "floor_profile": ([{
                "floor_label": r.floor_label, "height_m": r.height_m,
                "head_drop_m": r.head_drop_m, "after_prv_m": r.after_prv_m,
                "note": r.note,
            } for r in self.floor_profile.floors] if self.floor_profile else None),
            "project_title": self.project_title,
            "zone_name": self.zone_name,
            "building_name": self.building_name,
            "natural_fall_start_floor": self.natural_fall_start_floor,
            "machine_room_ceiling_m": self.machine_room_ceiling_m,
            "roof_tank_water_level_m": self.roof_tank_water_level_m,
            "fx_profile_key": self.fx_profile_key,
            "material_zones": list(self.material_zones),
            "ceiling_zones": list(self.ceiling_zones),
            "roof_run_to_riser_m": self.roof_run_to_riser_m,
            "roof_run_after_drop_m": self.roof_run_after_drop_m,
            "roof_to_top_floor_drop_m": self.roof_to_top_floor_drop_m,
            "tee_to_alarm_valve_m": self.tee_to_alarm_valve_m,
            "tee_branch_above_slab_m": self.tee_branch_above_slab_m,
            "top_floor_extra_height_m": self.top_floor_extra_height_m,
            "water_hammer_arrester": self.water_hammer_arrester,
            "sources": dict(self.sources),
        }

    @classmethod
    def from_dict(cls, data: dict) -> "ProjectContext":
        rows = data.get("floor_profile")
        profile = BuildingPressureProfile(
            building_name=str(data.get("building_name", "") or ""),
            floors=[row for row in map(_floor_row_from_mapping, rows) if row is not None],
        ) if isinstance(rows, list) and rows else None
        return cls(
            zone_spec=ZoneSpec(
                zone_type=ZoneType(data.get("zone_type", ZoneType.LSP_1STAGE.value)),
                target_floor=str(data.get("target_floor", "") or ""),
                prv1_target_pa=data.get("prv1_target_pa"),
                prv2_target_pa=data.get("prv2_target_pa"),
                pump_library_name=str(data.get("pump_library_name") or DEFAULT_PUMP_LIBRARY_NAME),
                pump_count=_to_int(data.get("pump_count"), 2),
                pump_rated_q_lpm=_to_opt_float(data.get("pump_rated_q_lpm")),
                pump_rated_h_m=_to_opt_float(data.get("pump_rated_h_m")),
            ),
            floor_profile=profile,
            project_title=str(data.get("project_title", "") or ""),
            zone_name=str(data.get("zone_name", "") or ""),
            building_name=str(data.get("building_name", "") or ""),
            natural_fall_start_floor=(data.get("natural_fall_start_floor") or None),
            machine_room_ceiling_m=data.get("machine_room_ceiling_m"),
            roof_tank_water_level_m=data.get("roof_tank_water_level_m"),
            fx_profile_key=str(data.get("fx_profile_key") or FX_DEFAULT_PROFILE),
            material_zones=list(data.get("material_zones") or []),
            ceiling_zones=list(data.get("ceiling_zones") or []),
            roof_run_to_riser_m=_to_float(data.get("roof_run_to_riser_m"),
                                          RISER_ROOF_RUN_TO_RISER_M),
            roof_run_after_drop_m=_to_float(data.get("roof_run_after_drop_m"),
                                            RISER_ROOF_RUN_AFTER_DROP_M),
            roof_to_top_floor_drop_m=_to_opt_float(data.get("roof_to_top_floor_drop_m")),
            tee_to_alarm_valve_m=_to_float(data.get("tee_to_alarm_valve_m"),
                                           TEE_TO_ALARM_VALVE_RISE_M),
            tee_branch_above_slab_m=_to_float(data.get("tee_branch_above_slab_m"),
                                              TEE_BRANCH_ABOVE_SLAB_M),
            top_floor_extra_height_m=_to_float(data.get("top_floor_extra_height_m"),
                                               TOP_FLOOR_EXTRA_HEIGHT_M),
            water_hammer_arrester=bool(data.get("water_hammer_arrester")),
            sources=dict(data.get("sources") or {}),
        )

    @classmethod
    def titled(cls, project_title: str, *, zone_type: ZoneType = ZoneType.LSP_1STAGE) -> "ProjectContext":
        """제목만 아는 호출자(구형 통합 경로)용 최소 컨텍스트.

        나머지 항목은 전부 미확정으로 남아 경고 블록에 그대로 드러난다.
        """
        return cls(zone_spec=ZoneSpec(zone_type=zone_type),
                   project_title=project_title,
                   sources={"project_title": SourceTag.USER_CONFIRMED.value})


@dataclass
class OverallInputs:
    """모듈 10 의 입력 일체."""
    dxf_path: Path
    zone_spec: ZoneSpec
    profile: BuildingPressureProfile | None = None
    alarm_xy: tuple[float, float] | None = None
    job_id: str = ""
    project_title: str = "Remote 30 전체 배관망 총괄"

    def to_context(self) -> ProjectContext:
        return ProjectContext(
            zone_spec=self.zone_spec, floor_profile=self.profile,
            project_title=self.project_title,
            building_name=(self.profile.building_name if self.profile else ""),
            sources={"project_title": SourceTag.USER_CONFIRMED.value},
        )


# ────────────────────────────────────────────────────────────────────────────
# RiserTables — Stage B 산출 / Stage C 입력
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class RiserTables:
    """라이저(펌프~AV) 구간의 PipeTables 확장.

    prototype.PipeTables 와 호환되는 nodes/pipes 형식 + pumps/valves 추가.
    각 dict 의 키는 emit_sdf 의 직렬화에 그대로 쓰일 수 있도록 PIPENET 속성명에 맞춤.
    """
    nodes: list[dict] = field(default_factory=list)
    pipes: list[dict] = field(default_factory=list)
    pumps: list[dict] = field(default_factory=list)   # Pump-fan 직렬화용
    valves: list[dict] = field(default_factory=list)  # Elastomeric-valve 직렬화용
    av_node_label: str = ""  # 라이저 끝점 = 헤드망 source 와 stitch 할 노드 라벨


# ────────────────────────────────────────────────────────────────────────────
# Stage A — 헤드망 추출 (remote30_prototype 재사용)
# ────────────────────────────────────────────────────────────────────────────

def run_stage_a(inputs: OverallInputs) -> Iterator[dict]:
    """Stage A — 평면도 DXF → 헤드 30개 + 배관망 추출 이벤트 스트림.

    호출 측(서버)은 마지막 ``stage2_complete`` 이벤트의 데이터를 보관 후
    사용자 헤드 편집 단계 → run_stages_3_5 → build_input_tables 까지 진행한다.
    """
    from remote30_prototype import run_stages_0_2
    yield from run_stages_0_2(
        dxf_path=inputs.dxf_path,
        job_id=inputs.job_id,
        alarm_xy=inputs.alarm_xy,
    )


# ────────────────────────────────────────────────────────────────────────────
# Stage B — zone별 라이저 템플릿
# ────────────────────────────────────────────────────────────────────────────
#
# 좌표 처리 방침:
# PIPENET SDF 의 Position x/y 는 isometric 시각 표시용일 뿐 수리계산엔 무관하다.
# emit_sdf 의 _xform 이 모든 노드 좌표를 bbox 중심 (0,0) + 약 3000 unit 으로 정규화하므로
# 라이저 노드 좌표의 절댓값은 의미가 없고 상대 배치(라이저 수직 진행, AV 가 적절한 위치)
# 만 유지하면 된다. 그러므로 빌딩 무관 logical 좌표계로 정의 — 모든 빌딩에서 동작.
#
# Logical 좌표 (mm 단위, 옥상 수원을 (0, 0) 으로):
#   Input(1)            (   0,    0)    옥상 수원
#   N2                  (-500, -100)    옥상 수평 분기
#   N3                  (-500, -300)    옥상 → 라이저 진입 (수직 강하 시작)
#   N4                  (-300, -1500)   라이저 중간 (수평 우회)
#   N7                  (-300, -2800)   PRV 진입 직전
#   N8 (PRV in)         (-200, -2900)
#   N89 (PRV out)       (-100, -2950)
#   N5 (AV 직전)         (-300, -3300)
#   N10 (AV ★)          ( 100, -3300)   헤드망 source 와 stitch
#   N87 (2차 PRV in)    ( -50, -3100)   LLSP_2STAGE 만
#   N88 (2차 PRV out)   (  50, -3150)
#   N100 (Pump Input)   ( 200,   100)   HSP_PUMP 만 (Input 1 보다 더 위)
#
# v1 구현 (대명동 201동 답안 SDF) 의 좌표는 git history 에서 확인 가능.
#
# 라이저 노드 라벨 컨벤션 (답안과 정합)::
#     1   : Input 노드 (옥상 수원, 1기압 boundary)
#     2,3,4,7  : 라이저 중간 노드 (옥상→하강→PRV 직전)
#     8   : 1차 PRV in
#     89  : 1차 PRV out
#     5   : AV 직전 노드
#     10  : AV 노드 (헤드망 source 와 stitch)
#     87  : 2차 PRV in (LLSP_2STAGE 만)
#     88  : 2차 PRV out (LLSP_2STAGE 만)
#     100 : Pump-fan 전 Input (HSP_PUMP 만 — Input 라벨이 1 → 100 으로 옮김)
#
# ────────────────────────────────────────────────────────────────────────────


def _node(label: str, elev: float, x: float, y: float, *,
          io_node: str = "No", pressure_pa: float | None = None) -> dict:
    """라이저 노드 dict — remote30_prototype.PipeTables.nodes 호환 형식."""
    d: dict = {
        "label": label, "elevation": elev, "io_node": io_node,
        "x": int(round(x)), "y": int(round(y)),
    }
    if pressure_pa is not None:
        d["pressure_pa"] = pressure_pa  # Input node 의 <Calculation-spec pressure="..."/>
    return d


def _pipe(label: str, in_lbl: str, out_lbl: str, bore_mm: int,
          length_m: float, rise_m: float = 0.0, c_factor: str = "120") -> dict:
    """라이저 파이프 dict — remote30_prototype.PipeTables.pipes 호환.

    PIPENET/K-solver 는 낙차가 관 길이보다 큰 관을 거부하므로 길이를 |낙차| 아래로
    내려보내지 않는다. 반올림도 이 관계를 깨지 않게 양쪽 모두 같은 자리로 맞춘다.
    """
    rise = round(rise_m, 2)
    return {
        "label": label, "in": in_lbl, "out": out_lbl, "type": "KSD 3507",
        "dia": bore_mm, "length": max(round(length_m, 2), abs(rise)), "elev": rise,
        "c": c_factor, "status": "Normal", "group": "Unset",
    }


def _pump_fan(label: str, in_lbl: str, out_lbl: str, *,
              library_pump: str, efficiency: int = 100, status: int = 1,
              rated_q: float | None = None, rated_h: float | None = None) -> dict:
    """Pump-fan dict — emit_full_sdf 가 <Pump-fan> 으로 직렬화.

    rated_q(L/min)/rated_h(m) 를 실으면 SLF 에 그 이름의 Pump-definition 이 없을 때
    _harden_slf_for_combined 가 NFPC 3점 곡선으로 만들어 주입한다. 없으면 주입하지
    않는다 — 곡선을 지어내면 PIPENET 이 양정을 스스로 고르는 것과 다를 바 없다.
    """
    fan = {
        "label": label, "in": in_lbl, "out": out_lbl,
        "efficiency": efficiency, "status": status,
        "library_pump": library_pump,
        "percentage_open": 1,
    }
    if rated_q and rated_h:
        fan["rated_q"] = float(rated_q)
        fan["rated_h"] = float(rated_h)
    return fan


def _pressure_valve(label: str, in_lbl: str, out_lbl: str, *,
                    target_pa: float, valve_type: str = "output") -> dict:
    """Elastomeric-valve dict — emit_full_sdf 가 <Elastomeric-valve> 로 직렬화."""
    return {
        "label": label, "in": in_lbl, "out": out_lbl,
        "target_value": float(target_pa), "type": valve_type,
    }


def _require_profile(profile: BuildingPressureProfile | None,
                     zone_label: str) -> BuildingPressureProfile:
    """압력표가 있어야만 라이저를 만든다 — 없으면 특정 현장 수치로 조용히 때우지 않는다."""
    if profile is None or not profile.floors:
        raise MissingProjectInputError(
            f"{zone_label} 라이저를 만들려면 층별 압력표가 필요합니다. "
            "계통도 DXF 자동 추출 · CSV/엑셀 업로드 · 직접 입력 중 하나로 채워주세요.")
    # 이 아래 표고 계산은 전부 "첫 행이 옥상" 을 전제로 한다. 표가 아래층부터
    # 적혀 있으면 라이저 전체가 뒤집히는데 결과만 봐서는 알 수 없으므로,
    # 누적 낙차가 위→아래로 커지는지 여기서 확인한다.
    for prev, cur in zip(profile.floors, profile.floors[1:]):
        if cur.head_drop_m < prev.head_drop_m:
            raise MissingProjectInputError(
                f"압력표가 옥상부터 아래로 정렬되어 있지 않습니다 — "
                f"'{prev.floor_label}'({prev.head_drop_m}m) 다음 행 "
                f"'{cur.floor_label}'({cur.head_drop_m}m) 의 누적 낙차가 더 작습니다.")
    return profile


def _elev_at_floor(profile: BuildingPressureProfile, floor_label: str) -> float:
    """압력표에서 층 라벨의 누적 낙차 (m) → SDF elevation (음수)."""
    row = profile.find_by_label(floor_label)
    if row is None:
        have = ", ".join(r.floor_label for r in profile.floors[:12])
        raise MissingProjectInputError(
            f"압력표에 '{floor_label}' 행이 없어 그 층의 낙차압을 알 수 없습니다. "
            f"표에 있는 층: {have}")
    return -float(row.head_drop_m)


def _elev_riser_top(profile: BuildingPressureProfile,
                    roof_drop_m: float | None = None) -> float:
    """라이저가 수원에서 내려오기 시작하는 표고.

    옥탑 수조에서 최상층까지의 낙차는 건물 고유값이라 도면에서 규칙으로 못 구한다.
    ``roof_drop_m`` 이 주어지면 그 값을 쓰고, 없으면 압력표 최상단 행(옥상)의 층고로
    대신한다 — 층고와 낙차가 같다는 보장은 없으니 어디까지나 대타다.
    """
    if roof_drop_m is not None:
        return -float(roof_drop_m)
    top = profile.floors[0]
    if top.height_m <= 0:
        raise MissingProjectInputError(
            f"압력표 최상단 행('{top.floor_label}')의 층고가 비어 있어 "
            "라이저가 내려오기 시작하는 표고를 정할 수 없습니다.")
    return -float(top.height_m)


def _prv_rows(profile: BuildingPressureProfile) -> list[FloorRow]:
    """감압이후 수두가 적힌 행 = 감압밸브 아래에 있는 층들."""
    return [r for r in profile.floors if r.after_prv_m is not None]


def _elev_prv1(profile: BuildingPressureProfile, zone_label: str) -> float:
    """1차 감압밸브 표고 — 감압 구간이 시작되는 첫 행."""
    rows = _prv_rows(profile)
    if not rows:
        raise MissingProjectInputError(
            f"{zone_label} 는 감압밸브가 있는 구간인데 압력표의 '감압이후' 열이 비어 있습니다. "
            "1차 감압밸브가 어느 층에 붙는지 정할 수 없습니다.")
    return -float(rows[0].head_drop_m)


def _elev_prv2(profile: BuildingPressureProfile) -> float:
    """2차 감압밸브 표고 — 감압이후 수두가 다시 낮아지는 첫 행.

    감압이후 수두는 밸브 출구를 기준으로 재므로 한 구간 안에서는 아래로 갈수록 커진다.
    다시 작아지는 자리가 새 밸브가 물린 층이다.
    """
    rows = _prv_rows(profile)
    for prev, cur in zip(rows, rows[1:]):
        if cur.after_prv_m < prev.after_prv_m:
            return -float(cur.head_drop_m)
    raise MissingProjectInputError(
        "압력표에서 2차 감압 시작 층을 찾지 못했습니다. 2차 감압 구간의 '감압이후' 수두는 "
        "밸브 출구 기준으로 다시 낮아져야 합니다.")


# ── 라이저 logical 좌표 — 화면 배치용(스키매틱)일 뿐 수리 계산에는 쓰이지 않는다.
#    표고는 elevation 필드가 따로 들고 있고, emit_sdf 의 _xform 이 이 x/y 를 정규화한다.
_COORDS_INPUT      = (   0,     0)    # 옥상 수원
_COORDS_N2         = (-500,  -100)    # 옥상 수평 분기
_COORDS_N3         = (-500,  -300)    # 옥상 → 라이저 진입
_COORDS_N4         = (-300, -1500)    # 라이저 중간 (수평 우회)
_COORDS_N7         = (-300, -2800)    # PRV 진입 직전
_COORDS_PRV_IN     = (-200, -2900)    # 노드 8 (1차 PRV in)
_COORDS_PRV_OUT    = (-100, -2950)    # 노드 89 (1차 PRV out)
_COORDS_PRV2_IN    = ( -50, -3100)    # 노드 87 (2차 PRV in, LLSP)
_COORDS_PRV2_OUT   = (  50, -3150)    # 노드 88 (2차 PRV out, LLSP)
_COORDS_N5_AV_PREV = (-300, -3300)    # AV 직전
_COORDS_AV         = ( 100, -3300)    # 노드 10 (헤드망 source ★)
_COORDS_PUMP_INPUT = ( 200,   100)    # HSP Pump-fan 앞 Input(100), Input(1) 보다 위


def _tee_to_av_pipe(label: str, ctx: ProjectContext) -> dict:
    """T분기(노드 5) → 알람밸브(노드 10). 수직 상승분 + 수평분의 L 자.

    알람밸브가 T분기점보다 위에 있다 — 수작업 모델 전량이 rise=+1.0 이다.
    """
    rise = float(ctx.tee_to_alarm_valve_m)
    return _pipe(label, "5", "10", 125, rise + TEE_TO_ALARM_VALVE_RUN_M, rise)


def _riser_upper(ctx: ProjectContext, elev_top: float,
                 elev_bottom: float) -> tuple[list[dict], list[dict]]:
    """수원(옥상) → 라이저 하강 → 노드 7 까지. 네 zone 공통 구간."""
    nodes = [
        _node("1", 0.0,       *_COORDS_INPUT, io_node="Input", pressure_pa=ATM_PA),
        _node("2", 0.0,       *_COORDS_N2),
        _node("3", elev_top,  *_COORDS_N3),
        _node("4", elev_top,  *_COORDS_N4),
        _node("7", elev_bottom, *_COORDS_N7),
    ]
    pipes = [
        _pipe("r1", "1", "2", 150, ctx.roof_run_to_riser_m, 0.0),
        _pipe("r2", "2", "3", 150, abs(elev_top), elev_top),
        _pipe("r3", "3", "4", 150, ctx.roof_run_after_drop_m, 0.0),
        _pipe("r4", "4", "7", 150, abs(elev_bottom - elev_top), elev_bottom - elev_top),
    ]
    return nodes, pipes


def _riser_prv_to_av(ctx: ProjectContext, elev_prv: float,
                     elev_av: float) -> tuple[list[dict], list[dict]]:
    """PRV 1개 → AV 까지. 노드 7 뒤에 붙는다 (LSP 1차감압 / HSP 펌프식 공통)."""
    elev_tee = elev_av - float(ctx.tee_to_alarm_valve_m)
    nodes = [
        _node("8",  elev_prv, *_COORDS_PRV_IN),
        _node("89", elev_prv, *_COORDS_PRV_OUT),
        _node("5",  elev_tee, *_COORDS_N5_AV_PREV),
        _node("10", elev_av,  *_COORDS_AV),
    ]
    pipes = [
        _pipe("r5", "7", "8",  150, RISER_PRV_APPROACH_M, 0.0),
        _pipe("r6", "89", "5", 150, abs(elev_tee - elev_prv), elev_tee - elev_prv),
        _tee_to_av_pipe("r7", ctx),
    ]
    return nodes, pipes


def build_riser_lsp_1stage(ctx: ProjectContext) -> RiserTables:
    """LSP 1차감압 라이저 — 자연낙차 + PRV 1개.

    노드: Input(1, elev=0) → 2 → 3 → 4 → 7 → 8(PRV in) → 89(PRV out) → 5 → 10(AV)
    파이프: 1→2, 2→3, 3→4, 4→7, 7→8, 89→5, 5→10
    Elastomeric-valve: 8 → 89 (target = spec.prv1_target_pa)
    """
    spec = ctx.zone_spec
    if spec.prv1_target_pa is None:
        raise MissingProjectInputError(
            "LSP_1STAGE 는 prv1_target_pa 가 필요합니다 (kg/cm² 또는 m 수두 → Pa).")
    profile = _require_profile(ctx.floor_profile, "LSP_1STAGE")
    elev_top = _elev_riser_top(profile, ctx.roof_to_top_floor_drop_m)
    elev_prv = _elev_prv1(profile, "LSP_1STAGE")
    elev_av = _elev_at_floor(profile, spec.target_floor)
    nodes, pipes = _riser_upper(ctx, elev_top, elev_prv)
    tail_nodes, tail_pipes = _riser_prv_to_av(ctx, elev_prv, elev_av)
    return RiserTables(
        nodes=nodes + tail_nodes,
        pipes=pipes + tail_pipes,
        pumps=[],
        valves=[_pressure_valve("1", "8", "89", target_pa=spec.prv1_target_pa)],
        av_node_label="10",
    )


def build_riser_hsp_pump(ctx: ProjectContext) -> RiserTables:
    """HSP 펌프식 라이저 — 자연낙차 부족 → Pump-fan 부스터 + PRV 1개.

    노드: Input(100, elev=0) → [Pump-fan 100→1] → 1 → 2 → 3 → 4 → 7 → 8(PRV in) → 89(PRV out) → 5 → 10(AV)
    Pump-fan: 100 → 1 (Library-pump = spec.pump_library_name, count=spec.pump_count)
    """
    spec = ctx.zone_spec
    if spec.prv1_target_pa is None:
        raise MissingProjectInputError("HSP_PUMP 는 prv1_target_pa 가 필요합니다.")
    profile = _require_profile(ctx.floor_profile, "HSP_PUMP")
    elev_top = _elev_riser_top(profile, ctx.roof_to_top_floor_drop_m)
    elev_prv = _elev_prv1(profile, "HSP_PUMP")
    elev_av = _elev_at_floor(profile, spec.target_floor)
    nodes, pipes = _riser_upper(ctx, elev_top, elev_prv)
    # 수원이 아니라 펌프 토출이 라이저 머리 — Input 지정을 노드 100 으로 옮긴다.
    nodes[0] = _node("1", 0.0, *_COORDS_INPUT)
    nodes.insert(0, _node("100", 0.0, *_COORDS_PUMP_INPUT,
                          io_node="Input", pressure_pa=ATM_PA))
    tail_nodes, tail_pipes = _riser_prv_to_av(ctx, elev_prv, elev_av)
    return RiserTables(
        nodes=nodes + tail_nodes,
        pipes=pipes + tail_pipes,
        pumps=[_pump_fan(str(i + 1), "100", "1", library_pump=spec.pump_library_name,
                         rated_q=spec.pump_rated_q_lpm, rated_h=spec.pump_rated_h_m)
               for i in range(max(1, spec.pump_count))],
        valves=[_pressure_valve("1", "8", "89", target_pa=spec.prv1_target_pa)],
        av_node_label="10",
    )


def build_riser_lsp_gravity(ctx: ProjectContext) -> RiserTables:
    """LSP 자연낙차 라이저 — 감압밸브 없음. PRV 자리(노드 8, 89)가 아예 없다.

    노드: Input(1, elev=0) → 2 → 3 → 4 → 7 → 5 → 10(AV)
    """
    spec = ctx.zone_spec
    profile = _require_profile(ctx.floor_profile, "LSP_GRAVITY")
    elev_top = _elev_riser_top(profile, ctx.roof_to_top_floor_drop_m)
    elev_av = _elev_at_floor(profile, spec.target_floor)
    elev_tee = elev_av - float(ctx.tee_to_alarm_valve_m)
    nodes, pipes = _riser_upper(ctx, elev_top, elev_tee)
    nodes += [
        _node("5",  elev_tee, *_COORDS_N5_AV_PREV),
        _node("10", elev_av,  *_COORDS_AV),
    ]
    pipes += [
        _pipe("r5", "7", "5", 150, RISER_PRV_APPROACH_M, 0.0),
        _tee_to_av_pipe("r6", ctx),
    ]
    return RiserTables(nodes=nodes, pipes=pipes, pumps=[], valves=[], av_node_label="10")


def build_riser_llsp_2stage(ctx: ProjectContext) -> RiserTables:
    """LLSP 2차감압 라이저 — 자연낙차 + 1차 PRV + 2차 PRV (지하주차장).

    노드: ... 89(1차 PRV out) → 87(2차 PRV in) → 88(2차 PRV out) → 5 → 10(AV)
    Elastomeric-valve: 8→89 (1차, target=prv1_target_pa), 87→88 (2차, target=prv2_target_pa)
    """
    spec = ctx.zone_spec
    if spec.prv1_target_pa is None or spec.prv2_target_pa is None:
        raise MissingProjectInputError(
            "LLSP_2STAGE 는 prv1_target_pa, prv2_target_pa 모두 필요합니다.")
    profile = _require_profile(ctx.floor_profile, "LLSP_2STAGE")
    elev_top = _elev_riser_top(profile, ctx.roof_to_top_floor_drop_m)
    elev_prv1 = _elev_prv1(profile, "LLSP_2STAGE")
    elev_prv2 = _elev_prv2(profile)
    elev_av = _elev_at_floor(profile, spec.target_floor)
    elev_tee = elev_av - float(ctx.tee_to_alarm_valve_m)
    nodes, pipes = _riser_upper(ctx, elev_top, elev_prv1)
    nodes += [
        _node("8",  elev_prv1, *_COORDS_PRV_IN),
        _node("89", elev_prv1, *_COORDS_PRV_OUT),
        _node("87", elev_prv2, *_COORDS_PRV2_IN),
        _node("88", elev_prv2, *_COORDS_PRV2_OUT),
        _node("5",  elev_tee,  *_COORDS_N5_AV_PREV),
        _node("10", elev_av,   *_COORDS_AV),
    ]
    pipes += [
        _pipe("r5", "7", "8",   150, RISER_PRV_APPROACH_M, 0.0),
        _pipe("r6", "89", "87", 150, abs(elev_prv2 - elev_prv1), elev_prv2 - elev_prv1),
        _pipe("r7", "88", "5",  150, abs(elev_tee - elev_prv2), elev_tee - elev_prv2),
        _tee_to_av_pipe("r8", ctx),
    ]
    return RiserTables(
        nodes=nodes, pipes=pipes, pumps=[],
        valves=[
            _pressure_valve("1", "8", "89", target_pa=spec.prv1_target_pa),
            _pressure_valve("2", "87", "88", target_pa=spec.prv2_target_pa),
        ],
        av_node_label="10",
    )


_RISER_BUILDERS = {
    ZoneType.HSP_PUMP: build_riser_hsp_pump,
    ZoneType.LSP_1STAGE: build_riser_lsp_1stage,
    ZoneType.LSP_GRAVITY: build_riser_lsp_gravity,
    ZoneType.LLSP_2STAGE: build_riser_llsp_2stage,
}


def build_riser(ctx: ProjectContext) -> RiserTables:
    """zone 분기 라우터. 폼 딕셔너리가 아니라 ProjectContext 하나만 받는다."""
    spec = ctx.zone_spec
    try:
        builder = _RISER_BUILDERS[spec.zone_type]
    except KeyError:
        raise ValueError(f"Unknown zone_type: {spec.zone_type}") from None
    return builder(ctx)


# ────────────────────────────────────────────────────────────────────────────
# Stage C — 라이저 ↔ 헤드망 stitch
# ────────────────────────────────────────────────────────────────────────────


@dataclass
class CombinedTables:
    """라이저 + 헤드망 결합 결과 — emit_full_sdf 입력."""
    nodes: list[dict] = field(default_factory=list)
    pipes: list[dict] = field(default_factory=list)
    nozzles: list[dict] = field(default_factory=list)
    fittings: list[dict] = field(default_factory=list)
    equipment: list[dict] = field(default_factory=list)
    pumps: list[dict] = field(default_factory=list)
    valves: list[dict] = field(default_factory=list)
    meta: list[tuple[str, str]] = field(default_factory=list)
    # 기계실 전체 평면 배관망 edge (시각화 전용, SDF 미포함). [[x1,y1,x2,y2], ...]
    machine_room_plan_edges: list[list[float]] = field(default_factory=list)


def _layout_riser_as_schematic(
    riser_nodes: list[dict],
    anchor_xy: tuple[float, float],
    head_yspan: float = 5000.0,
    descend: bool = False,
    min_span: float = 0.0,
) -> list[dict]:
    """라이저 노드를 PIPENET schematic 의 수직 막대 형태로 재배치.

    실제 계통도 DXF 의 라이저는 수십 m (60km mm) 길이라 헤드망 (5-10m schematic)
    과 통합하면 라이저가 너무 거대해서 그래프가 깨져 보임. PIPENET 답안처럼
    라이저는 헤드 군집 위로 수직 막대 (5m 정도) 로 배치해 한 화면에 깔끔히 보이게.

    Layout:
        - 모든 노드 X = anchor_xy[0] (수직 막대)
        - AV (마지막 노드) = anchor_xy 위치 (= 헤드 군집 가로 중앙 위)
        - 펌프 (첫 노드) = (anchor_xy[0], anchor_xy[1] + 라이저 yspan)
        - 중간 노드 = 균등 간격으로 사이 배치

    Args:
        riser_nodes: 라이저 노드 리스트. 인덱스 0 = 펌프, 마지막 = AV.
        anchor_xy: 막대 하단(AV) 을 놓을 좌표. 헤드망 source 노드 실좌표를 넘겨
            AV 를 source 에 정합하면 둘을 잇는 헤드 첫 배관(선언 길이 ~0)이
            긴 선으로 늘어나지 않음.
        head_yspan: 헤드망 bbox 의 y 범위 — 라이저 막대 길이는 이의 80%.
        descend: 펌프 가압(B1 펌프실) 모드. True 면 막대를 AV(헤드망) **아래**로
            내려, 펌프/수원이 화면 최하부에 오도록 한다(물이 B1→위로 가압되는
            물리 배치). 기본 False = 옥상수조(자연낙차) — 수원이 위.
        min_span: 막대 길이 하한. 기계실이 붙을 때 그 평면이 헤드 군집 밖에
            놓이도록 stitch 가 계산해 넘긴다.
    """
    n = len(riser_nodes)
    if n < 2:
        return list(riser_nodes)
    riser_yspan = max(2000.0, head_yspan * 0.8, float(min_span))
    step_y = riser_yspan / (n - 1)
    target_x = float(anchor_xy[0])
    target_y_av = float(anchor_xy[1])
    # 자연낙차: 수원 위(+). 펌프 가압: 수원 아래(-).
    y_dir = -1.0 if descend else 1.0

    out: list[dict] = []
    for i, node in enumerate(riser_nodes):
        # i=0: 펌프/수원, i=n-1: AV (head_av 위치, 막대 끝점)
        rank_from_av = (n - 1) - i
        out.append({
            **node,
            "x": int(round(target_x)),
            "y": int(round(target_y_av + y_dir * rank_from_av * step_y)),
        })
    return out


def _layout_machine_room_plan(
    mr_nodes: list[dict],
    plan_edges: list[list[float]] | None,
    pump_xy: tuple[float, float],
    head_yspan: float = 5000.0,
    conn_raw_xy: tuple[float, float] | None = None,
) -> tuple[list[dict], list[list[float]]]:
    """기계실(옥상수조) 전체 평면 배관망을 실제 x, y 형상 그대로 배치 — schematic 금지.

    기계실 DXF 는 평면도(옥상층 소방배관 평면도)라 실제 2D 배관망 형상을 가진다.
    라이저(계통도)처럼 수직 막대로 모사하면 그 형상이 뭉개지므로, 헤드망(평면도)
    과 동일하게 실제 x, y 로 보여준다. 수리계산 경로(mr_nodes, m*) 뿐 아니라 전체
    SP 배관망 edge(plan_edges)까지 **동일 변환**으로 배치해 완전한 평면도로 렌더.

        1. 경로 노드 + 전체 edge 끝점을 합친 bbox 기준 균등 스케일(aspect 보존),
           헤드 군집과 비슷한 크기로.
        2. 입상관 연결점(mK)의 raw 좌표를 펌프 노드 좌표로 보내는 평행이동.
           mK 는 라이저 Input 과 병합돼 사라진 노드라, 그 자리가 곧 펌프 노드다.
           bbox 중앙을 펌프에 맞추고 gap 만큼 띄우던 종전 방식은 이음매를 도면
           대각의 18% 만큼 벌려(대명동 실측) 두 망을 잇는 배관이 긴 사선으로
           그려졌다. 연결점을 영점으로 잡으면 이음매가 정확히 0 이 된다.

    Args:
        mr_nodes: 수리경로 노드(라벨 m*). 실제 DXF x, y 보유(원점 미변환 raw).
        plan_edges: 기계실 전체 SP 배관망 edge [[x1,y1,x2,y2], ...] (raw DXF).
        pump_xy: 펌프 junction(라이저 "1")의 schematic 좌표 — 부착 기준점.
        head_yspan: 헤드망 y-span — 기계실 평면 크기 산정 기준.
        conn_raw_xy: 입상관 연결점(mK)의 raw DXF 좌표. 미지정이면 bbox 중앙.

    Returns:
        (laid_nodes, laid_edges) — 동일 변환 적용된 경로 노드 + 전체망 edge.
    """
    pe = plan_edges or []
    pts: list[tuple[float, float]] = [
        (float(n.get("x", 0.0)), float(n.get("y", 0.0))) for n in mr_nodes
    ]
    for e in pe:
        pts.append((float(e[0]), float(e[1])))
        pts.append((float(e[2]), float(e[3])))
    if not pts:
        return list(mr_nodes), []
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    w = x_max - x_min
    h = y_max - y_min
    diag = max((w * w + h * h) ** 0.5, 1.0)
    target = max(2000.0, head_yspan * 0.7)   # 헤드 군집과 비슷한 크기로
    scale = target / diag
    if conn_raw_xy is not None:
        ax, ay = float(conn_raw_xy[0]), float(conn_raw_xy[1])
    else:
        ax, ay = (x_min + x_max) / 2.0, (y_min + y_max) / 2.0
    px, py = float(pump_xy[0]), float(pump_xy[1])

    def _tf(x: float, y: float) -> tuple[float, float]:
        return ((x - ax) * scale + px, (y - ay) * scale + py)

    laid_nodes: list[dict] = []
    for n in mr_nodes:
        nx, ny = _tf(float(n.get("x", 0.0)), float(n.get("y", 0.0)))
        laid_nodes.append({**n, "x": int(round(nx)), "y": int(round(ny))})
    laid_edges: list[list[float]] = []
    for e in pe:
        x1, y1 = _tf(float(e[0]), float(e[1]))
        x2, y2 = _tf(float(e[2]), float(e[3]))
        laid_edges.append([int(round(x1)), int(round(y1)),
                           int(round(x2)), int(round(y2))])
    return laid_nodes, laid_edges


def count_reducers_snapped_to_tee(nodes: list[dict], pipes: list[dict]) -> int:
    """관경 전이 지점 중 T분기점에 귀속되는 개수.

    현업은 분기점에서 ``REDUCER_SNAP_TO_TEE_MM`` 안에 들어오는 레듀서를 따로 세지
    않고 그 분기점에 붙은 것으로 본다. 지금 추출망은 관경이 노드에서만 바뀌므로
    이 판정이 실제로 형상을 바꾸지는 않는다 — 자동본이 수동본과 다를 때 "몇 곳을
    그렇게 봤는지" 를 말할 수 있게 세기만 한다 (작업지시서 4-3).
    """
    degree: dict[str, int] = {}
    bores: dict[str, set] = {}
    for p in pipes:
        # 라이저는 int, 헤드망은 문자열로 실어 오는 자리가 있어 그대로 비교하면
        # 150 과 "150" 이 다른 관경으로 잡힌다.
        dia = _to_opt_float(p.get("dia"))
        for key in ("in", "out"):
            lbl = str(p.get(key, ""))
            if not lbl:
                continue
            degree[lbl] = degree.get(lbl, 0) + 1
            bores.setdefault(lbl, set()).add(dia)

    xy: dict[str, tuple[float, float]] = {}
    for n in nodes:
        try:
            xy[str(n["label"])] = (float(n["x"]), float(n["y"]))
        except (KeyError, TypeError, ValueError):
            continue

    tees = [lbl for lbl, deg in degree.items() if deg >= 3]
    tee_xy = [xy[lbl] for lbl in tees if lbl in xy]
    tee_set = set(tees)
    limit_sq = REDUCER_SNAP_TO_TEE_MM ** 2

    snapped = 0
    for lbl, dias in bores.items():
        if len(dias) < 2:
            continue
        if lbl in tee_set:
            snapped += 1
            continue
        here = xy.get(lbl)
        if here is None:
            continue
        if any((here[0] - tx) ** 2 + (here[1] - ty) ** 2 <= limit_sq
               for tx, ty in tee_xy):
            snapped += 1
    return snapped


def stitch_riser_and_heads(
    riser: RiserTables,
    head_tables: Any,
    machine_room_labels: set[str] | list[str] | None = None,
    pump_junction_label: str | None = None,
    machine_room_plan_edges: list[list[float]] | None = None,
    machine_room_at_bottom: bool = False,
    machine_room_conn_xy: tuple[float, float] | None = None,
) -> CombinedTables:
    """라이저 끝점(AV node) ↔ 헤드망 source(=같은 label) 결합 + schematic 좌표 정렬.

    Args:
        riser: Stage B 산출 (legacy template) 또는 extract_system_path 결과.
        head_tables: remote30_prototype.PipeTables 인스턴스.

    Returns:
        CombinedTables — 좌표가 정렬된 한 schematic.

    충돌 처리:
        라이저 노드 라벨 = {1..9, ...} + AV(10) — 또는 v1 path: {1, n2..nN, 10}
        헤드망 노드 라벨 = {10, 11, 12, ...} (10 이 source = AV)
        AV(10) 만 공통 — 라이저 쪽 노드만 유지, 헤드망 사본은 skip.

    좌표 정렬 (PIPENET answer schematic 스타일):
        legacy 28F 템플릿은 6 노드의 작은 schematic 좌표 (0~15K 범위) 를 affine
        변환으로 사용자 클릭 위치에 맞춤. v1 path 는 실제 DXF 절대 좌표 (수십m)
        를 그대로 써서 통합 시 라이저가 헤드망보다 훨씬 거대해 보임.

        해결: 라이저를 _layout_riser_as_schematic 로 헤드 AV 위치 위 수직 막대
        로 재배치. 라이저 막대 길이는 헤드 bbox y-span 의 80%. 결과는 답안 SDF
        의 schematic 처럼 한 시각 영역 안에 컴팩트하게 부착.
    """
    av_lbl = riser.av_node_label
    riser_av_node = next((n for n in riser.nodes if n["label"] == av_lbl), None)
    if riser_av_node is None:
        raise ValueError(f"라이저에 AV 노드(label={av_lbl})가 없음")

    head_av_node = next((n for n in head_tables.nodes if n["label"] == av_lbl), None)

    # 헤드망 bbox 계산 (라이저 schematic 크기 결정용)
    head_xs = [float(nd.get("x", 0.0)) for nd in head_tables.nodes if "x" in nd]
    head_ys = [float(nd.get("y", 0.0)) for nd in head_tables.nodes if "y" in nd]
    head_yspan = (max(head_ys) - min(head_ys)) if head_ys else 5000.0

    # 노드를 진짜 라이저(계통도) 와 기계실(평면도) 로 분리.
    #   - 라이저: 실제 좌표가 수십 m 라 schematic 수직 막대로 재배치.
    #   - 기계실: 평면도이므로 실제 x, y 형상을 유지(평면 군집). 막대로 뭉개지 않음.
    mr_set = {str(l) for l in (machine_room_labels or [])}
    true_riser_nodes = [n for n in riser.nodes if str(n["label"]) not in mr_set]
    mr_nodes = [n for n in riser.nodes if str(n["label"]) in mr_set]

    # 라이저 좌표 schematic 재배치 — AV(막대 하단)를 헤드망 source 노드 실좌표에 snap.
    #   AV 노드는 (1) 라이저 막대 하단이자 (2) 헤드망 첫 메인(파이프 av_lbl)의 source 로
    #   같은 논리 분기점이다. 막대를 헤드 군집 중앙 위에 띄우면 AV 가 헤드 source 위치에서
    #   떨어져, 그 둘을 잇는 헤드 첫 배관(선언 길이 ~0)이 두 좌표계를 가로지르는 긴 선으로
    #   그려졌다. → anchor 를 head_av_node 좌표로 잡아 AV 를 source 에 정합, 연결선을 0 으로.
    # 기계실 평면은 이음매(mK)를 원점으로 먼저 배치해 크기를 잰다 — 막대를 그만큼
    # 늘려야 기계실이 헤드 군집 **밖**에 놓인다. 막대 길이를 헤드 y-span 의 80% 로
    # 고정하면 막대 끝(펌프)이 헤드 군집 안에 남아, 거기 붙는 기계실 평면(70% 크기)이
    # 헤드 배관과 통째로 겹쳐 그려졌다(대명동 실측: 기계실 78 edge 전부 헤드망 위).
    mr_rel: list[dict] = []
    plan_rel: list[list[float]] = []
    if mr_nodes and pump_junction_label is not None:
        try:
            mr_rel, plan_rel = _layout_machine_room_plan(
                mr_nodes, machine_room_plan_edges, (0.0, 0.0),
                head_yspan=head_yspan, conn_raw_xy=machine_room_conn_xy,
            )
        except (KeyError, TypeError, ValueError):
            mr_rel, plan_rel = [], []

    min_span = 0.0
    if mr_rel and head_ys and head_av_node is not None:
        rel_ys = [float(n.get("y", 0.0)) for n in mr_rel]
        for e in plan_rel:
            rel_ys.append(float(e[1]))
            rel_ys.append(float(e[3]))
        pad = max(1000.0, head_yspan * 0.08)
        try:
            anchor_y = float(head_av_node["y"])
        except (KeyError, TypeError, ValueError):
            anchor_y = 0.0
        if machine_room_at_bottom:   # 막대가 아래로 → 기계실 최상단이 헤드망 아래
            min_span = anchor_y + max(rel_ys) - min(head_ys) + pad
        else:                        # 막대가 위로 → 기계실 최하단이 헤드망 위
            min_span = max(head_ys) + pad - min(rel_ys) - anchor_y
        min_span = max(0.0, min_span)

    if head_av_node is not None and head_xs and head_ys:
        try:
            anchor_xy = (float(head_av_node["x"]), float(head_av_node["y"]))
            translated_riser_nodes = _layout_riser_as_schematic(
                true_riser_nodes, anchor_xy, head_yspan=head_yspan,
                descend=machine_room_at_bottom, min_span=min_span,
            )
        except (KeyError, TypeError, ValueError):
            translated_riser_nodes = list(true_riser_nodes)
    else:
        translated_riser_nodes = list(true_riser_nodes)

    # 기계실 평면 배치 — 원점 배치본을 펌프 junction("1")의 schematic 좌표로 평행이동.
    #   수리경로 노드(mr_nodes) + 전체 SP 배관망 edge(plan_edges)가 **동일 변환**이라
    #   완전한 평면도로 렌더된다. plan_laid 는 시각화 전용(SDF 미포함).
    plan_laid: list[list[float]] = []
    mr_laid = list(mr_nodes)
    if mr_rel:
        pump_node = next((n for n in translated_riser_nodes
                          if str(n["label"]) == str(pump_junction_label)), None)
        if pump_node is not None:
            px = int(round(float(pump_node.get("x", 0.0))))
            py = int(round(float(pump_node.get("y", 0.0))))
            mr_laid = [{**n, "x": n["x"] + px, "y": n["y"] + py} for n in mr_rel]
            plan_laid = [[e[0] + px, e[1] + py, e[2] + px, e[3] + py]
                         for e in plan_rel]

    translated_riser_nodes = translated_riser_nodes + mr_laid

    # 헤드망 표고를 라이저 표고계로 rebase — AV 는 두 추출의 유일한 공통 노드이므로
    # 그 지점의 z 가 양쪽에서 완전히 같아야 한다. 평면도 추출은 전 노드에 평탄한
    # 상대표고를 주고 계통도 추출은 수원=0 기준이라, 그대로 합치면 AV 를 잇는 배관
    # 하나에 두 기준점 차이만큼의 가짜 낙차가 실린다(대명동 실측 80.95m).
    # 배관의 "elev" 는 구간 낙차(delta)라 기준 이동의 영향을 받지 않는다.
    # head_tables 는 캐시되어 재사용되므로 사본으로 만든다.
    head_elev_shift = 0.0
    if head_av_node is not None:
        head_elev_shift = (float(riser_av_node.get("elevation", 0.0))
                           - float(head_av_node.get("elevation", 0.0)))
    head_nodes_filtered = []
    for n in head_tables.nodes:
        if n["label"] == av_lbl:
            # AV 는 라이저 쪽에서 이미 포함 — 헤드망 쪽 사본 skip
            continue
        if head_elev_shift:
            n = {**n, "elevation": round(float(n.get("elevation", 0.0))
                                         + head_elev_shift, 3)}
        head_nodes_filtered.append(n)

    # 파이프 라벨 전역 유일화 — 계통도·평면도·기계실이 전부 r1.. 컨벤션을 쓰므로
    # 합치면 라벨이 충돌한다. parse_sdf 의 plabel_to_pid 는 라벨 문자열로 keying 해
    # 같은 라벨 두 파이프를 한 K-solver pid 로 접어버려 KFP 토폴로지가 붕괴된다.
    # → 충돌하는 두 번째 이후 항목만 개명(원본 dict 불변, 사본 생성).
    combined_pipes: list[dict] = []
    seen_pipe_labels: set[str] = set()
    for p in (list(riser.pipes) + list(head_tables.pipes)):
        lbl = str(p.get("label", ""))
        if lbl and lbl not in seen_pipe_labels:
            seen_pipe_labels.add(lbl)
            combined_pipes.append(p)
            continue
        base = lbl or "p"
        k = 2
        new_lbl = f"{base}_{k}"
        while new_lbl in seen_pipe_labels:
            k += 1
            new_lbl = f"{base}_{k}"
        seen_pipe_labels.add(new_lbl)
        combined_pipes.append({**p, "label": new_lbl})

    # ── 표준 소화배관 밸브(Fitting) 주입.
    # 라이저 빌더는 fitting 을 생성하지 않아(통합 fitting 은 head_tables.fittings 만),
    # 정답 SDF 의 펌프 토출부 개폐/체크밸브·알람밸브 1차측 버터플라이가 통째로 빠진다.
    # 화재안전기준 표준 배치를 결정론적으로 부착한다:
    #   · 수원/펌프 토출 배관(Input 경계 노드 직결): gate + check
    #   · 알람밸브 배관(A/V Equipment 보유): butterfly
    # (정답 2. Pipenet_hand.sdf 의 pipe"1"=gate+check, pipe"9"[A/V]=butterfly 와 정합.)
    combined_fittings = list(head_tables.fittings)

    def _pipe_by_label(lbl: str) -> dict | None:
        return next((p for p in combined_pipes if str(p["label"]) == str(lbl)), None)

    def _add_valve(pipe: dict, vtype: str) -> None:
        combined_fittings.append({
            "pipe": pipe["label"], "in": pipe["in"], "out": pipe["out"],
            "type": vtype, "count": "1",
        })

    av_eq = next((e for e in head_tables.equipment
                  if str(e.get("desc", "")).upper() == "A/V" and e.get("pipe")), None)
    if av_eq is not None:
        av_pipe = _pipe_by_label(av_eq["pipe"]) or next(
            (p for p in combined_pipes
             if str(p["in"]) == str(av_eq.get("in")) and str(p["out"]) == str(av_eq.get("out"))),
            None)
        if av_pipe is not None:
            _add_valve(av_pipe, "butterfly")

    input_label = next((str(n["label"]) for n in (translated_riser_nodes + head_nodes_filtered)
                        if str(n.get("io_node", "")).lower() == "input"), None)
    if input_label is not None:
        src_pipe = next((p for p in combined_pipes
                         if str(p["in"]) == input_label or str(p["out"]) == input_label), None)
        if src_pipe is not None:
            _add_valve(src_pipe, "gate")
            _add_valve(src_pipe, "check")

    combined_nodes = translated_riser_nodes + head_nodes_filtered
    meta = list(getattr(head_tables, "meta", []))
    snapped = count_reducers_snapped_to_tee(combined_nodes, combined_pipes)
    meta.append(("레듀서 T분기 귀속",
                 f"{snapped} 곳 (기준 {REDUCER_SNAP_TO_TEE_MM:.0f}mm)"))

    return CombinedTables(
        nodes=combined_nodes,
        pipes=combined_pipes,
        nozzles=list(head_tables.nozzles),
        fittings=combined_fittings,
        equipment=list(head_tables.equipment),
        pumps=list(riser.pumps),
        valves=list(riser.valves),
        meta=meta,
        machine_room_plan_edges=plan_laid,
    )


def prepend_machine_room_to_riser(
    machine_room: dict, riser: RiserTables, *,
    at_bottom: bool = False, source_drop_below_lowest_m: float = 0.0,
) -> tuple[RiserTables, bool]:
    """기계실 경로를 라이저 Input 노드 앞에 prepend → 확장 RiserTables.

    수원 경계가 라이저 top('1')이 아니라 그보다 상류인 기계실 수면(m1)으로 이동한다.
    기계실 경로 = 수원(m1, Input) → 입상관 연결점(mK). mK 를 라이저 Input 노드와
    병합하고 라이저 Input 은 일반 분기로 강등(경계는 이제 m1 하나). 이렇게 기계실부
    배관의 마찰손실과 고저차가 통합망 수리계산에 반영된다.

    좌표 정합 (x, y): 기계실 노드의 raw DXF x, y 는 그대로 보존한다. 통합 캔버스
    배치는 stitch 단계의 _layout_machine_room_plan 이 수리경로 노드 + 전체 SP
    배관망 edge(plan_edges, 동일 raw 좌표계)를 한 변환으로 묶어 펌프 위에 부착하므로,
    여기서 미리 translation 하면 오히려 plan_edges 와 어긋난다. → x, y 무변환.

    elevation (가압방식에 따라 기준이 다름 — 수리결과를 바꾸는 핵심):
      • at_bottom=False (고가수조/자연낙차): 기계실(옥상수조)은 라이저 top 과
        동일한 옥상 레벨(수평)이므로 기계실 노드 elev 를 라이저 Input(옥상) elev 로
        offset. 수원이 망 최상부 → 자연낙차로 하류에 양압 공급.
      • at_bottom=True (펌프 가압): 기계실/수원이 최하부(예: B1)다. 라이저 Input
        (옥상)이 아니라 **라이저 최저 고도(=서비스 최저층, 보통 AV)** 를 기준으로
        삼되, ``source_drop_below_lowest_m`` 만큼 그 아래로 더 내려 수원을 둔다.
        도면(DXF)에는 z 가 없어 기계실의 실제 깊이를 추출할 수 없으므로, 이 깊이는
        사용자가 가압방식 패널에서 직접 입력한다(미지정 시 0 = 최저헤드와 동일 고도).
        그 결과 수원→최저헤드 사이에 ``source_drop_below_lowest_m`` 만큼의 양(+)의
        정수두(lift)가 생겨 펌프가 극복해야 할 실양정으로 계산에 반영된다.
        (옥상 기준이면 이 lift 가 0 으로 사라져 펌프 실양정이 과소평가됨.)

    Args:
        source_drop_below_lowest_m: at_bottom 일 때만 사용. 수원(기계실)이 라이저
            최저 노드보다 몇 m 더 아래에 있는지(>0). 헤드 대비 실제 흡입 고저차.

    반환: (확장 RiserTables, attached) — attached 는 실제 병합 성공 여부.
    machine_room 가 비었거나 라이저 Input 을 못 찾으면 (원본 riser, False) 반환(안전).
    """
    mr_nodes = list(machine_room.get("nodes", []))
    mr_pipes = list(machine_room.get("pipes", []))
    if not mr_nodes or not mr_pipes:
        return riser, False

    conn_label = str(machine_room.get("conn_node_label") or mr_nodes[-1]["label"])

    riser_input = next(
        (n for n in riser.nodes if str(n.get("io_node", "")).lower() == "input"), None)
    if riser_input is None:
        riser_input = next((n for n in riser.nodes if str(n["label"]) == "1"), None)
    if riser_input is None:
        return riser, False  # 정합 불가 — 기계실 skip
    riser_input_label = str(riser_input["label"])
    riser_input_elev = float(riser_input.get("elevation", 0.0))

    # 기계실 고도 기준(offset) — 가압방식에 따라 분기.
    #  · 고가수조: 라이저 Input(옥상) 고도. 수원이 최상부.
    #  · 펌프 가압(at_bottom): 라이저 최저 고도에서 source_drop_below_lowest_m 만큼
    #    더 아래. 수원이 최저헤드보다 아래 → 양(+)의 실양정(lift)이 계산에 반영됨.
    if at_bottom:
        lowest = min(
            (float(n.get("elevation", 0.0)) for n in riser.nodes),
            default=riser_input_elev,
        )
        mr_ref_elev = lowest - abs(float(source_drop_below_lowest_m))
    else:
        mr_ref_elev = riser_input_elev

    # 기계실 노드: x,y 는 raw DXF 좌표 그대로(무변환 — plan_edges 와 동일 좌표계 유지),
    # elev 는 mr_ref_elev 기준으로 offset + conn 노드(mK) 제거(병합)
    new_mr_nodes: list[dict] = []
    for n in mr_nodes:
        if str(n["label"]) == conn_label:
            continue  # mK ≡ riser_input — 중복 제거
        nn = dict(n)
        nn["elevation"] = round(mr_ref_elev + float(n.get("elevation", 0.0)), 3)
        new_mr_nodes.append(nn)

    # 기계실 pipe: conn(mK) 향하던 끝점을 riser_input_label 로 재지정
    new_mr_pipes: list[dict] = []
    for p in mr_pipes:
        pp = dict(p)
        if str(pp.get("out")) == conn_label:
            pp["out"] = riser_input_label
        if str(pp.get("in")) == conn_label:
            pp["in"] = riser_input_label
        new_mr_pipes.append(pp)

    # 라이저 Input 강등: Input→No, pressure 제거 (경계는 이제 m1)
    new_riser_nodes: list[dict] = []
    for n in riser.nodes:
        nn = dict(n)
        if str(n["label"]) == riser_input_label:
            nn["io_node"] = "No"
            nn.pop("pressure_pa", None)
        new_riser_nodes.append(nn)

    return RiserTables(
        nodes=new_mr_nodes + new_riser_nodes,
        pipes=new_mr_pipes + list(riser.pipes),
        pumps=list(riser.pumps),
        valves=list(riser.valves),
        av_node_label=riser.av_node_label,
    ), True


# ────────────────────────────────────────────────────────────────────────────
# 펌프 가압 방식 — 통합망 수원(Input 경계) 직후 펌프 요소 삽입
# ────────────────────────────────────────────────────────────────────────────

# 화재안전기준(NFPC 103) 펌프 성능시험 곡선 기준:
#   체절운전(Q=0)      : 양정 ≤ 정격양정의 140%
#   정격(Q=정격)        : 정격양정
#   150% 유량(Q=1.5정격): 양정 ≥ 정격양정의 65%
PUMP_SHUTOFF_HEAD_RATIO = 1.40
PUMP_OVERLOAD_Q_RATIO = 1.50
PUMP_OVERLOAD_HEAD_RATIO = 0.65


def insert_source_pump(
    combined: CombinedTables,
    *,
    rated_q_lpm: float,
    rated_h_m: float,
    count: int = 1,
    pump_name: str = "FP",
    efficiency: int = 100,
) -> CombinedTables:
    """통합망 수원(Input 경계) 직후에 펌프 요소를 삽입한다 (펌프 가압 방식).

    물리 모델: 수원(Input, 대기압) → [Pump-fan] → 토출노드 → (기존 배관).
    수원에서 출발하던 모든 파이프의 시작점을 새 토출노드로 옮기고, 펌프 요소가
    수원→토출노드를 잇는다. 수원 노드는 그대로 대기압 경계(Calculation-spec)로
    남으므로 parse_sdf 의 pressure_bar 가 None 이 아니어서 펌프 양정이 boundary
    압력으로 잘못 주입되는 일이 없다.

    화재안전기준 표준 3점 성능곡선(체절 140% / 정격 / 150% 65%)을 펌프 dict 에
    담아 emit_full_sdf 의 Pump-fan attribute + has_converter 의 pumpFlowDataTable
    로 전파한다. 이로써 SDF/KFP/HAS 모두 동일 곡선으로 연산 가능해진다.

    Args:
        combined: stitch_riser_and_heads 산출. in-place 로 펌프/토출노드 추가.
        rated_q_lpm: 정격 토출량 (L/min).
        rated_h_m:   정격 양정 (m).
        count:       병렬 펌프 개수 (기본 1 — 운전 듀티 펌프 1대).
        pump_name:   Library-pump / HAS PumpType 이름 (기본 "FP").

    Returns:
        combined (동일 객체) — 펌프/토출노드가 추가됨.

    Raises:
        ValueError: 통합망에 Input 경계 노드가 없을 때.
    """
    src = next((n for n in combined.nodes
                if str(n.get("io_node", "")).lower() == "input"), None)
    if src is None:
        raise ValueError("통합망에 Input 경계 노드(수원)가 없어 펌프를 삽입할 수 없습니다.")
    src_label = str(src["label"])

    # 토출 노드 — 수원과 같은 고도(수평), 경계 아님. 라벨 유일 보장.
    existing = {str(n["label"]) for n in combined.nodes}
    disch_label = f"{src_label}_pd"
    _k = 1
    while disch_label in existing:
        _k += 1
        disch_label = f"{src_label}_pd{_k}"
    disch = {
        "label": disch_label,
        "x": int(src.get("x", 0)) + 400,   # 캔버스에서 펌프가 보이도록 약간 오프셋
        "y": int(src.get("y", 0)),
        "elevation": float(src.get("elevation", 0.0)),
        "io_node": "No",
    }
    combined.nodes.append(disch)

    # 수원에서 나가던 파이프(in==수원)의 시작점을 토출노드로 재지정.
    # (트리 root 인 수원은 모든 배관이 out 방향 → in==src 만 검사하면 충분)
    for p in combined.pipes:
        if str(p.get("in")) == src_label:
            p["in"] = disch_label

    q = float(rated_q_lpm)
    h = float(rated_h_m)
    shutoff_h = round(h * PUMP_SHUTOFF_HEAD_RATIO, 3)
    peak_q = round(q * PUMP_OVERLOAD_Q_RATIO, 3)
    peak_h = round(h * PUMP_OVERLOAD_HEAD_RATIO, 3)

    # 소화펌프는 주+예비(또는 N대) 구성이나, 예비는 동시 운전하지 않는 신뢰성
    # 이중화 → 수리계산은 운전 듀티 1대 곡선으로 한다(화재안전기준). 같은 in/out
    # 노드에 N개 Pump-fan 을 직렬화하면 PIPENET/KFP 가 N대 병렬 = N배 용량으로
    # 잘못 계산하므로, Pump-fan 은 단 하나만 만들고 대수는 count 메타로 보존한다.
    n_pumps = max(1, int(count))
    combined.pumps.append({
        "label": pump_name,
        "in": src_label,
        "out": disch_label,
        "efficiency": int(efficiency),
        "status": 1,
        "library_pump": pump_name,
        "percentage_open": 1,
        "pump_type": pump_name,
        "count": n_pumps,            # 설치 대수(주+예비) — 표시/문서용, 수리계산 비반영
        # 성능곡선 (정격 + 체절 + 150% 과부하) — m / L·min 단위
        "rated_q": q,
        "rated_h": h,
        "shutoff_h": shutoff_h,
        "peak_q": peak_q,
        "peak_h": peak_h,
    })
    return combined


# ────────────────────────────────────────────────────────────────────────────
# Stage D — emit_full_sdf (PIPENET-native 후처리 + Pump-fan / Elastomeric-valve)
# ────────────────────────────────────────────────────────────────────────────

# 물 ρg (kg/m³ × m/s²) — 양정(m) → 압력(Pa). 표준 SLF 펌프점과 동일 계수.
# M_TO_PA 와 같은 물리량(1 m 수두 → Pa). 의미가 다른 두 이름이라 alias 로 단일화.
_WATER_RHO_G = M_TO_PA


def _harden_slf_for_combined(
    slf_path: Path,
    opt_flow_by_lib: dict[str, float],
    pumps: list[dict],
) -> list[dict]:
    """동봉 SLF 라이브러리를 통합망에 맞게 보정 — PIPENET 연산 경고/에러 제거.

    1) Nozzle 최소운전압력 ↓ : 표준 SLF 의 SP-HEAD minimum-pressure 가 헤드
       설계유량(optimum flow)에 해당하는 압력 (Q/k)² 보다 높으면 모든 헤드에
       "optimum flow below minimum operating pressure" 경고가 발생한다. 각
       노즐 정의의 minimum-pressure 를 설계유량 압력의 90% 이하로 낮춘다.
    2) Pump 라이브러리 주입 : Pump-fan 이 참조하는 library_pump(예 "FP") 가
       SLF Pump-section 에 없으면 곡선 범위가 미정의되어 "Minimum flowrate
       should be less than maximum" 에러가 난다. 정격유량·양정이 실려 있으면
       NFPC 3점 곡선(체절 140% / 정격 / 150% 65%)으로 Pump-definition 을 만들어
       주입하고, 없으면 지어내지 않고 미해결로 돌려준다.

    SLF 는 DOCTYPE(<!DOCTYPE Library SYSTEM "Library.dtd">) 를 요구하므로
    ElementTree 직렬화 후 XML 선언 + DOCTYPE 를 직접 앞에 붙여 보존한다.

    Returns:
        곡선을 확보하지 못한 Pump-fan 참조 ``[{"name":.., "reason":..}]``.
        빈 리스트는 "모든 참조가 SLF 에 실재하거나 주입됐다" 를 뜻한다 —
        대조를 못 한 경우(파일 없음·파싱 실패)는 빈 리스트가 아니라 그 사실을 담는다.
    """
    import xml.etree.ElementTree as ET
    refs = sorted({str(p.get("library_pump", "")).strip() for p in pumps} - {""})
    try:
        tree = ET.parse(slf_path)
    except (OSError, ET.ParseError) as exc:
        return [{"name": n, "reason": f"SLF 대조 불가 ({type(exc).__name__})"} for n in refs]
    root = tree.getroot()
    changed = False

    # ── (1) 노즐 최소운전압력 보정
    for ndef in root.iter("Nozzle-definition"):
        name_el = ndef.find("Item-name")
        lib = (name_el.text or "").strip() if name_el is not None else ""
        q_opt = opt_flow_by_lib.get(lib)
        if not q_opt:
            continue
        try:
            k = float(ndef.get("k-value", "0"))
        except ValueError:
            k = 0.0
        if k <= 0:
            continue
        p_opt = (q_opt / k) ** 2  # 설계유량에 필요한 노즐 압력 (Pa)
        try:
            p_min = float(ndef.get("minimum-pressure", "0"))
        except ValueError:
            p_min = 0.0
        if p_min > p_opt:
            ndef.set("minimum-pressure", f"{p_opt * 0.9:.2f}")
            changed = True

    # ── (2) 펌프 라이브러리 주입
    unresolved: list[dict] = []
    pump_sec = root.find("Pump-section")
    if pump_sec is None:
        unresolved += [{"name": n, "reason": "SLF 에 Pump-section 이 없음"} for n in refs]
    elif refs:
        existing = {
            (pd.find("Item-name").text or "").strip()
            for pd in pump_sec.findall("Pump-definition")
            if pd.find("Item-name") is not None
        }
        rated_by_lib = {
            str(p.get("library_pump", "")).strip(): p
            for p in pumps if p.get("rated_q") and p.get("rated_h")
        }
        for lib in refs:
            if lib in existing:
                continue
            pump = rated_by_lib.get(lib)
            if pump is None:
                # 곡선을 지어내지 않는다 — 정격유량·양정 없이 주입하면 PIPENET 이
                # 양정을 스스로 고르는 것과 똑같이 근거 없는 계산서가 된다.
                unresolved.append({"name": lib, "reason": "SLF 에 없음, 정격유량·양정 미입력으로 주입 실패"})
                continue
            q = float(pump["rated_q"])  # L/min
            h = float(pump["rated_h"])  # m
            q_si = q / 60000.0  # L/min → m³/s
            peak_q = float(pump.get("peak_q", q * PUMP_OVERLOAD_Q_RATIO) or q * PUMP_OVERLOAD_Q_RATIO)
            peak_q_si = peak_q / 60000.0
            shutoff_h = float(pump.get("shutoff_h", h * PUMP_SHUTOFF_HEAD_RATIO) or h * PUMP_SHUTOFF_HEAD_RATIO)
            peak_h = float(pump.get("peak_h", h * PUMP_OVERLOAD_HEAD_RATIO) or h * PUMP_OVERLOAD_HEAD_RATIO)
            pdef = ET.SubElement(pump_sec, "Pump-definition", {
                "curve-type": "quadratic",
                "flowrate-unit": "l-min",
                "max-degeneration-factor": "0",
                "max-flow": f"{peak_q_si:.9g}",
                "min-degeneration-factor": "0",
                "min-flow": "0",
                "pressure-unit": "metres",
            })
            ET.SubElement(pdef, "Item-name").text = lib
            ET.SubElement(pdef, "Description").text = lib
            pts = ET.SubElement(pdef, "Set-of-pump-points")
            # 압력은 평문 소수로 (과학표기 e+06 회피 — 표준 SLF 펌프점 포맷과 정합).
            ET.SubElement(pts, "Pump-point", {"flow": "0", "pressure": f"{shutoff_h * _WATER_RHO_G:.2f}"})
            ET.SubElement(pts, "Pump-point", {"flow": f"{q_si:.9g}", "pressure": f"{h * _WATER_RHO_G:.2f}"})
            ET.SubElement(pts, "Pump-point", {"flow": f"{peak_q_si:.9g}", "pressure": f"{peak_h * _WATER_RHO_G:.2f}"})
            changed = True

    if changed:
        body = ET.tostring(root, encoding="unicode")
        slf_path.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE Library SYSTEM "Library.dtd">\n' + body,
            encoding="utf-8",
        )
    return unresolved


def emit_full_sdf(combined: CombinedTables, out_path: Path, *,
                  ctx: ProjectContext) -> Path:
    """완성 SDF 직렬화.

    1단계: ``remote30_prototype.emit_sdf`` 호출 — PIPENET-native 후처리
           (빈 Pipe-set placeholder, 6 schedule embed, SLF 동봉, Template Graphics)
           가 모두 그대로 적용됨.
    2단계: 결과 SDF 를 다시 열어:
           - Input 노드에 ``<Calculation-spec pressure="ATM_PA"/>`` 추가
           - ``<Pump-fan>`` element 추가 (combined.pumps)
           - ``<Elastomeric-valve>`` element 추가 (combined.valves)
    """
    from remote30_prototype import emit_sdf, PipeTables, write_sdf_tree

    # 1단계: PipeTables 로 캐스팅 후 emit_sdf 호출
    tables = PipeTables(
        nodes=list(combined.nodes),
        pipes=list(combined.pipes),
        nozzles=list(combined.nozzles),
        fittings=list(combined.fittings),
        equipment=list(combined.equipment),
        meta=list(combined.meta),
    )
    emit_sdf(tables, out_path, project_title=ctx.report_title())

    # 2단계: SDF 재오픈 → Pump-fan / Elastomeric-valve / Calculation-spec 추가
    import xml.etree.ElementTree as ET
    tree = ET.parse(out_path)
    root = tree.getroot()

    # Input 노드에 boundary pressure 명시 (Calculation-spec)
    pressure_by_label: dict[str, float] = {
        n["label"]: float(n["pressure_pa"])
        for n in combined.nodes if n.get("pressure_pa") is not None
    }
    for node_el in root.iter("Node"):
        lbl = node_el.get("label", "")
        if lbl in pressure_by_label and node_el.find("Calculation-spec") is None:
            ET.SubElement(node_el, "Calculation-spec",
                          {"pressure": str(int(pressure_by_label[lbl]))})

    # Links 안에 Pump-fan + Elastomeric-valve 삽입 (Nozzle 다음, 또는 끝쪽)
    for links in root.iter("Links"):
        for pump in combined.pumps:
            pf_attrib = {
                "efficiency": str(pump["efficiency"]),
                "input": pump["in"],
                "label": pump["label"],
                "output": pump["out"],
                "status": str(pump["status"]),
            }
            # 성능곡선이 있으면 attribute 로 직렬화 → parse_sdf 가 읽어 .has/KFP 로 전파.
            # rated-p/shutoff-p/peak-p 는 양정[m]. rated-p-unit="metres" 로 명시 →
            # parse_sdf 가 bar 로 정규화(KFP→SDF 의 bar 와 구분). (수원 노드는 이미
            # 대기압 boundary 라 parse_sdf 의 pressure_bar fallback 은 발동하지 않음.)
            if pump.get("rated_q") and pump.get("rated_h"):
                pf_attrib.update({
                    "rated-q": f"{float(pump['rated_q']):g}",
                    "rated-p": f"{float(pump['rated_h']):g}",
                    "rated-p-unit": "metres",
                    "shutoff-p": f"{float(pump.get('shutoff_h', float(pump['rated_h']) * 1.4)):g}",
                    "peak-q": f"{float(pump.get('peak_q', float(pump['rated_q']) * 1.5)):g}",
                    "peak-p": f"{float(pump.get('peak_h', float(pump['rated_h']) * 0.65)):g}",
                })
            pf = ET.Element("Pump-fan", pf_attrib)
            ET.SubElement(pf, "Description")
            lib = ET.SubElement(pf, "Library-pump")
            lib.text = pump["library_pump"]
            ET.SubElement(pf, "Pump-setting",
                          {"percentage-open": str(pump["percentage_open"])})
            links.append(pf)

        for v in combined.valves:
            ev = ET.Element("Elastomeric-valve", {
                "input": v["in"],
                "label": v["label"],
                "output": v["out"],
                "target-value": f"{v['target_value']:.6g}",
                "type": v["type"],
            })
            links.append(ev)
        break  # 첫 Links 만

    write_sdf_tree(tree, out_path)

    # 동봉 SLF 보정 — 노즐 최소운전압력 / 펌프 곡선 라이브러리 (PIPENET 연산 경고·에러 제거).
    # SDF 가 PIPENET 에 넘기는 실제 optimum flow(<Flow-define flow=>) 를 라이브러리별로
    # 모아 SLF 의 minimum-pressure 와 비교 보정한다.
    opt_flow_by_lib: dict[str, float] = {}
    for nz_el in root.iter("Nozzle"):
        fd = nz_el.find("Flow-define")
        li = nz_el.find("Library-item")
        if fd is None or li is None:
            continue
        lib = (li.text or "").strip()
        try:
            q = float(fd.get("flow", "0"))
        except ValueError:
            continue
        if lib and q > 0:
            opt_flow_by_lib[lib] = min(opt_flow_by_lib.get(lib, q), q)
    # 방출 게이트 — Pump-fan 이 가리키는 Library-pump 가 보정 뒤 SLF 에 실재하는가.
    # 없는데도 조용히 넘어가면 PIPENET 이 양정을 임의 선정한 계산서가 나오고,
    # 출력물만 봐서는 그 사실을 알 수 없다. 방출은 막지 않고(기존 미확정 관행)
    # 경고 블록에 올린다.
    for miss in _harden_slf_for_combined(out_path.with_suffix(".slf"),
                                         opt_flow_by_lib, list(combined.pumps)):
        ctx.note_emit_issue("pump_library_name", f"펌프 성능곡선 {miss['name']}", miss["reason"])
    return out_path


# ────────────────────────────────────────────────────────────────────────────
# 통합 파이프라인 — 한 번에 모든 Stage 진행 (서버 API 에서 호출)
# ────────────────────────────────────────────────────────────────────────────


def run_full_pipeline(inputs: OverallInputs, head_tables: Any, out_path: Path) -> Path:
    """Stage B + C + D 일괄 실행. Stage A 는 별도 (서버 측에서 사용자 헤드 편집 단계 필요).

    Args:
        inputs: OverallInputs (zone_spec, profile, project_title 사용).
        head_tables: remote30_prototype.PipeTables — Stage A 의 결과.
        out_path: 출력 SDF 경로.
    """
    ctx = inputs.to_context()
    riser = build_riser(ctx)
    combined = stitch_riser_and_heads(riser, head_tables)
    return emit_full_sdf(combined, out_path, ctx=ctx)


# ────────────────────────────────────────────────────────────────────────────
# 직접 입력 폼 → ZoneSpec / BuildingPressureProfile 변환
# ────────────────────────────────────────────────────────────────────────────

def zone_spec_from_form(form: dict[str, Any]) -> ZoneSpec:
    """HTML 폼 데이터 → ZoneSpec.

    필수 폼 필드::

        zone_type           : "hsp_pump" / "lsp_1stage" / "llsp_2stage" / "lsp_gravity"
        target_floor        : "16층" 등 라벨

    선택 폼 필드 (감압 zone)::

        prv1_target_kgf     : 1차 PRV 출구압 (kg/cm²) — Pa 변환됨
        prv2_target_kgf     : 2차 PRV 출구압 (kg/cm²) — LLSP_2STAGE 만
        prv1_target_m       : 1차 PRV 출구압 (m 수두) — kg/cm² 와 둘 중 하나만
        prv2_target_m       : 2차 PRV 출구압 (m 수두)
        pump_library_name   : Library-pump 이름 (HSP_PUMP, 기본 DEFAULT_PUMP_LIBRARY_NAME)
        pump_count          : Pump-fan 개수 (HSP_PUMP, 기본 2)
        pump_rated_q_lpm    : 정격유량 (L/min) — 없으면 곡선 주입 불가, 미확정 경고
        pump_rated_h_m      : 정격양정 (m) — 없으면 곡선 주입 불가, 미확정 경고
    """
    zone_type = ZoneType(form.get("zone_type", "lsp_1stage"))

    def _to_pa(kgf_key: str, m_key: str) -> float | None:
        kgf = form.get(kgf_key, "")
        m = form.get(m_key, "")
        if kgf:
            try:
                return float(kgf) * KGF_CM2_TO_PA
            except ValueError:
                pass
        if m:
            try:
                return float(m) * M_TO_PA
            except ValueError:
                pass
        return None

    return ZoneSpec(
        zone_type=zone_type,
        target_floor=str(form.get("target_floor", "")).strip(),
        prv1_target_pa=_to_pa("prv1_target_kgf", "prv1_target_m"),
        prv2_target_pa=_to_pa("prv2_target_kgf", "prv2_target_m"),
        pump_library_name=str(form.get("pump_library_name", DEFAULT_PUMP_LIBRARY_NAME)).strip(),
        pump_count=_to_int(form.get("pump_count"), 2),
        pump_rated_q_lpm=_to_opt_float(form.get("pump_rated_q_lpm")),
        pump_rated_h_m=_to_opt_float(form.get("pump_rated_h_m")),
    )


def parse_system_diagram_dxf(dxf_path: Path, *, default_height_m: float = 2.9,
                              roof_height_m: float = 6.0) -> BuildingPressureProfile:
    """계통도 DXF 의 텍스트 라벨 → BuildingPressureProfile 자동 추정.

    추출 패턴 (대명동 201동 계통도 분석 기반):
        "지상N층"   → "N층" label
        "지하N층"   → "지하N층" label
        "옥상", "PH", "PH N F"  → "옥상층" label
        "Nf" / "NF" / "NFL"  (fallback)

    elevation 계산:
        TEXT 의 Y 좌표를 정렬 (큰 Y = 위쪽) → 최상부부터 default_height_m 씩 누적.
        도면의 Y 단위가 도면별로 달라 직접 사용은 불안정. 층고는 default 로 채우고
        사용자가 검토/수정.

    Args:
        dxf_path: 계통도 DXF.
        default_height_m: 표준 층고 (기본 2.9m).
        roof_height_m: 옥상층의 층고 (기본 6m — PDF 압력표 기준).

    Returns:
        BuildingPressureProfile — 최상부 → 최하부 순. 비어있을 수도 (라벨 0건).
    """
    import re
    try:
        import ezdxf as _ezdxf
    except ImportError as exc:
        raise RuntimeError("ezdxf 가 설치되지 않아 계통도 DXF 파싱 불가") from exc

    doc = _ezdxf.readfile(str(dxf_path))
    msp = doc.modelspace()

    PAT_ABOVE = re.compile(r"지상\s*(\d+)\s*층")
    PAT_BELOW = re.compile(r"지하\s*(\d+)\s*층")
    PAT_ROOF  = re.compile(r"옥상|PH(?:\s*\d+)?\s*F?")
    PAT_NF    = re.compile(r"^\s*(\d{1,2})\s*F\s*$")

    candidates: list[tuple[str, float]] = []
    for e in msp:
        if e.dxftype() not in ("TEXT", "MTEXT"):
            continue
        try:
            v = e.dxf.text if e.dxftype() == "TEXT" else e.text
        except Exception:
            continue
        if not v or not v.strip():
            continue
        v = v.strip()
        try:
            pos = e.dxf.insert
            y = float(pos.y)
        except Exception:
            continue
        if (m := PAT_ABOVE.search(v)):
            candidates.append((f"{m.group(1)}층", y))
            continue
        if (m := PAT_BELOW.search(v)):
            candidates.append((f"지하{m.group(1)}층", y))
            continue
        if (m := PAT_NF.search(v)):
            candidates.append((f"{m.group(1)}층", y))
            continue
        if PAT_ROOF.search(v):
            candidates.append(("옥상층", y))

    # 같은 label 중복 시 가장 큰 Y 사용 (도면 상단 라벨이 가장 신뢰)
    by_label: dict[str, float] = {}
    for label, y in candidates:
        if label not in by_label or y > by_label[label]:
            by_label[label] = y

    # Y 내림차순 정렬 — 최상부 (Y 큰 것) → 최하부
    sorted_floors = sorted(by_label.items(), key=lambda kv: -kv[1])

    rows: list[FloorRow] = []
    cumulative_drop = 0.0
    for i, (label, _y) in enumerate(sorted_floors):
        height = roof_height_m if label == "옥상층" else default_height_m
        if i > 0:
            cumulative_drop += height
        rows.append(FloorRow(
            floor_label=label,
            height_m=height,
            head_drop_m=round(cumulative_drop, 1),
            note="(계통도 자동 추출 — 층고/낙차 검토 필요)",
        ))

    return BuildingPressureProfile(
        building_name=dxf_path.stem,
        floors=rows,
    )


def profile_from_form(form: dict[str, Any]) -> BuildingPressureProfile | None:
    """HTML 폼의 row 배열(JSON) → BuildingPressureProfile.

    폼 필드 ``pressure_table_json`` 이 있으면 파싱, 없으면 None.
    JSON 형식 예::

        [{"floor_label": "옥상층", "height_m": 6,   "head_drop_m": 3.1},
         {"floor_label": "49층",   "height_m": 3.1, "head_drop_m": 6.2},
         ...]
    """
    raw = str(form.get("pressure_table_json") or "").strip()
    if not raw:
        return None
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return None
    if not isinstance(data, list):
        return None
    rows = [row for row in map(_floor_row_from_mapping, data) if row is not None]
    return BuildingPressureProfile(
        building_name=str(form.get("building_name", "")).strip(),
        floors=rows,
    )


def project_context_from_form(
    form: dict[str, Any],
    *,
    floor_profile: BuildingPressureProfile | None = None,
    material_zones: list[dict] | None = None,
    ceiling_zones: list[dict] | None = None,
) -> ProjectContext:
    """run 폼 + (이미 해석된) 압력표 → ProjectContext.

    폼에 실제로 값이 들어온 항목만 ``user_confirmed`` 로 표시한다. 빈칸은
    태그 없이 남아 :meth:`ProjectContext.unconfirmed` 에 잡힌다 — 기본값이
    쓰였다는 사실을 리포트가 숨기지 않게 하려는 것이다.
    """
    sources: dict[str, str] = {}

    def _text(key: str) -> str:
        value = str(form.get(key, "") or "").strip()
        if value:
            sources[key] = SourceTag.USER_CONFIRMED.value
        return value

    def _number(key: str) -> float | None:
        raw = str(form.get(key, "") or "").strip()
        if not raw:
            return None
        try:
            value = float(raw)
        except ValueError:
            return None
        sources[key] = SourceTag.USER_CONFIRMED.value
        return value

    project_title = _text("project_title")
    zone_name = _text("zone_name")
    building_name = _text("building_name")
    natural_fall_start_floor = _text("natural_fall_start_floor")
    machine_room_ceiling_m = _number("machine_room_ceiling_m")
    roof_tank_water_level_m = _number("roof_tank_water_level_m")

    # 라이저 형상 — 폼이 비면 상수의 대명동 실측값이 그대로 나간다. 그 사실은
    # 태그를 안 달아 [미확정] 로 드러낸다.
    riser_geometry: dict[str, Any] = {}
    for key in ("roof_run_to_riser_m", "roof_run_after_drop_m",
                "roof_to_top_floor_drop_m", "tee_to_alarm_valve_m",
                "tee_branch_above_slab_m", "top_floor_extra_height_m"):
        value = _number(key)
        if value is not None:
            riser_geometry[key] = value
    arrester = str(form.get("water_hammer_arrester", "") or "").strip()
    if arrester:
        sources["water_hammer_arrester"] = SourceTag.USER_CONFIRMED.value
        riser_geometry["water_hammer_arrester"] = arrester.lower() not in {"0", "false", "no", "없음"}

    fx_profile_key = str(form.get("fx_profile_key", "") or "").strip()
    if fx_profile_key in FX_SPEC_PROFILES:
        sources["fx_profile_key"] = SourceTag.USER_CONFIRMED.value
    else:
        fx_profile_key = FX_DEFAULT_PROFILE

    if floor_profile is not None and floor_profile.floors:
        sources["floor_profile"] = SourceTag.USER_CONFIRMED.value
    if material_zones:
        sources["material_zones"] = SourceTag.USER_CONFIRMED.value
    if ceiling_zones:
        sources["ceiling_zones"] = SourceTag.USER_CONFIRMED.value

    return ProjectContext(
        zone_spec=zone_spec_from_form(form),
        floor_profile=floor_profile,
        project_title=project_title,
        zone_name=zone_name,
        building_name=building_name or (floor_profile.building_name if floor_profile else ""),
        natural_fall_start_floor=natural_fall_start_floor or None,
        machine_room_ceiling_m=machine_room_ceiling_m,
        roof_tank_water_level_m=roof_tank_water_level_m,
        fx_profile_key=fx_profile_key,
        material_zones=list(material_zones or []),
        ceiling_zones=list(ceiling_zones or []),
        sources=sources,
        **riser_geometry,
    )


def project_context_from_job(job: dict[str, Any]) -> ProjectContext:
    """overall job 상태 → ProjectContext. 폼 딕셔너리를 읽는 유일한 자리다.

    압력표 우선순위는 업로드 CSV → 업로드 XLSX → 폼 JSON.
    ``context_override`` 는 run 이후에 사용자가 확정한 값(자연낙차 시작층 등)으로,
    run 폼보다 나중에 정해졌으니 폼 값을 덮는다.
    """
    form = {**(job.get("spec_form") or {}), **(job.get("context_override") or {})}
    building_name = str(form.get("building_name", "") or "").strip()
    profile: BuildingPressureProfile | None = None
    if job.get("pressure_table_csv"):
        profile = BuildingPressureProfile.from_csv(
            Path(job["pressure_table_csv"]), building_name=building_name)
    elif job.get("pressure_table_xlsx"):
        profile = BuildingPressureProfile.from_xlsx(
            Path(job["pressure_table_xlsx"]), building_name=building_name)
    else:
        profile = profile_from_form(form)
    return project_context_from_form(
        form,
        floor_profile=profile,
        material_zones=job.get("material_zones"),
        ceiling_zones=job.get("ceiling_zones"),
    )
