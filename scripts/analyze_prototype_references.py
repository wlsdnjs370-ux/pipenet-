"""Remote 30 프로토타입 reference 5개 파일 정밀 분석."""

from __future__ import annotations

import io
import json
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path

import ezdxf
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PROJECT = Path(r"C:\Users\admin\PycharmProjects\JupyterProject")

REF = {
    "input_dxf": PROJECT / "대명동201동 단위세대_layer정리.dxf",
    "pipe_only_dxf": PROJECT / "대명동201동 단위세대_layer정리_배관망.dxf",
    "selected_dxf": PROJECT / "대명동201동 단위세대_layer정리_배관망_파이프넷버전.dxf",
    "input_xlsx": PROJECT / "3-1형_자연낙차_LSP_4F_알람밸브_1.xlsx",
    "output_sdf": PROJECT / "3-1형_자연낙차_LSP_4F_OA_지하층포함_120m~200m미만_6.6K로 감압_알람밸브.sdf",
}


def dxf_summary(path: Path) -> dict:
    doc = ezdxf.readfile(str(path))
    msp = doc.modelspace()
    type_cnt = Counter()
    layer_cnt = Counter()
    hidden_layers = set()
    for ly in doc.layers:
        if ly.is_off() or ly.is_frozen() or int(ly.dxf.color) < 0:
            hidden_layers.add(ly.dxf.name)
    for e in msp:
        t = e.dxftype()
        l = getattr(e.dxf, "layer", "")
        type_cnt[t] += 1
        layer_cnt[l] += 1
    return {
        "size_kb": path.stat().st_size / 1024,
        "total_entities": sum(type_cnt.values()),
        "types": dict(type_cnt),
        "layers": dict(layer_cnt.most_common()),
        "hidden_layers": sorted(hidden_layers),
    }


def xlsx_summary(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    info = {"sheets": {}}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        first_data_rows = [r for r in rows if any(c is not None for c in r)]
        header = first_data_rows[0] if first_data_rows else None
        sample = first_data_rows[1:4] if len(first_data_rows) > 1 else []
        info["sheets"][sheet_name] = {
            "total_rows": len(rows),
            "non_empty_rows": len(first_data_rows),
            "max_col": ws.max_column,
            "header": list(header) if header else [],
            "sample_rows": [list(r) for r in sample],
        }
    return info


def sdf_summary(path: Path) -> dict:
    raw = path.read_text(encoding="utf-8", errors="replace")
    root = ET.fromstring(raw)
    tag_cnt = Counter()
    network_info: dict[str, dict] = {}
    for elem in root.iter():
        tag_cnt[elem.tag] += 1
    # network 안 노드 카운트
    networks = root.findall(".//Network-spray") + root.findall(".//Network")
    nets = []
    for net in networks:
        net_summary = {"tag": net.tag, "children": {}}
        for child in net:
            net_summary["children"][child.tag] = len(list(child))
        nets.append(net_summary)
    # 첫 Pipe / Node / Nozzle 한 개씩 attrib 추출
    first_pipe = root.find(".//Pipes/*")
    first_node = root.find(".//Nodes/*")
    first_nozzle = root.find(".//Nozzles/*")
    samples = {}
    for label, e in [("first_pipe", first_pipe), ("first_node", first_node), ("first_nozzle", first_nozzle)]:
        if e is not None:
            samples[label] = {
                "tag": e.tag,
                "attrib": dict(e.attrib),
                "children": [(c.tag, dict(c.attrib), c.text.strip() if c.text else "") for c in e][:8],
            }
    return {
        "size_kb": path.stat().st_size / 1024,
        "root_tag": root.tag,
        "total_elements": sum(tag_cnt.values()),
        "top_tags": dict(tag_cnt.most_common(20)),
        "networks": nets,
        "samples": samples,
    }


print("=" * 90)
for label, path in REF.items():
    print(f"\n### {label}: {path.name}")
    print(f"  exists: {path.exists()}  size: {path.stat().st_size / 1024:.1f} KB" if path.exists() else "  ! MISSING")
    if not path.exists():
        continue
    suffix = path.suffix.lower()
    if suffix == ".dxf":
        s = dxf_summary(path)
        print(f"  total_entities: {s['total_entities']}")
        print(f"  types: {s['types']}")
        print(f"  layers (top 12):")
        for l, c in list(s["layers"].items())[:12]:
            tag = " (HIDDEN)" if l in s["hidden_layers"] else ""
            print(f"    {c:6d}  |{l}|{tag}")
        if s["hidden_layers"]:
            print(f"  hidden_layers: {s['hidden_layers']}")
    elif suffix == ".xlsx":
        s = xlsx_summary(path)
        for sheet_name, info in s["sheets"].items():
            print(f"  Sheet [{sheet_name}]: total_rows={info['total_rows']} non_empty={info['non_empty_rows']} max_col={info['max_col']}")
            print(f"    header: {info['header']}")
            for i, row in enumerate(info["sample_rows"]):
                print(f"    sample[{i}]: {row}")
    elif suffix == ".sdf":
        s = sdf_summary(path)
        print(f"  root: <{s['root_tag']}>  total_elements: {s['total_elements']}")
        print(f"  top tags (count):")
        for tag, c in list(s["top_tags"].items())[:15]:
            print(f"    {tag}: {c}")
        for net in s["networks"]:
            print(f"  network <{net['tag']}>:")
            for ch, n in net["children"].items():
                print(f"    {ch}: {n} 개")
        for label_s, sample in s["samples"].items():
            print(f"  {label_s}: <{sample['tag']}> attrib={sample['attrib']}")
            for tag, attr, text in sample["children"]:
                v = (text[:30] + "..") if len(text) > 32 else text
                print(f"    <{tag}> {attr}  text={v!r}")
print()
print("=" * 90)
print("DONE")
