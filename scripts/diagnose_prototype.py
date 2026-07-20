"""참조 5종 vs 현재 prototype 출력 비교 — 99% 유사도 수렴까지 반복용."""

from __future__ import annotations

import csv
import io
import json
import math
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PROJECT = Path(r"C:\Users\admin\PycharmProjects\JupyterProject")
REF_XLSX = PROJECT / "3-1형_자연낙차_LSP_4F_알람밸브_1.xlsx"
REF_SDF = PROJECT / "assets" / "3-1형_자연낙차_LSP_4F_OA_지하층포함_120m~200m미만_6.6K로 감압_알람밸브.sdf"


def load_xlsx_counts(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        if not rows:
            continue
        out[sheet] = len(rows) - 1  # excl. header
    return out


def load_xlsx_pipes(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Pipes" not in wb.sheetnames:
        return []
    ws = wb["Pipes"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = list(rows[0])
    pipes = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        d = {h: v for h, v in zip(header, r)}
        pipes.append(d)
    return pipes


def load_xlsx_nodes(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Nodes" not in wb.sheetnames:
        return []
    ws = wb["Nodes"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = list(rows[0])
    return [{h: v for h, v in zip(header, r)} for r in rows[1:] if any(c is not None for c in r)]


def load_xlsx_equipment(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Equipment" not in wb.sheetnames:
        return []
    ws = wb["Equipment"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = list(rows[0])
    return [{h: v for h, v in zip(header, r)} for r in rows[1:] if any(c is not None for c in r)]


def load_xlsx_fittings(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Fittings" not in wb.sheetnames:
        return []
    ws = wb["Fittings"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = list(rows[0])
    return [{h: v for h, v in zip(header, r)} for r in rows[1:] if any(c is not None for c in r)]


def sdf_counts(path: Path) -> dict:
    root = ET.fromstring(path.read_text(encoding="utf-8", errors="replace"))
    return {
        # 참조 SDF 는 Network-spray 의 children 으로 Node/Pipe/Nozzle 이 직접 들어있어서 .//Node 로 카운트
        "nodes": len(root.findall(".//Node")),
        "pipes": len(root.findall(".//Pipe")),
        "nozzles": len(root.findall(".//Nozzle")),
        "equipment": len(root.findall(".//Equipment")),
        "fittings": len(root.findall(".//Fitting")),
        "file_kb": path.stat().st_size / 1024,
    }


def pipe_stats(pipes: list[dict]) -> dict:
    if not pipes:
        return {}
    lengths = [float(p.get("Length (m)") or 0) for p in pipes]
    dias = Counter(p.get("Diameter (mm)") for p in pipes)
    return {
        "count": len(pipes),
        "total_length_m": round(sum(lengths), 1),
        "avg_length_m": round(sum(lengths) / len(lengths), 2),
        "min_length_m": round(min(lengths), 2),
        "max_length_m": round(max(lengths), 2),
        "diameter_dist": dict(dias.most_common()),
    }


def node_stats(nodes: list[dict]) -> dict:
    if not nodes:
        return {}
    xs = [float(n.get("Position X (mm)") or 0) for n in nodes]
    ys = [float(n.get("Position Y (mm)") or 0) for n in nodes]
    return {
        "count": len(nodes),
        "bbox_x": (min(xs), max(xs)),
        "bbox_y": (min(ys), max(ys)),
        "x_span_mm": round(max(xs) - min(xs), 0),
        "y_span_mm": round(max(ys) - min(ys), 0),
    }


def equip_stats(equipment: list[dict]) -> dict:
    return {
        "count": len(equipment),
        "by_desc": dict(Counter(e.get("Description") for e in equipment)),
    }


def fit_stats(fittings: list[dict]) -> dict:
    return {
        "count": len(fittings),
        "by_type": dict(Counter(f.get("Fitting type") for f in fittings)),
    }


def diff_pct(cur: float, ref: float) -> float:
    if ref == 0:
        return 0.0 if cur == 0 else 999.0
    return abs(cur - ref) / ref * 100.0


def main(prototype_run_dir: Path | None = None):
    if prototype_run_dir is None:
        # 가장 최근 prototype_runs 찾기
        runs = sorted((PROJECT / "data" / "prototype_runs").glob("*"), key=lambda p: -p.stat().st_mtime if p.exists() else 0)
        if not runs:
            print("ERROR: no prototype run found")
            return
        prototype_run_dir = runs[0]
    print(f"=== 진단 대상 ===")
    print(f"  cur run  : {prototype_run_dir}")
    print(f"  ref xlsx : {REF_XLSX.name}")
    print(f"  ref sdf  : {REF_SDF.name}")

    # 현재 출력 찾기
    cur_xlsx = next(prototype_run_dir.glob("*.xlsx"), None)
    cur_sdf = next(prototype_run_dir.glob("*.sdf"), None)
    if not cur_xlsx or not cur_sdf:
        print("ERROR: prototype xlsx/sdf 없음")
        return

    print(f"\n=== XLSX 시트별 행수 ===")
    ref_counts = load_xlsx_counts(REF_XLSX)
    cur_counts = load_xlsx_counts(cur_xlsx)
    print(f"  {'sheet':12s}  {'ref':>6s}  {'cur':>6s}  {'diff':>8s}  {'%':>7s}")
    for sheet in ("Pipes", "Nodes", "Nozzles", "Fittings", "Equipment", "Meta"):
        r = ref_counts.get(sheet, 0)
        c = cur_counts.get(sheet, 0)
        d = c - r
        p = diff_pct(c, r)
        print(f"  {sheet:12s}  {r:6d}  {c:6d}  {d:+8d}  {p:>6.1f}%")

    print(f"\n=== Pipes 통계 ===")
    rp = pipe_stats(load_xlsx_pipes(REF_XLSX))
    cp = pipe_stats(load_xlsx_pipes(cur_xlsx))
    for k in ("count", "total_length_m", "avg_length_m", "min_length_m", "max_length_m"):
        print(f"  {k:18s}  ref={rp.get(k,'')}  cur={cp.get(k,'')}  diff%={diff_pct(cp.get(k,0), rp.get(k,0)):.1f}")
    print(f"  diameter dist ref: {rp.get('diameter_dist',{})}")
    print(f"  diameter dist cur: {cp.get('diameter_dist',{})}")

    print(f"\n=== Nodes 통계 ===")
    rn = node_stats(load_xlsx_nodes(REF_XLSX))
    cn = node_stats(load_xlsx_nodes(cur_xlsx))
    for k in ("count", "x_span_mm", "y_span_mm"):
        print(f"  {k:14s}  ref={rn.get(k,'')}  cur={cn.get(k,'')}  diff%={diff_pct(cn.get(k,0), rn.get(k,0)):.1f}")

    print(f"\n=== Equipment ===")
    re_ = equip_stats(load_xlsx_equipment(REF_XLSX))
    ce_ = equip_stats(load_xlsx_equipment(cur_xlsx))
    print(f"  ref: count={re_['count']} by_desc={re_['by_desc']}")
    print(f"  cur: count={ce_['count']} by_desc={ce_['by_desc']}")

    print(f"\n=== Fittings ===")
    rf = fit_stats(load_xlsx_fittings(REF_XLSX))
    cf = fit_stats(load_xlsx_fittings(cur_xlsx))
    print(f"  ref: count={rf['count']} by_type={rf['by_type']}")
    print(f"  cur: count={cf['count']} by_type={cf['by_type']}")

    print(f"\n=== SDF 요소수 ===")
    rs = sdf_counts(REF_SDF)
    cs = sdf_counts(cur_sdf)
    for k in ("nodes", "pipes", "nozzles", "equipment", "fittings"):
        print(f"  {k:12s}  ref={rs[k]:6d}  cur={cs[k]:6d}  diff%={diff_pct(cs[k], rs[k]):.1f}")
    print(f"  file_kb     ref={rs['file_kb']:.1f}  cur={cs['file_kb']:.1f}")

    # 종합 점수
    print(f"\n=== 종합 유사도 점수 (낮을수록 좋음) ===")
    metrics = [
        ("Pipes count", cp.get("count", 0), rp.get("count", 0)),
        ("Nodes count", cn.get("count", 0), rn.get("count", 0)),
        ("Nozzles", cur_counts.get("Nozzles", 0), ref_counts.get("Nozzles", 0)),
        ("Fittings", cur_counts.get("Fittings", 0), ref_counts.get("Fittings", 0)),
        ("Equipment", cur_counts.get("Equipment", 0), ref_counts.get("Equipment", 0)),
        ("Total len (m)", cp.get("total_length_m", 0), rp.get("total_length_m", 0)),
        ("X span (mm)", cn.get("x_span_mm", 0), rn.get("x_span_mm", 0)),
        ("Y span (mm)", cn.get("y_span_mm", 0), rn.get("y_span_mm", 0)),
        ("SDF nodes", cs["nodes"], rs["nodes"]),
        ("SDF pipes", cs["pipes"], rs["pipes"]),
        ("SDF nozzles", cs["nozzles"], rs["nozzles"]),
    ]
    diffs = []
    print(f"  {'metric':16s}  {'cur':>10s}  {'ref':>10s}  {'diff%':>7s}")
    for name, cur, ref in metrics:
        d = diff_pct(cur, ref)
        diffs.append(d)
        marker = " ✓" if d <= 5 else " ⚠" if d <= 20 else " ✗"
        print(f"  {name:16s}  {cur:>10.1f}  {ref:>10.1f}  {d:>6.1f}%{marker}")
    score = sum(diffs) / len(diffs) if diffs else 0
    print(f"\n  평균 오차: {score:.1f}%  →  99% 유사도 = 평균 1% 이하 (목표)")
    print(f"  ±5% 이내 metric: {sum(1 for d in diffs if d <= 5)}/{len(diffs)}")
    return diffs


if __name__ == "__main__":
    main()
